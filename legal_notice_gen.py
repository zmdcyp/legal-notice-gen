"""
Legal Notice Batch Generator (Web)

Placeholder-neutral batch processor:
- The tool only replaces {{placeholder}} tokens in a docx template with
  values from Excel columns whose headers match the placeholder names.
- It does NOT interpret, validate, or understand the business content of
  the template or spreadsheet. Any docx + any Excel that share matching
  field names will work.

Scale-oriented architecture:
- Async task model: /generate returns a task_id immediately; the client
  polls /status and fetches the final zip from /download. Avoids HTTP
  request-level timeouts (Nginx/Gunicorn) on long batches.
- Parallel docx rendering via ThreadPoolExecutor.
- Parallel PDF conversion via multiple soffice workers, each with its own
  UserInstallation profile dir; files are batched per worker so soffice
  starts only N times rather than once per file.

Single-file delivery:
- Everything (backend, HTML, CSS, JS) is in this one file. Ship by copying
  legal_notice_gen.py + requirements.txt to the target machine.
"""

import os
import re
import io
import json
import math
import base64
import random
import zipfile
import uuid
import shutil
import tempfile
import subprocess
import datetime
import threading
import traceback
import concurrent.futures
import time
import html as _htmllib
import hashlib
import functools
import sqlite3
from collections import defaultdict

from flask import (Flask, request, send_file, jsonify, session,
                   Response, after_this_request, redirect, send_from_directory)
import hmac as _hmac
import openpyxl

# ── security overlay deps (watermark + guilloche + QR) ─────────
# These are optional: if any are unavailable the overlay features are
# silently skipped so the tool still renders plain DOCX/PDFs.
try:
    from PIL import Image, ImageDraw, ImageFont
    import qrcode
    import arabic_reshaper
    from bidi.algorithm import get_display
    _SECURITY_DEPS_OK = True
except ImportError:
    _SECURITY_DEPS_OK = False

# ── app setup ───────────────────────────────────────────────────

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
UPLOAD_DIR = os.path.join(BASE_DIR, "uploads")

# Point Flask's built-in /static route at templates/static so the base
# HTML template's absolute paths (/static/fonts.css, /static/images/...)
# resolve inside the app without a separate http.server.
app = Flask(__name__,
            static_folder=os.path.join(BASE_DIR, "templates", "static"),
            static_url_path="/static")
os.makedirs(UPLOAD_DIR, exist_ok=True)

# Persist secret_key across restarts so /upload -> /generate sessions survive.
_SECRET_FILE = os.path.join(UPLOAD_DIR, ".secret_key")
if os.path.exists(_SECRET_FILE):
    with open(_SECRET_FILE, "rb") as _f:
        app.secret_key = _f.read()
else:
    app.secret_key = os.urandom(32)
    with open(_SECRET_FILE, "wb") as _f:
        _f.write(app.secret_key)

PLACEHOLDER_RE = re.compile(r"\{\{(.+?)\}\}")
SAFE_NAME_RE = re.compile(r'[\\/*?:"<>|\r\n\t]+')

# Valid template-slug regex. Gates every route that takes <slug> so an
# attacker cannot escape TEMPLATES_ROOT with `..`/slashes/etc.
_SLUG_VALID_RE = re.compile(r"[A-Za-z0-9_-]{1,50}")


def _valid_slug(slug):
    return bool(slug) and bool(_SLUG_VALID_RE.fullmatch(slug))


def _name_seed(name):
    """Stable 64-bit seed for per-document RNG.

    Python's built-in `hash()` is salted per process (PYTHONHASHSEED
    defaults to `random`), so `hash(name)` gives DIFFERENT values on
    different server restarts AND across the separate Python processes
    spawned by ProcessPoolExecutor. That breaks the anti-counterfeit
    contract ("same respondent → same serial / same watermark nonce")
    the moment the server reboots or a batch spans multiple workers.
    SHA-256 of the UTF-8 bytes is deterministic everywhere."""
    digest = hashlib.sha256((name or "").encode("utf-8")).digest()[:8]
    return int.from_bytes(digest, "big")

# ── access auth ─────────────────────────────────────────────────
#
# Single-password gate. Override at runtime with the LEGAL_NOTICE_PASSWORD
# environment variable (strongly recommended in production). The default
# baked into the file is a randomly generated 20-character password; change
# it by setting the env var on the VPS/systemd unit.
_DEFAULT_PASSWORD = "RT%L6IXoXT*^r=z6%npe"
APP_PASSWORD = os.environ.get("LEGAL_NOTICE_PASSWORD") or _DEFAULT_PASSWORD

# Paths that never require auth (the login form itself + logout).
_AUTH_EXEMPT_PATHS = {"/login", "/logout"}

# API endpoints return JSON 401 instead of redirecting, so the fetch() calls
# on the front-end don't silently follow a redirect to HTML.
_AUTH_API_PREFIXES = ("/upload", "/generate", "/status", "/download", "/api/")

# ── machine profiles ────────────────────────────────────────────
#
# Concurrency used to be a single module constant. But this project runs in
# two very different environments:
#   1. Locally on a powerful Mac for big interactive batches.
#   2. On a small VPS for always-on access.
# The webpage lets the user pick which profile to use at generate time. The
# default is conservative (VPS); picking "mac" unlocks process-pool-based
# rendering that scales linearly with CPU cores — the real speedup lever for
# DOCX output, which is this project's primary mode.
_CPU = os.cpu_count() or 4

MACHINE_PROFILES = {
    "vps": {
        "render_workers":  1,    # 1 Chromium ≈ 250 MB RAM — safe on small VPS
        "label": "VPS (1 Chromium worker)",
    },
    "mac": {
        "render_workers":  min(_CPU, 6),
        "label": f"Mac (up to {min(_CPU, 6)} Chromium workers)",
    },
}
DEFAULT_MACHINE = "vps"


def _get_machine_profile(name):
    return MACHINE_PROFILES.get((name or "").lower()) or MACHINE_PROFILES[DEFAULT_MACHINE]


# soffice timeout: generous per-file allowance with a floor so small batches
# don't time out prematurely.


# ── template storage ─────────────────────────────────────────────
#
# Single built-in base template with 5 data-purpose blocks. The 4
# user-editable blocks (subject / body-text / consequences / payment) are
# stored as PLAIN TEXT. Rendering to HTML happens per-block in
# `BLOCK_RENDERERS` — each renderer wraps paragraphs in the right tags
# (<p>, <li>, etc.) and re-applies inline formatting (italic/bold/etc.)
# to known phrases via `FORMAT_RULES`. This guarantees consistent
# formatting regardless of what the user types, and keeps the editor
# UI a simple plain-text textarea per block.
#
# Directory layout for saved templates:
#   uploads/templates/<slug>/
#     meta.json     {name, created, updated}
#     blocks.json   {notice-subject: "...", notice-body-text: "...", ...}
#     assets/
#       logo.png        (optional per-template override)
#       seal.png        (optional per-template override)
#       signature.png   (optional per-template override)
# The built-in "default" template lives in code (DEFAULT_BLOCKS_TEXT)
# and uses the PNGs shipped at templates/static/images/.

TEMPLATES_ROOT = os.path.join(UPLOAD_DIR, "templates")
os.makedirs(TEMPLATES_ROOT, exist_ok=True)

TEMPLATE_BASE_DIR = os.path.join(BASE_DIR, "templates")
TEMPLATE_BASE_HTML_PATH = os.path.join(TEMPLATE_BASE_DIR, "legal_notice_full.html")
TEMPLATE_STATIC_DIR = os.path.join(TEMPLATE_BASE_DIR, "static")

EDITABLE_PURPOSES = (
    "letterhead-firm",
    "letterhead-partners",
    "notice-subject",
    "notice-body-text",
    "legal-consequences",
    "payment-instructions",
    "page-footer",
)

# The amounts table is fixed structure; user references it from the body
# text via this marker so they never edit HTML.
BODY_TABLE_MARKER = "[[AMOUNTS_TABLE]]"
# The callout-pay paragraph is the prominent bold "called upon to pay"
# line. The next non-marker paragraph after [[CALLOUT]] gets the
# callout-pay class.
BODY_CALLOUT_MARKER = "[[CALLOUT]]"

AMOUNTS_TABLE_HTML = (
    '<table class="amounts">\n'
    '  <tr><td class="item">Principal</td><td class="cur">PKR</td>'
    '<td class="amt">{{Principal_Amount}}</td></tr>\n'
    '  <tr><td class="item">Interest</td><td class="cur">PKR</td>'
    '<td class="amt">{{Interest}}</td></tr>\n'
    '  <tr><td class="item">Penalty</td><td class="cur">PKR</td>'
    '<td class="amt">{{Penalty}}</td></tr>\n'
    '  <tr class="total"><td class="item">Total payable</td>'
    '<td class="cur">PKR</td><td class="amt">{{Payable}}</td></tr>\n'
    '</table>'
)


DEFAULT_BLOCKS_TEXT = {
    "letterhead-firm": (
        "S&S LAW FIRM\n"
        "Legal & Corporate Consultant"
    ),
    "letterhead-partners": (
        "Muhammad Junaid Abbasi | Adv High Court (Managing Partner)\n"
        "Khizar Ayoub Chaudhary | Adv High Court\n"
        "Shahid Abbas Khokhar | Adv High Court\n"
        "Raja Usman Hameed Abbasi | Adv High Court\n"
        "Shabab Zahir | Adv High Court\n"
        "Habibullah | Adv High Court"
    ),
    "page-footer": (
        "Office: Plot# 81, I&T Center, G-9/4 Islamabad. "
        "| Email: sandslawfirm3187@gmail.com\n"
        "Phone: +92 303 9175939 | +92 333 5649275"
    ),
    "notice-subject": (
        "SUBJECT: FINAL LEGAL NOTICE FOR RECOVERY OF OUTSTANDING FINANCE "
        "UNDER SECTIONS 3, 9, 15 & 20 OF THE FINANCIAL INSTITUTIONS "
        "(RECOVERY OF FINANCES) ORDINANCE, 2001"
    ),
    "notice-body-text": (
        "Under instructions from our client, M/s Zanda Financial Services (Pvt.) Limited, "
        "this FINAL LEGAL NOTICE is served upon you as follows:\n\n"
        "That you availed a loan/finance facility through the MoneyTap Application "
        "disbursed on {{disb_date}} under agreed terms and conditions.\n\n"
        "That as per agreed terms, the due date for repayment was {{Due_date}}, "
        "which you have failed to honour.\n\n"
        "That an amount comprising the following:\n\n"
        "[[AMOUNTS_TABLE]]\n\n"
        "is now outstanding, bringing the total payable to PKR {{Payable}}.\n\n"
        "Transaction history:\n"
        "Transaction ID, {{Transaction_id}}, and EasyPaisa account number: {{easypaisa_account}}.\n\n"
        "That despite repeated reminders, you have willfully evaded payment, "
        "reflecting mala fide intention and unlawful retention of funds.\n\n"
        "[[CALLOUT]]\n\n"
        "You are hereby finally called upon to pay PKR {{Payable}} within "
        "03 (three) days from receipt of this notice."
    ),
    "legal-consequences": (
        "Initiation of recovery proceedings before the competent Banking Court "
        "under Sections 9, 15 and 20 of the Financial Institutions "
        "(Recovery of Finances) Ordinance, 2001.\n"
        "Attachment, garnishee, and freezing of your bank accounts, assets, "
        "and properties under the relevant provisions of law.\n"
        "Application before the competent forum for directions to NADRA for "
        "blocking/suspension of CNIC, subject to court orders.\n"
        "In case any misrepresentation, fraudulent activity, or concealment "
        "of material facts is found, our client shall initiate legal proceedings, "
        "including filing of a complaint and registration of FIR under Clause (b) "
        "of Section 20 of the Financial Institutions (Recovery of Finances) "
        "Ordinance, 2001, at your sole risk and cost.\n"
        "Recovery of full litigation costs, legal expenses, markup, damages, "
        "and any other relief available under law."
    ),
    "payment-instructions": (
        "Payment shall be made through the same mobile application (MoneyTap App) "
        "through which the loan facility was originally availed.\n\n"
        "This is your FINAL AND LAST OPPORTUNITY to settle the matter amicably. "
        "Failure to comply will result in strict legal action at your sole risk and cost.\n\n"
        "For any query, you may contact: 0303-9175939 | 0309-0548645. "
        "For and on behalf of the Client."
    ),
}


# Per-block phrase-to-tag mapping. After block renderer emits HTML, scan
# for these phrases and wrap them. Longest phrases first so a shorter
# phrase that's a substring doesn't "steal" the match.
FORMAT_RULES = {
    "notice-body-text": [
        ('em class="client"', "M/s Zanda Financial Services (Pvt.) Limited"),
        ("em",                "MoneyTap Application"),
        ("em",                "{{Transaction_id}}"),
        ("em",                "{{easypaisa_account}}"),
        ('span class="total-amt"', "{{Payable}}"),
        ('span class="deadline"',  "03 (three) days"),
    ],
    "payment-instructions": [
        ("em", "MoneyTap App"),
    ],
    "page-footer": [
        ('span class="lbl"', "Office:"),
        ('span class="lbl"', "Email:"),
        ('span class="lbl"', "Phone:"),
    ],
}


def _esc_text(text):
    """HTML-escape but keep placeholders ({{...}}) readable as-is (they
    will be substituted at render time with already-safe values)."""
    return _htmllib.escape(text, quote=False)


def _apply_format_rules(html_fragment, purpose):
    """Wrap each configured phrase with the configured tag. Phrases are
    matched verbatim (case-sensitive) on the already-escaped HTML; the
    phrase list is ordered longest-first to avoid nested-match mishaps."""
    for tag_spec, phrase in FORMAT_RULES.get(purpose, ()):
        tag_name = tag_spec.split(" ", 1)[0]
        html_fragment = html_fragment.replace(
            phrase, f"<{tag_spec}>{phrase}</{tag_name}>")
    return html_fragment


def _render_subject(text):
    body = _apply_format_rules(_esc_text(text.strip()), "notice-subject")
    return f'<div class="subject">{body}</div>'


def _render_body_text(text):
    """[[AMOUNTS_TABLE]] becomes the fixed amounts table. [[CALLOUT]] on
    its own line flags the NEXT paragraph as `callout-pay`. All other
    paragraphs are wrapped in <p>. Single newlines inside a paragraph
    become <br> so multi-line items like "Transaction history:\\nid, X"
    render as one paragraph with a line break."""
    paragraphs = [p.strip() for p in text.split("\n\n") if p.strip()]
    out = []
    callout_next = False
    for para in paragraphs:
        if para == BODY_TABLE_MARKER:
            out.append(AMOUNTS_TABLE_HTML)
            continue
        if para == BODY_CALLOUT_MARKER:
            callout_next = True
            continue
        escaped = _esc_text(para).replace("\n", "<br>\n")
        escaped = _apply_format_rules(escaped, "notice-body-text")
        cls = ' class="callout-pay"' if callout_next else ""
        out.append(f"<p{cls}>{escaped}</p>")
        callout_next = False
    return "\n".join(out)


def _render_consequences(text):
    items = [ln.strip() for ln in text.split("\n") if ln.strip()]
    lis = "\n".join(f"<li>{_esc_text(x)}</li>" for x in items)
    return (
        '<h3 class="consequences-heading">LEGAL CONSEQUENCES IN CASE OF NON-COMPLIANCE</h3>\n'
        f'<ol class="consequences">\n{lis}\n</ol>'
    )


def _render_payment(text):
    paras = [p.strip() for p in text.split("\n\n") if p.strip()]
    ps = []
    for p in paras:
        esc = _esc_text(p).replace("\n", "<br>\n")
        esc = _apply_format_rules(esc, "payment-instructions")
        ps.append(f"<p>{esc}</p>")
    return (
        '<h3 class="payment-heading">PAYMENT INSTRUCTIONS</h3>\n'
        + "\n".join(ps)
    )


def _render_letterhead_firm(text):
    """Line 1 → firm h1 heading; line 2+ → italic tag below. Extra lines
    after the 2nd get joined with <br> into the same tag line."""
    lines = [ln.strip() for ln in text.split("\n") if ln.strip()]
    if not lines:
        return ""
    firm = _esc_text(lines[0])
    tag_html = ""
    if len(lines) > 1:
        tag = "<br>".join(_esc_text(ln) for ln in lines[1:])
        tag_html = f'<div class="tag">{tag}</div>'
    return f'<h1 class="firm">{firm}</h1>\n{tag_html}'


def _render_letterhead_partners(text):
    """Each non-empty line is one partner. Format: 'Name | Role'.
    If no '|' is present the whole line is treated as the name."""
    out = []
    for ln in text.split("\n"):
        ln = ln.strip()
        if not ln:
            continue
        if "|" in ln:
            name, role = ln.split("|", 1)
            name_html = _esc_text(name.strip())
            role_html = _esc_text(role.strip())
        else:
            name_html = _esc_text(ln)
            role_html = ""
        if role_html:
            out.append(
                f'<div class="person"><span class="name">{name_html}</span>'
                f'<span class="role">{role_html}</span></div>')
        else:
            out.append(
                f'<div class="person"><span class="name">{name_html}</span></div>')
    return "\n".join(out)


def _render_page_footer(text):
    """Each non-empty line → <p>. Known labels (Office:/Email:/Phone:)
    are wrapped with <span class='lbl'> via FORMAT_RULES."""
    out = []
    for ln in text.split("\n"):
        ln = ln.strip()
        if not ln:
            continue
        esc = _esc_text(ln)
        esc = _apply_format_rules(esc, "page-footer")
        out.append(f"<p>{esc}</p>")
    return "\n".join(out)


BLOCK_RENDERERS = {
    "letterhead-firm":      _render_letterhead_firm,
    "letterhead-partners":  _render_letterhead_partners,
    "notice-subject":       _render_subject,
    "notice-body-text":     _render_body_text,
    "legal-consequences":   _render_consequences,
    "payment-instructions": _render_payment,
    "page-footer":          _render_page_footer,
}


def render_blocks_to_html(blocks_text):
    """Return {purpose: html_fragment} from {purpose: plain_text}."""
    out = {}
    for purpose in EDITABLE_PURPOSES:
        text = blocks_text.get(purpose, DEFAULT_BLOCKS_TEXT.get(purpose, ""))
        out[purpose] = BLOCK_RENDERERS[purpose](text)
    return out


# ── saved-template CRUD ──

_SLUG_RE = re.compile(r"[^A-Za-z0-9_-]+")


def _slugify(name):
    s = _SLUG_RE.sub("-", (name or "").strip()).strip("-").lower()
    return s[:50] or "template"


def list_templates():
    """Return all templates — built-in default first, then saved ones."""
    out = [{
        "slug": "default",
        "name": "S&S Law Firm — Default",
        "builtin": True,
        "created": None,
        "updated": None,
    }]
    if os.path.isdir(TEMPLATES_ROOT):
        for slug in sorted(os.listdir(TEMPLATES_ROOT)):
            meta_path = os.path.join(TEMPLATES_ROOT, slug, "meta.json")
            if os.path.isfile(meta_path):
                try:
                    with open(meta_path, "r", encoding="utf-8") as f:
                        meta = json.load(f)
                    out.append({
                        "slug": slug,
                        "name": meta.get("name", slug),
                        "builtin": False,
                        "created": meta.get("created"),
                        "updated": meta.get("updated"),
                    })
                except (OSError, json.JSONDecodeError):
                    pass
    return out


ASSET_KINDS = ("logo", "seal", "signature_seal")

# Default per-asset placement: size (mm), horizontal/vertical offset from
# the default anchor (mm), and user-set base rotation (degrees). These
# defaults reproduce the baseline CSS numbers in templates/legal_notice_full.html.
# The per-document random rotation from the security overlay is applied
# ON TOP of the user-set base rotation (seal / signature_seal).
DEFAULT_ASSETS_CONFIG = {
    "logo":           {"size": 22, "dx": 0, "dy": 0, "rot": 0},
    "seal":           {"size": 22, "dx": 0, "dy": 0, "rot": 0},
    # signature_seal PNG is a square 1000×1000 containing the handwritten
    # signature + printed partner name + role baked in, so the width is
    # also the full block height. 44mm leaves enough room above the page
    # footer line and keeps the signature comparable in scale to the
    # 22mm office seal on the left.
    "signature_seal": {"size": 44, "dx": 0, "dy": 0, "rot": 0},
}


def _merge_assets_config(user_cfg):
    """Fill any missing per-kind values from DEFAULT_ASSETS_CONFIG and
    coerce numeric types. Tolerates partial user overrides."""
    out = {}
    for kind in ASSET_KINDS:
        base = DEFAULT_ASSETS_CONFIG[kind]
        u = (user_cfg or {}).get(kind) or {}
        try:
            out[kind] = {
                "size": float(u.get("size", base["size"])),
                "dx":   float(u.get("dx",   base["dx"])),
                "dy":   float(u.get("dy",   base["dy"])),
                "rot":  float(u.get("rot",  base["rot"])),
            }
        except (TypeError, ValueError):
            out[kind] = dict(base)
    return out


def _read_template_dir(slug):
    """Read on-disk overrides (blocks, assets, security, assets_config,
    meta) for a slug from uploads/templates/<slug>/. Missing files return
    empty defaults; this helper is shared by load_template for both the
    built-in default and user-saved templates."""
    dir_path = os.path.join(TEMPLATES_ROOT, slug)
    out = {"blocks": {}, "assets": {}, "security": None,
           "assets_config": None,
           "name": None, "created": None, "updated": None}
    if not os.path.isdir(dir_path):
        return out
    for fname, key in [("blocks.json", "blocks"),
                       ("security.json", "security"),
                       ("assets_config.json", "assets_config"),
                       ("meta.json", "meta")]:
        p = os.path.join(dir_path, fname)
        if os.path.isfile(p):
            try:
                with open(p, "r", encoding="utf-8") as f:
                    data = json.load(f)
                if key == "meta":
                    out["name"] = data.get("name")
                    out["created"] = data.get("created")
                    out["updated"] = data.get("updated")
                else:
                    out[key] = data
            except (OSError, json.JSONDecodeError):
                pass
    assets_dir = os.path.join(dir_path, "assets")
    for kind in ASSET_KINDS:
        p = os.path.join(assets_dir, f"{kind}.png")
        if os.path.isfile(p):
            out["assets"][kind] = p
    return out


def load_template(slug):
    """Return {slug, name, blocks, assets, assets_config, security, builtin}
    or None.

    For 'default' we start from the built-in hardcoded blocks/security
    and overlay any on-disk overrides under uploads/templates/default/.
    For any other slug we load straight from disk.
    """
    if slug == "default":
        disk = _read_template_dir("default")
        blocks = dict(DEFAULT_BLOCKS_TEXT)
        blocks.update({p: v for p, v in (disk["blocks"] or {}).items()
                       if p in EDITABLE_PURPOSES})
        return {
            "slug": "default",
            "name": disk["name"] or "S&S Law Firm — Default",
            "blocks": blocks,
            "assets": disk["assets"],
            "assets_config": _merge_assets_config(disk["assets_config"]),
            "security": _merge_security_config(disk["security"])
                        if disk["security"] else load_default_security(),
            "builtin": True,
        }
    disk = _read_template_dir(slug)
    if not disk["blocks"]:
        return None
    return {
        "slug": slug,
        "name": disk["name"] or slug,
        "blocks": {p: (disk["blocks"].get(p) or DEFAULT_BLOCKS_TEXT.get(p, ""))
                   for p in EDITABLE_PURPOSES},
        "assets": disk["assets"],
        "assets_config": _merge_assets_config(disk["assets_config"]),
        "security": _merge_security_config(disk["security"])
                    if disk["security"] else load_default_security(),
        "builtin": False,
    }


def save_template(name, blocks=None, security=None, assets_config=None,
                  overwrite_slug=None):
    """Persist blocks / security / assets_config / name under a slug.

    - If `overwrite_slug` is given and the dir exists (or overwrite_slug
      == 'default'), update in place.
    - Otherwise allocate a fresh slug derived from `name`; append -2,-3…
      if one already exists with that base.
    - `blocks` / `security` / `assets_config` may be None to keep
      existing values.
    """
    if overwrite_slug == "default":
        slug = "default"
    else:
        base = _slugify(name or "template")
        if base == "default":
            base = "template"
        slug = overwrite_slug or base
        dir_path = os.path.join(TEMPLATES_ROOT, slug)
        if overwrite_slug is None:
            n = 1
            while os.path.isdir(dir_path):
                n += 1
                slug = f"{base}-{n}"
                dir_path = os.path.join(TEMPLATES_ROOT, slug)

    dir_path = os.path.join(TEMPLATES_ROOT, slug)
    os.makedirs(dir_path, exist_ok=True)
    os.makedirs(os.path.join(dir_path, "assets"), exist_ok=True)

    now = datetime.datetime.now().isoformat(timespec="seconds")
    meta_path = os.path.join(dir_path, "meta.json")
    created = now
    existing_name = None
    if os.path.isfile(meta_path):
        try:
            with open(meta_path, "r", encoding="utf-8") as f:
                m = json.load(f)
            created = m.get("created", now)
            existing_name = m.get("name")
        except (OSError, json.JSONDecodeError):
            pass
    meta = {
        "name": name or existing_name
                 or ("S&S Law Firm — Default" if slug == "default" else slug),
        "created": created,
        "updated": now,
    }
    with open(meta_path, "w", encoding="utf-8") as f:
        json.dump(meta, f, indent=2, ensure_ascii=False)

    if blocks is not None:
        # Merge with existing on-disk blocks so we never lose sections.
        existing_blocks = _read_template_dir(slug)["blocks"]
        merged_blocks = dict(existing_blocks)
        for p in EDITABLE_PURPOSES:
            if p in blocks:
                merged_blocks[p] = blocks[p]
        with open(os.path.join(dir_path, "blocks.json"), "w",
                  encoding="utf-8") as f:
            json.dump(merged_blocks, f, indent=2, ensure_ascii=False)

    if security is not None:
        merged = _merge_security_config(security)
        with open(os.path.join(dir_path, "security.json"), "w",
                  encoding="utf-8") as f:
            json.dump(merged, f, indent=2, ensure_ascii=False)

    if assets_config is not None:
        merged_ac = _merge_assets_config(assets_config)
        with open(os.path.join(dir_path, "assets_config.json"), "w",
                  encoding="utf-8") as f:
            json.dump(merged_ac, f, indent=2, ensure_ascii=False)
    return slug


def delete_template(slug):
    if slug == "default":
        raise ValueError("cannot delete the built-in default template")
    dir_path = os.path.join(TEMPLATES_ROOT, slug)
    if not os.path.isdir(dir_path):
        return False
    shutil.rmtree(dir_path)
    return True


def save_template_asset(slug, kind, uploaded_file):
    """Store an uploaded PNG as <slug>/assets/<kind>.png. Works for
    any slug including 'default' — overrides the shipped defaults."""
    if kind not in ASSET_KINDS:
        raise ValueError(
            f"unknown asset kind: {kind!r} "
            f"(valid: {', '.join(ASSET_KINDS)})")
    dir_path = os.path.join(TEMPLATES_ROOT, slug, "assets")
    os.makedirs(dir_path, exist_ok=True)
    out_path = os.path.join(dir_path, f"{kind}.png")
    uploaded_file.save(out_path)
    return out_path


def delete_template_asset(slug, kind):
    """Remove a per-template asset override (revert to shipped default)."""
    if kind not in ASSET_KINDS:
        raise ValueError(f"unknown asset kind: {kind!r}")
    p = os.path.join(TEMPLATES_ROOT, slug, "assets", f"{kind}.png")
    if os.path.isfile(p):
        os.remove(p)
        return True
    return False


# ── security config store (for default template) ─────────────────
#
# The built-in default template doesn't own a meta.json; its security
# overrides live here so users can iterate in the /editor without
# promoting to a named saved template.

_DEFAULT_SECURITY_FILE = os.path.join(UPLOAD_DIR, "default_security.json")


def load_default_security():
    """Return the stored security config for the default template, or
    the baseline DEFAULT_SECURITY_CONFIG if nothing has been saved yet."""
    if os.path.isfile(_DEFAULT_SECURITY_FILE):
        try:
            with open(_DEFAULT_SECURITY_FILE, "r", encoding="utf-8") as f:
                return _merge_security_config(json.load(f))
        except (OSError, json.JSONDecodeError):
            pass
    return _merge_security_config(None)


def save_default_security(config):
    merged = _merge_security_config(config)
    with open(_DEFAULT_SECURITY_FILE, "w", encoding="utf-8") as f:
        json.dump(merged, f, indent=2, ensure_ascii=False)
    return merged


# ── task registry ────────────────────────────────────────────────

# In-process task store. REQUIRES Gunicorn --workers 1 --threads N so that
# /status and /download reach the same process that started the task.
TASKS = {}
TASKS_LOCK = threading.Lock()
TASK_TTL_SECONDS = 3600  # prune finished/stale tasks after 1 hour


def _new_task():
    _prune_tasks()
    tid = uuid.uuid4().hex
    with TASKS_LOCK:
        TASKS[tid] = {
            "status": "pending",     # pending | running | done | error
            "stage": "queued",       # queued | rendering | converting | packing | done
            "progress": 0,
            "total": 0,
            "message": "",
            "result_path": None,
            "error": None,
            "created": time.time(),
        }
    return tid


def _update_task(tid, **fields):
    with TASKS_LOCK:
        if tid in TASKS:
            TASKS[tid].update(fields)


def _get_task(tid):
    with TASKS_LOCK:
        task = TASKS.get(tid)
        return dict(task) if task else None


def _drop_task(tid):
    with TASKS_LOCK:
        return TASKS.pop(tid, None)


def _prune_tasks():
    """Remove old finished tasks and their leftover part files."""
    cutoff = time.time() - TASK_TTL_SECONDS
    stale_paths = []
    with TASKS_LOCK:
        for tid, t in list(TASKS.items()):
            if t["created"] < cutoff and t["status"] in ("done", "error"):
                for p in (t.get("ready_parts") or []):
                    if p.get("path"):
                        stale_paths.append(p["path"])
                TASKS.pop(tid, None)
    for path in stale_paths:
        if path and os.path.exists(path):
            try:
                os.remove(path)
            except OSError:
                pass


def _format_value(v):
    """Stringify a cell value for substitution.

    - Integer-valued floats become plain integers so IDs / CNICs / phone
      numbers don't render as '1234.0'.
    - Non-integer floats are treated as monetary amounts and rendered with
      thousands separators and 2 decimal places ('1234.5' -> '1,234.50').
    - Money columns are usually formatted at read_excel() time based on the
      column header (see _is_money_header) — by the time a value reaches
      this function it's already a pre-formatted string in that case.
    """
    if v is None:
        return ""
    if isinstance(v, float):
        if v.is_integer():
            return str(int(v))
        return f"{v:,.2f}"
    return str(v)


# ── money column detection ───────────────────────────────────────
#
# The user's real workload has columns like Principal_Amount, Interest,
# Penalty, Payable that are stored in Excel as integer-valued numbers
# (e.g. 500000, not 500000.00). The integer-valued-float fast path in
# _format_value would drop them to plain integers with no commas. To get
# the formal "1,234.56" style the user wants, we match money columns by
# header name and always format them as {value:,.2f}, regardless of
# whether the cell was stored as int or float.
#
# The tool is still placeholder-neutral: this list is a generic set of
# money words in English and Chinese. If your template uses a money
# column not covered here, add the keyword below.

MONEY_KEYWORDS = (
    # English
    "amount", "interest", "penalty", "payable", "fee", "balance",
    "total", "principal", "charge", "payment", "debt", "price", "cost",
    # Chinese
    "金额", "利息", "罚息", "罚金", "应付", "应还", "应缴", "费用",
    "总额", "本金", "欠款", "滞纳金", "款项",
)


def _is_money_header(header):
    """Return True if the column header looks like a monetary field."""
    h = (header or "").strip().lower()
    if not h:
        return False
    # Never treat explicit date columns as money, even if another money
    # keyword happens to appear elsewhere in the string.
    if h == "date" or h.endswith("_date") or h.endswith(" date"):
        return False
    return any(kw in h for kw in MONEY_KEYWORDS)


# ── placeholder extraction ───────────────────────────────────────



# ── Excel reading ────────────────────────────────────────────────

def read_excel(excel_path: str):
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return [], []
    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    data_rows = []
    for row in rows[1:]:
        if all(cell is None for cell in row):
            continue
        record = {}
        for header, cell in zip(headers, row):
            if not header:
                continue
            if isinstance(cell, (datetime.datetime, datetime.date)):
                record[header] = cell.strftime("%d/%m/%Y")
            elif _is_money_header(header) and isinstance(cell, (int, float)) and not isinstance(cell, bool):
                # Money columns are formatted up-front to "1,234.56" style
                # regardless of whether Excel stored the value as int or
                # float — integer-valued amounts (500000) and float
                # amounts (500000.5) both land in the same formal shape.
                record[header] = f"{float(cell):,.2f}"
            elif isinstance(cell, float) and cell.is_integer():
                # Integer-valued numeric cells (CNIC, phone numbers, etc.)
                # should not carry a trailing .0 into the notice.
                record[header] = int(cell)
            else:
                record[header] = cell if cell is not None else ""
        data_rows.append(record)
    return headers, data_rows


# ── filename helpers ─────────────────────────────────────────────

def _safe_name(value) -> str:
    s = _format_value(value)
    s = SAFE_NAME_RE.sub("_", s).strip().strip(".")
    return s or "unnamed"


def _build_filename(record, filename_fields, idx):
    """Join the selected field values with '_' to form the base filename."""
    if filename_fields:
        parts = []
        for f in filename_fields:
            v = record.get(f, "")
            if v != "" and v is not None:
                parts.append(_format_value(v))
        if parts:
            return _safe_name("_".join(parts))
    return f"notice_{idx:04d}"


# ── security overlay: watermark + guilloche + QR code ──────────
#
# Each generated DOCX gets three layered security artifacts:
#   1. A full-page PNG overlay in the header (behindDoc=1) containing a
#      light guilloche pattern plus a diagonal bilingual watermark
#      "SS Legal Firm; for Respondent {name}" in English + Urdu.
#   2. A QR code anchored to the last body paragraph, absolute-positioned
#      at the bottom-right of the page (inside the body area, above the
#      footer). The QR payload is the English watermark text only — the
#      Urdu line is skipped to keep the code dense enough to scan.
#
# The overlay PNG is generated per-row because the name is personalized;
# PIL + arabic_reshaper + python-bidi handle Urdu shaping and bi-di
# reordering. If any of those deps are missing, the whole security pass
# is silently skipped (placeholder replacement still runs).

ENGLISH_WATERMARK_TEMPLATE = "SS Legal Firm; for Respondent {name}"
URDU_WATERMARK_TEMPLATE = "ایس ایس لیگل فرم؛ جواب دہندہ {name} کے لیے"

# Prefer a single font that covers both Arabic script (Urdu) and Latin.
# Mac ships Arial Unicode MS which is ideal; Debian/Ubuntu gets DejaVu
# Sans from fonts-dejavu-core or Noto Sans from fonts-noto-core.
_MIXED_FONT_CANDIDATES = (
    "/System/Library/Fonts/Supplemental/Arial Unicode.ttf",
    "/Library/Fonts/Arial Unicode.ttf",
    "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
    "/usr/share/fonts/truetype/noto/NotoSans-Regular.ttf",
    "/usr/share/fonts/TTF/DejaVuSans.ttf",
)


def _pick_font_path():
    for p in _MIXED_FONT_CANDIDATES:
        if os.path.exists(p):
            return p
    return None


def _shape_urdu(text):
    """Reshape + bi-di reorder so PIL's left-to-right drawer renders Urdu correctly."""
    return get_display(arabic_reshaper.reshape(text))


# ── security overlay config (watermark + anti-counterfeit pattern) ──
#
# Each template stores its own `security` block in meta.json; callers pass
# it to _render_security_overlay_png() to drive the generator. Missing
# keys fall back to these defaults.

DEFAULT_SECURITY_CONFIG = {
    "watermark": {
        "enabled": True,
        "english_template": "SS Legal Firm; for Respondent {name}",
        "urdu_template":    "ایس ایس لیگل فرم؛ جواب دہندہ {name} کے لیے",
        "font_size":  22,     # px in the 1200-wide overlay canvas
        "opacity":    67,     # percent, 0–100
        "count":      25,     # total tiles per page; mapped to a near-square grid
        "color":      "#323255",
    },
    "pattern": {
        "enabled": True,
        "opacity": 22,        # percent, 0–100 (avg across curves)
        "density": "medium",  # low | medium | high | ultra
    },
}

# Density → (sine curve count, spiral count, ring step px).
PATTERN_DENSITY_PROFILES = {
    "low":    (20,  4, 12),
    "medium": (40,  8,  8),
    "high":   (60, 12,  6),
    "ultra":  (80, 16,  5),
}


def _merge_security_config(user_cfg):
    """Deep-merge user-supplied security config on top of defaults."""
    cfg = {k: dict(v) for k, v in DEFAULT_SECURITY_CONFIG.items()}
    if not user_cfg:
        return cfg
    for section in ("watermark", "pattern"):
        if section in user_cfg and isinstance(user_cfg[section], dict):
            cfg[section].update(user_cfg[section])
    return cfg


def _hex_to_rgb(hex_str):
    """'#RRGGBB' → (r, g, b). Silently ignore bad input, return slate."""
    try:
        s = hex_str.lstrip("#")
        if len(s) == 3:
            s = "".join(ch * 2 for ch in s)
        return (int(s[0:2], 16), int(s[2:4], 16), int(s[4:6], 16))
    except Exception:
        return (50, 50, 85)


def _render_security_overlay_png(name, config=None,
                                 width_px=1200, height_px=1697):
    """Return PNG bytes: transparent full-page overlay with pattern + watermark.

    `config` follows DEFAULT_SECURITY_CONFIG shape. Per-document random
    seed derives from `name` so each notice has its own unique guilloche
    fingerprint (harder to forge by copy-paste across documents).

    Thin wrapper around `_render_security_overlay_png_cached` — canonical-
    JSON of the config + name + dimensions form the cache key, so the live-
    preview POST (fires on every 300 ms debounced keystroke) no longer
    re-draws the 40+ curve guilloche and 25 watermark tiles on each tick.
    Cache is per-process; in the ProcessPool batch path each worker has
    its own cache and gets identical hits row-after-row within a batch
    since the config doesn't change.
    """
    cfg = _merge_security_config(config)
    cache_key = json.dumps(cfg, sort_keys=True, ensure_ascii=False)
    return _render_security_overlay_png_cached(
        cache_key, name, width_px, height_px)


@functools.lru_cache(maxsize=64)
def _render_security_overlay_png_cached(cfg_key, name, width_px, height_px):
    cfg = json.loads(cfg_key)
    return _render_security_overlay_png_impl(name, cfg, width_px, height_px)


def _render_security_overlay_png_impl(name, cfg, width_px, height_px):
    import random
    rng = random.Random(_name_seed(name))

    img = Image.new("RGBA", (width_px, height_px), (255, 255, 255, 0))
    draw = ImageDraw.Draw(img)

    # ─── Pattern layers ─────────────────────────────────────────
    if cfg["pattern"].get("enabled", True):
        p_op = max(0, min(100, int(cfg["pattern"].get("opacity", 22))))
        # Map 0–100% to a ±15% alpha band centered on the requested level.
        a_mid = int(p_op * 2.55)
        pa_min = max(8,   a_mid - 12)
        pa_max = min(220, a_mid + 18)
        density = cfg["pattern"].get("density", "medium")
        n_curves, n_spirals, ring_step = PATTERN_DENSITY_PROFILES.get(
            density, PATTERN_DENSITY_PROFILES["medium"])

        # Layer 1: sine curves with 3rd-harmonic wobble
        for _ in range(n_curves):
            alpha = rng.randint(pa_min, pa_max)
            color = (rng.randint(60, 110), rng.randint(110, 160),
                     rng.randint(160, 210), alpha)
            f1 = rng.uniform(0.0028, 0.012)
            f2 = f1 * rng.uniform(1.5, 3.0)
            f3 = f1 * rng.uniform(4.0, 6.0)
            amp = rng.uniform(80, 320)
            pa_ = rng.uniform(0, 6.283)
            pb_ = rng.uniform(0, 6.283)
            pc_ = rng.uniform(0, 6.283)
            base = rng.uniform(0, height_px)
            prev = None
            for x in range(0, width_px, 2):
                y = (base
                     + math.sin(x * f1 + pa_) * amp
                     + math.sin(x * f2 + pb_) * (amp * 0.30)
                     + math.sin(x * f3 + pc_) * (amp * 0.08))
                pt = (x, y)
                if prev is not None:
                    draw.line([prev, pt], fill=color, width=1)
                prev = pt

        # Layer 2: Archimedean spirals scattered around
        for _ in range(n_spirals):
            cx = rng.uniform(0, width_px)
            cy = rng.uniform(0, height_px)
            alpha = rng.randint(max(4, pa_min - 10), max(20, pa_max - 10))
            color = (rng.randint(60, 110), rng.randint(110, 160),
                     rng.randint(160, 210), alpha)
            prev = None
            t_end = rng.uniform(18.0, 32.0)
            a = rng.uniform(2.0, 5.5)
            t = 0.0
            while t < t_end:
                r = a * t
                x = cx + math.cos(t) * r
                y = cy + math.sin(t) * r
                if 0 <= x < width_px and 0 <= y < height_px:
                    pt = (x, y)
                    if prev is not None:
                        draw.line([prev, pt], fill=color, width=1)
                    prev = pt
                else:
                    prev = None
                t += 0.06

        # Layer 3: concentric rings around jittered center
        rcx = width_px / 2 + rng.uniform(-100, 100)
        rcy = height_px / 2 + rng.uniform(-120, 120)
        ring_color = (100, 140, 200,
                      rng.randint(max(4, pa_min - 15),
                                  max(20, pa_max - 15)))
        for r in range(14, 1300, ring_step):
            draw.ellipse([rcx - r, rcy - r, rcx + r, rcy + r],
                         outline=ring_color, width=1)

    # ─── Tiled bilingual watermark ──────────────────────────────
    if cfg["watermark"].get("enabled", True):
        wm = cfg["watermark"]
        font_px  = max(10, min(64, int(wm.get("font_size", 22))))
        w_op     = max(0, min(100, int(wm.get("opacity", 67))))
        w_alpha  = int(w_op * 2.55)
        w_rgb    = _hex_to_rgb(wm.get("color", "#323255"))
        count    = max(1, min(100, int(wm.get("count", 25))))

        # Grid geometry: use a near-square grid that's slightly taller
        # than wide to match A4 aspect.
        cols = max(1, int(math.sqrt(count * (width_px / height_px))))
        rows = max(1, int(math.ceil(count / cols)))

        font_path = _pick_font_path()
        font = (ImageFont.truetype(font_path, font_px)
                if font_path else ImageFont.load_default())

        english = wm.get("english_template", "").format(name=name)
        try:
            urdu = _shape_urdu(wm.get("urdu_template", "").format(name=name))
        except Exception:
            urdu = wm.get("urdu_template", "").format(name=name)

        en_bb = draw.textbbox((0, 0), english, font=font)
        ur_bb = draw.textbbox((0, 0), urdu, font=font)
        tile_w = max(en_bb[2] - en_bb[0], ur_bb[2] - ur_bb[0]) + 8
        tile_h = (en_bb[3] - en_bb[1]) + (ur_bb[3] - ur_bb[1]) + 6
        if tile_w > 0 and tile_h > 0:
            tile = Image.new("RGBA", (tile_w, tile_h), (255, 255, 255, 0))
            td = ImageDraw.Draw(tile)
            stamp = (*w_rgb, w_alpha)
            td.text(((tile_w - (en_bb[2] - en_bb[0])) / 2, 0),
                    english, font=font, fill=stamp)
            if urdu.strip():
                td.text(((tile_w - (ur_bb[2] - ur_bb[0])) / 2,
                         (en_bb[3] - en_bb[1]) + 4),
                        urdu, font=font, fill=stamp)
            rotated = tile.rotate(30, expand=True, resample=Image.BICUBIC)
            rw, rh = rotated.size

            col_spacing = width_px / cols
            row_spacing = height_px / rows
            for row in range(rows):
                for col in range(cols):
                    if (row * cols + col) >= count:
                        break
                    stagger = (col_spacing / 2) if (row % 2) else 0
                    cx_ = col * col_spacing + col_spacing * 0.5 + stagger
                    cy_ = row * row_spacing + row_spacing * 0.5
                    jx = rng.uniform(-18, 18)
                    jy = rng.uniform(-12, 12)
                    img.paste(rotated,
                              (int(cx_ - rw / 2 + jx),
                               int(cy_ - rh / 2 + jy)),
                              rotated)

    buf = io.BytesIO()
    img.save(buf, format="PNG", optimize=True)
    return buf.getvalue()


@functools.lru_cache(maxsize=256)
def _render_qr_png(text):
    """Return PNG bytes for a medium-error-correction QR code.

    Cached because batch rendering of 5000 rows produces 5000 QRs whose
    text only varies in the `{name}` slot — identical text → identical
    PNG, and QR encoding is the slowest per-row step after rasterization.
    The cache lives per-process so each ProcessPool worker warms
    independently."""
    qr = qrcode.QRCode(
        error_correction=qrcode.constants.ERROR_CORRECT_M,
        box_size=10, border=2)
    qr.add_data(text)
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white")
    buf = io.BytesIO()
    img.save(buf, format="PNG")
    return buf.getvalue()


# ── HTML render pipeline (task 9) ────────────────────────────────
#
# Per row, produce a final non-extractable PDF by:
#   1. Assembling a self-contained HTML string (base template with
#      block markup swapped in, placeholders filled, security overlay
#      + QR code + seal rotations baked into inline <style>/<img>).
#   2. Rendering via headless Chromium to an intermediate vector PDF.
#   3. Rasterizing every page at 300 dpi through PyMuPDF and re-embedding
#      each page as a single image — this strips every text object so
#      no PDF editor can select/copy/modify the text layer.
#   4. Applying a permissions lock (pikepdf AES-256) that signals
#      no-copy/no-modify and sets clean metadata.
#
# The pipeline is process-pool safe: each worker boots its own Chromium
# once at startup (initializer) and reuses the browser across rows.

try:
    from playwright.sync_api import sync_playwright as _sync_playwright
    import pymupdf as _pymupdf
    import pikepdf as _pikepdf
    _HTML_RENDER_DEPS_OK = True
except ImportError:
    _HTML_RENDER_DEPS_OK = False


PDF_PRODUCER = "S&S Law Firm Legal Notice Generator"
PDF_TITLE_DEFAULT = "Legal Notice"


# ── single-notice Chromium pool ─────────────────────────────────
#
# /generate_one renders one PDF per request. Originally it called
# render_notice_row_pdf(browser=None) which booted a fresh Playwright +
# Chromium on every request — ~1s of pure startup tax per click. Keep a
# shared singleton browser protected by a lock so the hot path is
# render-only. Gunicorn runs --workers 1, so one singleton per app.
_SINGLE_BROWSER = None
_SINGLE_PW = None
_SINGLE_BROWSER_LOCK = threading.Lock()

# Concurrency cap on /generate_one. Each in-flight render is ~300 MB RSS
# + a CPU core for rasterization — without this gate a loop of clicks
# exhausts memory on small VPSes. 4 is a reasonable ceiling even for a
# 1-vCPU box because each request is mostly I/O-bound once rendering
# hands off to Chromium. 429 on contention so the client retries.
_SINGLE_GEN_SEM = threading.Semaphore(4)


def _get_shared_single_browser():
    """Lazy-launch a single reusable Chromium for /generate_one."""
    global _SINGLE_BROWSER, _SINGLE_PW
    with _SINGLE_BROWSER_LOCK:
        if _SINGLE_BROWSER is None:
            _SINGLE_PW = _sync_playwright().start()
            _SINGLE_BROWSER = _SINGLE_PW.chromium.launch()
        return _SINGLE_BROWSER


# ── notice-generation history (audit + reverse-lookup) ──────────
#
# Every successful PDF render writes one row into SQLite:
#   (serial, name, principal, generated_at)
# The 16-char serial `XXXX-XXXX-XXXX-XXXX` printed under the QR is the
# lookup key — the firm can type/paste a serial from a received letter
# into /verify and confirm WHO it was issued to and WHEN. Capped at
# HISTORY_MAX_ROWS (2M); oldest rows roll off when the cap is exceeded.
# 2M rows ≈ ~200 MB on disk, still O(ms) indexed lookup by serial.

HISTORY_DB_PATH = os.path.join(UPLOAD_DIR, "history.db")
HISTORY_MAX_ROWS = 2_000_000
_HISTORY_LOCK = threading.Lock()


def _init_history_db():
    """Create the table + index on first boot. Idempotent."""
    with sqlite3.connect(HISTORY_DB_PATH) as conn:
        # WAL mode lets readers and writers coexist; important because
        # /verify reads while batch workers write.
        conn.execute("PRAGMA journal_mode=WAL")
        conn.execute("""
            CREATE TABLE IF NOT EXISTS notice_history (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                serial TEXT NOT NULL,
                name TEXT NOT NULL,
                principal TEXT,
                generated_at TEXT NOT NULL
            )
        """)
        conn.execute("CREATE INDEX IF NOT EXISTS idx_history_serial "
                     "ON notice_history(serial)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_history_created "
                     "ON notice_history(generated_at)")
        conn.commit()


_init_history_db()


def _log_notice_record(name, principal, serial=None, when=None):
    """Append one row to the history log.

    `serial` defaults to generate_qr_serial(name) so call sites that only
    have (name, principal) don't have to know about the crypto scheme.
    """
    if serial is None:
        serial = generate_qr_serial(name)
    if when is None:
        when = datetime.datetime.now().isoformat(timespec="seconds")
    with _HISTORY_LOCK:
        try:
            with sqlite3.connect(HISTORY_DB_PATH) as conn:
                conn.execute(
                    "INSERT INTO notice_history "
                    "(serial, name, principal, generated_at) "
                    "VALUES (?, ?, ?, ?)",
                    (serial, str(name or ""), str(principal or ""), when))
                # Cap check only runs when we risk being over — cheap.
                count = conn.execute(
                    "SELECT COUNT(*) FROM notice_history").fetchone()[0]
                if count > HISTORY_MAX_ROWS:
                    over = count - HISTORY_MAX_ROWS
                    conn.execute(
                        "DELETE FROM notice_history WHERE id IN "
                        "(SELECT id FROM notice_history "
                        "ORDER BY id ASC LIMIT ?)", (over,))
                conn.commit()
        except sqlite3.Error as e:
            # History logging must never break the render pipeline.
            print(f"[history] log failed: {e}", flush=True)


def lookup_notice_by_serial(serial):
    """Return {name, principal, generated_at} for a serial, or None."""
    if not serial or len(serial) > 40:
        return None
    with sqlite3.connect(HISTORY_DB_PATH) as conn:
        conn.row_factory = sqlite3.Row
        row = conn.execute(
            "SELECT name, principal, generated_at FROM notice_history "
            "WHERE serial = ? ORDER BY id DESC LIMIT 1",
            (serial.strip(),)).fetchone()
    return dict(row) if row else None


# ── cases store (inventory mode) ──────────────────────────────────
#
# Inventory mode keeps every imported case row in a SQLite table so the
# user can bring in a whole portfolio once (hundreds of thousands of rows
# from the same-format batch Excel) and then pick subsets by
# CNIC / name / phone on demand — without re-uploading the spreadsheet
# each time a letter needs to go out.
#
# Uniqueness key is the "订单编号" (order number) column. CNIC repeats
# across multiple loans for the same borrower, but every loan has its
# own order number, so that's what drives conflict handling on import.
#
# `row_json` stores the full original row as-is so render-time code can
# consume it unchanged — including columns the current template doesn't
# use (e.g. 负责人). The redundant cnic / name / phone columns exist only
# to index the three supported search fields; they duplicate data from
# row_json but let LIKE queries hit an index on a 500 k-row table.

CASES_DB_PATH = os.path.join(UPLOAD_DIR, "cases.db")
_CASES_LOCK = threading.Lock()   # serializes writers; WAL lets reads bypass

CASES_IMPORTS_DIR = os.path.join(UPLOAD_DIR, "case_imports")
os.makedirs(CASES_IMPORTS_DIR, exist_ok=True)

# Excel header aliases. The real workload has Chinese-headed columns
# (订单编号, 姓名, CNIC, 注册手机号) on one sheet and Latin-headed
# (name, cnic, phone) on another — both come through the same importer.
# First exact-match hit wins; a case-insensitive pass runs second.
CASE_COL_ALIASES = {
    "order_id": ("订单编号", "order_id", "Order_ID", "OrderID",
                 "order id", "order no", "order_no"),
    "cnic":     ("cnic", "CNIC", "身份证", "身份证号"),
    "name":     ("name", "Name", "姓名"),
    "phone":    ("phone", "Phone", "注册手机号", "手机号",
                 "mobile", "Mobile"),
}


def _find_col(headers, aliases):
    header_set = {h for h in headers if h}
    for a in aliases:
        if a in header_set:
            return a
    lower_map = {h.lower(): h for h in headers if h}
    for a in aliases:
        hit = lower_map.get(a.lower())
        if hit:
            return hit
    return None


def _init_cases_db():
    with sqlite3.connect(CASES_DB_PATH) as conn:
        # WAL so search queries don't block behind a 40-万-row import tx.
        conn.execute("PRAGMA journal_mode=WAL")
        conn.execute("""
            CREATE TABLE IF NOT EXISTS cases (
                order_id      TEXT PRIMARY KEY,
                cnic          TEXT,
                name          TEXT,
                original_name TEXT,
                phone         TEXT,
                row_json      TEXT NOT NULL,
                created_at    TEXT NOT NULL,
                updated_at    TEXT NOT NULL
            )
        """)
        # Backfill for DBs created before original_name existed. The
        # ALTER is a no-op on fresh installs; the UPDATE seeds existing
        # rows so "orig:" shows a real value after upgrade even for
        # cases that were imported before the audit trail was added.
        cols = {r[1] for r in conn.execute("PRAGMA table_info(cases)")}
        if "original_name" not in cols:
            conn.execute("ALTER TABLE cases ADD COLUMN original_name TEXT")
        conn.execute(
            "UPDATE cases SET original_name = name "
            "WHERE original_name IS NULL")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_cases_cnic  ON cases(cnic)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_cases_name  ON cases(name)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_cases_phone ON cases(phone)")
        conn.commit()


_init_cases_db()


def _normalize_order_id(v):
    """Coerce an order-id cell to a clean non-empty string, or None.

    Real order numbers are 19-digit bigints (e.g. 7410924208336667649)
    which overflow SQLite's INTEGER; always store as TEXT. Integer-valued
    floats are de-scientific-notation'd first so the key survives
    Excel's auto-typing ('3674103' stays '3674103', not '3674103.0').
    """
    if v is None:
        return None
    if isinstance(v, float):
        if v.is_integer():
            v = int(v)
    s = str(v).strip()
    return s or None


def _cases_count():
    with sqlite3.connect(CASES_DB_PATH) as conn:
        return conn.execute("SELECT COUNT(*) FROM cases").fetchone()[0]


def _cases_existing_ids(ids):
    """Return the subset of `ids` already present in the DB, chunked to
    stay under SQLite's 999-parameter limit."""
    if not ids:
        return set()
    found = set()
    with sqlite3.connect(CASES_DB_PATH) as conn:
        for i in range(0, len(ids), 500):
            chunk = ids[i:i + 500]
            qs = ",".join("?" * len(chunk))
            for (oid,) in conn.execute(
                    f"SELECT order_id FROM cases WHERE order_id IN ({qs})",
                    tuple(chunk)):
                found.add(oid)
    return found


INVENTORY_BATCH_DIR = os.path.join(UPLOAD_DIR, "inventory_batches")
os.makedirs(INVENTORY_BATCH_DIR, exist_ok=True)
_BATCH_COUNTER_LOCK = threading.Lock()

# Match "2026-04-18_batch-017.zip" — used to find the next counter
# value for today's batches when naming the outer zip.
_BATCH_NAME_RE = re.compile(r"^\d{4}-\d{2}-\d{2}_batch-(\d{3})\.zip$")


def _next_batch_number(date_str):
    """Scan `INVENTORY_BATCH_DIR` for today's batches and return the
    next sequence number. Serialized under `_BATCH_COUNTER_LOCK` so two
    generate tasks firing back-to-back don't collide on the same N."""
    with _BATCH_COUNTER_LOCK:
        used = []
        try:
            for fn in os.listdir(INVENTORY_BATCH_DIR):
                if not fn.startswith(date_str + "_"):
                    continue
                m = _BATCH_NAME_RE.match(fn)
                if m:
                    used.append(int(m.group(1)))
        except OSError:
            pass
        return (max(used) + 1) if used else 1


def _wrap_inventory_batch(task_id):
    """Post-process an inventory-mode generate task: bundle every group
    zip into a single outer zip named `YYYY-MM-DD_batch-NNN.zip`,
    replace the task's ready_parts with that single entry, and delete
    the now-redundant per-group zips from the working dir. No-op if the
    task isn't in a terminal `done` state or has no parts.

    The outer zip is ZIP_STORED because the inner zips are already
    DEFLATED — re-compressing them wastes CPU for no size gain.
    """
    task = _get_task(task_id)
    if not task or task.get("status") != "done":
        return
    parts = list(task.get("ready_parts") or [])
    if not parts:
        return

    today = datetime.date.today().isoformat()
    batch_num = _next_batch_number(today)
    outer_name = f"{today}_batch-{batch_num:03d}.zip"
    outer_path = os.path.join(INVENTORY_BATCH_DIR, outer_name)

    inner_taken = set()
    with zipfile.ZipFile(outer_path, "w", zipfile.ZIP_STORED) as outer:
        for p in parts:
            src = p.get("path")
            if not src or not os.path.exists(src):
                continue
            # Respect user-visible inner name (already includes the
            # 负责人 value, sanitized). Dedupe if two groups collapsed
            # to the same safe name.
            arcname = p.get("name") or os.path.basename(src)
            base = arcname
            n = 1
            while arcname in inner_taken:
                n += 1
                stem, ext = os.path.splitext(base)
                arcname = f"{stem}_{n}{ext}"
            inner_taken.add(arcname)
            outer.write(src, arcname)

    # Delete the per-group zips now that they're bundled.
    for p in parts:
        src = p.get("path")
        if src and os.path.exists(src):
            try: os.remove(src)
            except OSError: pass

    # Replace ready_parts with a single persistent entry. `persistent`
    # tells /download/<tid>/<idx> to skip the 10-minute cleanup timer
    # so the batch zip stays available for re-download.
    with TASKS_LOCK:
        t = TASKS.get(task_id)
        if t is not None:
            t["ready_parts"] = [{
                "index": 0,
                "name": outer_name,
                "group": "batch",
                "path": outer_path,
                "persistent": True,
                "inner_count": len(inner_taken),
            }]
            t["message"] = (
                f"Packed {len(inner_taken)} group zip(s) into {outer_name}")


def _parse_multi_values(q):
    """Split a search input into a deduped list of exact values.

    One value falls through to the existing prefix-LIKE fast path; two
    or more trigger an IN-style exact-match lookup. Splits on commas,
    whitespace, and newlines so the same input works whether the user
    typed a comma-separated list or pasted from a spreadsheet column.

    Capped at 10k entries so a runaway paste can't eat the backend."""
    if not q:
        return []
    parts = re.split(r"[\s,]+", q)
    seen, out = set(), []
    for p in parts:
        p = p.strip()
        if not p or p in seen:
            continue
        seen.add(p)
        out.append(p)
        if len(out) >= 10_000:
            break
    return out


def _cases_temp_values_table(conn, values):
    """Materialize `values` into a per-connection TEMP table so queries
    can JOIN against it — this bypasses SQLite's 999-bind-parameter cap
    that an inline `IN (?,?,...)` would hit for multi-thousand lists.
    The table is re-created each call so it's safe to reuse the same
    connection for back-to-back searches."""
    conn.execute(
        "CREATE TEMP TABLE IF NOT EXISTS _srch_q (v TEXT PRIMARY KEY)")
    conn.execute("DELETE FROM _srch_q")
    conn.executemany("INSERT OR IGNORE INTO _srch_q VALUES (?)",
                     [(v,) for v in values])


def _cases_fetch_rows(ids):
    """Load row_json for a list of order_ids, preserving input order.
    Returns (rows_in_order, missing_ids)."""
    by_id = {}
    if ids:
        with sqlite3.connect(CASES_DB_PATH) as conn:
            for i in range(0, len(ids), 500):
                chunk = ids[i:i + 500]
                qs = ",".join("?" * len(chunk))
                for oid, payload in conn.execute(
                        f"SELECT order_id, row_json FROM cases "
                        f"WHERE order_id IN ({qs})", tuple(chunk)):
                    by_id[oid] = json.loads(payload)
    rows, missing = [], []
    for oid in ids:
        row = by_id.get(oid)
        if row is None:
            missing.append(oid)
        else:
            rows.append(row)
    return rows, missing


_ANY_OPEN_DIV = re.compile(r'<div\b[^>]*>')
_ANY_CLOSE_DIV = re.compile(r'</div\s*>')


def _swap_block_inner(html, purpose, inner_html):
    """Replace the innerHTML of every <div data-purpose="PURPOSE"> ... </div>.

    Handles nested <div>s by walking the source and counting depth —
    a naive regex can't match the correct </div> when the block body
    contains its own divs (meta grid, amounts table wrappers, etc.).
    Replaces ALL occurrences, so blocks that repeat on every page
    (like the contact footer) stay in sync.
    """
    open_re = re.compile(
        r'<div[^>]*\sdata-purpose="' + re.escape(purpose) + r'"[^>]*>')
    chunks = []
    cursor = 0
    while True:
        m = open_re.search(html, cursor)
        if not m:
            chunks.append(html[cursor:])
            break
        chunks.append(html[cursor:m.end()])
        i = m.end()
        depth = 1
        close_pos = None
        while i < len(html) and depth > 0:
            mo = _ANY_OPEN_DIV.search(html, i)
            mc = _ANY_CLOSE_DIV.search(html, i)
            if mc is None:
                return html  # malformed — bail without mutating
            if mo is not None and mo.start() < mc.start():
                depth += 1
                i = mo.end()
            else:
                depth -= 1
                if depth == 0:
                    close_pos = mc.start()
                    break
                i = mc.end()
        if close_pos is None:
            return html  # unbalanced — bail without mutating
        chunks.append("\n" + inner_html + "\n")
        cursor = close_pos
    return "".join(chunks)


def _rewrite_static_to_relative(html):
    """Rewrite absolute /static/… → relative static/… so the HTML can
    be written anywhere inside templates/ and loaded via file://,
    letting Chromium resolve fonts/images from the sibling static/
    directory without a server. Flask still serves /static/ for the
    live browser preview because @app.route("/static/...") matches."""
    return (html
            .replace('"/static/', '"static/')
            .replace("'/static/", "'static/"))


def _swap_asset_src(html, slot, src_url):
    """Swap the src of <img data-slot="SLOT" src="..."> to src_url."""
    return re.sub(
        r'(<img[^>]*\sdata-slot="' + re.escape(slot) + r'"[^>]*\ssrc=")[^"]*(")',
        lambda m: m.group(1) + src_url + m.group(2),
        html, count=1)


# 32-char alphabet used for the anti-counterfeit serial printed under
# the QR. 0/O and 1/I are omitted so the string stays readable when
# someone transcribes it from a paper copy.
_QR_SN_ALPHABET = "ABCDEFGHJKLMNPQRSTUVWXYZ23456789"


def generate_qr_serial(name):
    """Deterministic 16-character (4×4 groups) anti-counterfeit code.

    Seeded by hash(name) so the SAME respondent always gets the SAME
    code — lets the firm reverse-lookup authenticity — but different
    respondents get statistically different codes, so an attacker can't
    copy a known-good notice and swap names."""
    srng = random.Random(_name_seed(name))
    chars = [srng.choice(_QR_SN_ALPHABET) for _ in range(16)]
    return "-".join("".join(chars[i:i + 4]) for i in (0, 4, 8, 12))


def build_notice_html(template, row_data, manual_fields=None,
                     base_html=None, static_dir=None,
                     fill_placeholders=True, strip_unfilled=None):
    """Produce a fully self-contained HTML string for one row.

    * template          — dict from load_template() (blocks, assets, security)
    * row_data          — placeholder dict from an Excel row
    * manual_fields     — optional overrides for unmatched placeholders
    * base_html         — cached base template text (re-read if None)
    * static_dir        — absolute path to templates/static (uses default)
    * fill_placeholders — run the `{{foo}}` → value substitution pass.
                          False means the bare template is returned
                          (literal `{{name}}` etc.).
    * strip_unfilled    — after substitution, regex-strip any leftover
                          `{{foo}}` to empty string. Defaults to
                          fill_placeholders so real renders drop
                          unfilled keys but the manual live-preview
                          (partial fill) keeps them visible as literal
                          placeholders for fields the user hasn't typed
                          yet.
    """
    if base_html is None:
        with open(TEMPLATE_BASE_HTML_PATH, "r", encoding="utf-8") as f:
            base_html = f.read()
    if static_dir is None:
        static_dir = TEMPLATE_STATIC_DIR
    if strip_unfilled is None:
        strip_unfilled = fill_placeholders

    merged = {**(manual_fields or {}), **(row_data or {})}
    # Name resolution: drives watermark, QR text, per-doc random
    # rotations, and the anti-counterfeit serial.
    #   - bare template preview (fill_placeholders=False) → "{name}"
    #   - real render (fill + strip) → typed value or "Respondent"
    #   - manual live preview (fill + no strip) → typed value or
    #     "{name}" so unfilled placeholders stay visible
    if not fill_placeholders:
        name = "{name}"
    else:
        typed = str(merged.get("name", "") or "").strip()
        if typed:
            name = typed
        elif strip_unfilled:
            name = "Respondent"
        else:
            name = "{name}"

    # 1. Swap editable blocks with rendered HTML fragments.
    blocks_text = template.get("blocks") or dict(DEFAULT_BLOCKS_TEXT)
    rendered = render_blocks_to_html(blocks_text)
    html = base_html
    for purpose, inner in rendered.items():
        html = _swap_block_inner(html, purpose, inner)

    # 2. Swap any per-template asset overrides. These are absolute
    # filesystem paths, not /static URLs; we inline them as data URIs
    # so they travel with the HTML into Chromium's sandboxed renderer.
    for kind, path in (template.get("assets") or {}).items():
        if not path or not os.path.isfile(path):
            continue
        try:
            with open(path, "rb") as f:
                data = f.read()
            b64 = base64.b64encode(data).decode()
            html = _swap_asset_src(html, kind, f"data:image/png;base64,{b64}")
        except OSError:
            pass

    # 3. Fill {{placeholders}}. Substitute every key in merged, then
    # optionally strip leftover `{{foo}}` tokens to empty strings. The
    # strip is what makes the final PDF clean of unfilled placeholders;
    # the manual live-preview keeps strip_unfilled=False so un-typed
    # fields stay visible as literal `{{foo}}`.
    #
    # SECURITY: values come from Excel cells, Manual form fields, or the
    # preview POST body — all attacker-influenceable. We MUST HTML-escape
    # before splicing into the DOM, otherwise `<img src=x onerror=...>`
    # in a name column executes in the live-preview iframe (same origin,
    # full session cookie) and inside Chromium during PDF render (file://
    # origin, can read local files). The structured block renderers
    # already escape via `_esc_text`; this flat replace was the leak.
    if fill_placeholders:
        for k, v in merged.items():
            safe = _htmllib.escape(_format_value(v), quote=False)
            html = html.replace("{{" + k + "}}", safe)
        if strip_unfilled:
            html = PLACEHOLDER_RE.sub("", html)

    # 4. Rewrite /static → static so the HTML loads fonts/images
    # relative to wherever we park it on disk (e.g., inside templates/).
    html = _rewrite_static_to_relative(html)

    # 4.5 Anti-counterfeit serial under the QR. Deterministic per-name,
    # so the SAME respondent always gets the SAME code across reruns
    # (lets the firm reverse-lookup authenticity). Generated regardless
    # of whether the security/QR deps are installed.
    html = html.replace("__QR_SN__", generate_qr_serial(name))

    # 5. Inject security overlay + QR + per-doc seal rotations +
    #    per-template asset placement vars.
    assets_cfg = template.get("assets_config") or DEFAULT_ASSETS_CONFIG
    def _asset_vars():
        # CSS var prefix per asset kind — sig/seal/logo
        prefix = {"logo": "logo", "seal": "seal", "signature_seal": "sig"}
        parts = []
        for kind, cfg in assets_cfg.items():
            px = prefix[kind]
            parts.append(f"--{px}-size: {cfg['size']}mm;")
            parts.append(f"--{px}-dx:   {cfg['dx']}mm;")
            parts.append(f"--{px}-dy:   {cfg['dy']}mm;")
            parts.append(f"--{px}-rot-base: {cfg['rot']}deg;")
        return " ".join(parts)

    if _SECURITY_DEPS_OK:
        sec_cfg = template.get("security") or load_default_security()
        overlay_png = _render_security_overlay_png(name, config=sec_cfg)
        overlay_b64 = ("data:image/png;base64,"
                       + base64.b64encode(overlay_png).decode())

        wm = sec_cfg.get("watermark", {})
        qr_text = (wm.get("english_template") or "").format(name=name)
        if not qr_text.strip():
            qr_text = f"SS Legal Firm; for Respondent {name}"
        qr_png = _render_qr_png(qr_text)
        qr_b64 = "data:image/png;base64," + base64.b64encode(qr_png).decode()

        srng = random.Random(_name_seed(name))
        firm_rot = srng.uniform(-15, 15)
        sig_rot  = srng.uniform(-15, 15)
        injection = (
            f'<style id="security-overlay">'
            f'  .page {{'
            f'    background-image: url("{overlay_b64}");'
            f'    background-size: 100% 100%;'
            f'    background-repeat: no-repeat;'
            f'    --firm-rot: {firm_rot:.2f}deg;'
            f'    --sig-rot:  {sig_rot:.2f}deg;'
            f'    {_asset_vars()}'
            f'  }}'
            f'</style>'
        )
        html = html.replace("</head>", injection + "</head>", 1)
        html = html.replace(
            '<div class="placeholder">QR</div>',
            f'<img src="{qr_b64}" alt="QR" '
            f'style="width:100%;height:100%;object-fit:contain;">')
    else:
        # Security deps missing (no watermark/QR), but still expose the
        # asset-placement vars so resizing/positioning still works.
        injection = (
            f'<style id="asset-vars">'
            f'  .page {{ {_asset_vars()} }}'
            f'</style>'
        )
        html = html.replace("</head>", injection + "</head>", 1)
    return html


def _playwright_render_html_to_pdf(html, output_pdf, browser=None):
    """Render HTML string → A4 PDF via headless Chromium.

    The HTML is dropped as a temp file inside templates/ (so sibling
    `static/` folder resolves correctly) and loaded via file:// URL,
    avoiding the about:blank base-URL issue that breaks relative
    asset loading with page.set_content()."""
    pw_ctx = None
    own_browser = False
    if browser is None:
        pw_ctx = _sync_playwright().start()
        browser = pw_ctx.chromium.launch()
        own_browser = True

    # Write HTML into the templates dir so relative `static/…` resolves.
    tmp_name = f"_render_{uuid.uuid4().hex}.html"
    tmp_path = os.path.join(TEMPLATE_BASE_DIR, tmp_name)
    try:
        with open(tmp_path, "w", encoding="utf-8") as f:
            f.write(html)
        url = "file://" + tmp_path
        page = browser.new_page()
        try:
            page.goto(url, wait_until="networkidle")
            page.emulate_media(media="print")
            page.pdf(path=output_pdf, format="A4",
                     margin={"top": "0", "right": "0",
                             "bottom": "0", "left": "0"},
                     print_background=True)
        finally:
            page.close()
    finally:
        try:
            os.remove(tmp_path)
        except OSError:
            pass
        if own_browser:
            browser.close()
            pw_ctx.stop()


def _rasterize_pdf(in_pdf, out_pdf, dpi=300):
    """Render every page at `dpi`, rebuild as image-only PDF.
    After this pass the PDF has no text objects — copying or editing
    the text layer is impossible in any PDF viewer/editor."""
    src = _pymupdf.open(in_pdf)
    dst = _pymupdf.open()
    try:
        for page in src:
            pix = page.get_pixmap(dpi=dpi, alpha=False)
            page_w, page_h = page.rect.width, page.rect.height
            new = dst.new_page(width=page_w, height=page_h)
            new.insert_image(new.rect, stream=pix.tobytes("png"))
        dst.save(out_pdf, garbage=4, deflate=True)
    finally:
        src.close()
        dst.close()


def _apply_pdf_lock(in_pdf, out_pdf, title=PDF_TITLE_DEFAULT,
                    producer=PDF_PRODUCER):
    """AES-256-encrypt with no-copy / no-modify permissions + metadata.
    The user password is empty so anyone can *open* the PDF; the owner
    password is a random 32-byte string nobody holds, so no one can
    downgrade the permissions. Printing stays allowed."""
    owner_pw = base64.b64encode(os.urandom(24)).decode()
    try:
        pdf = _pikepdf.open(in_pdf)
    except _pikepdf.PdfError:
        # If the input is already encrypted or malformed, fall back to
        # copying raw bytes rather than failing the whole row.
        shutil.copy(in_pdf, out_pdf)
        return
    try:
        with pdf.open_metadata() as meta:
            meta["dc:title"] = title
            meta["pdf:Producer"] = producer
        pdf.docinfo["/Title"] = title
        pdf.docinfo["/Producer"] = producer
        pdf.docinfo["/CreationDate"] = (
            "D:" + datetime.datetime.now().strftime("%Y%m%d%H%M%S") + "Z")
        permissions = _pikepdf.Permissions(
            accessibility=True,
            extract=False,              # disallow copy
            modify_annotation=False,
            modify_assembly=False,
            modify_form=False,
            modify_other=False,
            print_highres=True,
            print_lowres=True,
        )
        pdf.save(out_pdf,
                 encryption=_pikepdf.Encryption(
                     owner=owner_pw, user="", allow=permissions))
    finally:
        pdf.close()


def render_notice_row_pdf(template, row_data, output_pdf,
                          manual_fields=None, base_html=None,
                          static_dir=None, browser=None,
                          rasterize=True, lock=True,
                          dpi=300, title=PDF_TITLE_DEFAULT):
    """End-to-end: build HTML → Chromium → rasterize → lock.

    Returns `output_pdf`. `browser` is an optional pooled Chromium
    instance (supplied by process-pool initializer). `rasterize`/`lock`
    can be disabled for faster dev iteration; production keeps both on
    so copy/edit are blocked."""
    html = build_notice_html(template, row_data,
                             manual_fields=manual_fields,
                             base_html=base_html,
                             static_dir=static_dir)
    with tempfile.TemporaryDirectory(prefix="legal_notice_") as tmp:
        vector_pdf = os.path.join(tmp, "vector.pdf")
        _playwright_render_html_to_pdf(html, vector_pdf, browser=browser)
        if not rasterize and not lock:
            shutil.copy(vector_pdf, output_pdf)
            return output_pdf
        raster_pdf = vector_pdf
        if rasterize:
            raster_pdf = os.path.join(tmp, "raster.pdf")
            _rasterize_pdf(vector_pdf, raster_pdf, dpi=dpi)
        if lock:
            _apply_pdf_lock(raster_pdf, output_pdf, title=title)
        else:
            shutil.copy(raster_pdf, output_pdf)
    return output_pdf


# ── process-pool worker for batched HTML→PDF generation ──────────
#
# Each worker process boots ONE Chromium + ONE Playwright context at
# init time and reuses the browser across all jobs it handles. This
# amortizes the ~1-second browser-startup cost across the whole
# batch. Workers are re-spawned if the pool is re-initialized.

_WORKER_PW = None
_WORKER_BROWSER = None


def _html_worker_init():
    global _WORKER_PW, _WORKER_BROWSER
    _WORKER_PW = _sync_playwright().start()
    _WORKER_BROWSER = _WORKER_PW.chromium.launch()


def _html_worker_job(args):
    """Process-pool worker: render one row, return the output path."""
    (template, row_data, output_pdf, manual_fields,
     base_html, static_dir, dpi, title) = args
    return render_notice_row_pdf(
        template, row_data, output_pdf,
        manual_fields=manual_fields,
        base_html=base_html, static_dir=static_dir,
        browser=_WORKER_BROWSER,
        rasterize=True, lock=True, dpi=dpi, title=title)


def extract_template_placeholders(template=None):
    """Return placeholders in the order they visually appear in the
    rendered notice — so the manual-single form's inputs match the
    on-page reading order (date / TO: block → subject → narrative →
    amounts table → page 2).

    We swap the editable blocks into the base HTML (without touching
    `{{...}}` tokens) and then scan the result once. `dict.fromkeys`
    keeps each placeholder at its first appearance."""
    try:
        with open(TEMPLATE_BASE_HTML_PATH, "r", encoding="utf-8") as f:
            html = f.read()
    except OSError:
        return []
    if template:
        blocks_text = template.get("blocks") or dict(DEFAULT_BLOCKS_TEXT)
        rendered = render_blocks_to_html(blocks_text)
        for purpose, inner in rendered.items():
            html = _swap_block_inner(html, purpose, inner)
    ordered = PLACEHOLDER_RE.findall(html)
    return list(dict.fromkeys(ordered))


def _process_group_html(template, rows, group_name, manual_fields,
                        filename_fields, task_id, cumulative_before,
                        grand_total, pool, base_html, static_dir,
                        dpi, title):
    """Render every row in one group, pack into a zip, return zip path."""
    tmp = tempfile.mkdtemp(prefix=f"notice_{_safe_name(group_name)[:32]}_")
    try:
        # Build unique per-row output filenames.
        taken = set()
        jobs = []
        for i, record in enumerate(rows, start=1):
            merged = {**(manual_fields or {}), **record}
            base = _build_filename(merged, filename_fields, i)
            unique = base
            n = 1
            while unique in taken:
                n += 1
                unique = f"{base}_{n}"
            taken.add(unique)
            output = os.path.join(tmp, f"{unique}.pdf")
            jobs.append((unique, output, record))

        group_total = len(jobs)
        if task_id:
            _update_task(task_id, stage="rendering",
                         current_group=group_name,
                         group_total=group_total, group_progress=0,
                         message=f"rendering {group_name}: 0/{group_total}")

        args_list = [
            (template, record, output, manual_fields,
             base_html, static_dir, dpi, title)
            for (_, output, record) in jobs
        ]
        # Map each future back to the row it rendered so we can log the
        # history record once the future resolves (only successful rows
        # get logged — failed futures raise in fut.result()).
        fut_to_record = {
            pool.submit(_html_worker_job, a): record
            for a, (_, _, record) in zip(args_list, jobs)
        }
        futures = list(fut_to_record.keys())
        done = 0
        for fut in concurrent.futures.as_completed(futures):
            fut.result()  # propagate exceptions
            # Notice-history audit log — one row per successful render.
            record = fut_to_record[fut]
            merged = {**(manual_fields or {}), **record}
            _log_notice_record(
                name=str(merged.get("name", "") or ""),
                principal=str(merged.get("Principal_Amount", "") or ""),
            )
            done += 1
            if task_id and (done % 3 == 0 or done == group_total):
                _update_task(task_id,
                             progress=cumulative_before + done,
                             group_progress=done,
                             message=f"rendering {group_name}: {done}/{group_total}")
        if task_id:
            _update_task(task_id, stage="packing",
                         message=f"packing {group_name}.zip")

        safe = _safe_name(group_name)
        zip_path = os.path.join(
            UPLOAD_DIR, f"part_{uuid.uuid4().hex}_{safe}.zip")
        with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
            for (unique, output, _) in jobs:
                if os.path.exists(output):
                    zf.write(output, f"{unique}.pdf")
        return zip_path
    finally:
        shutil.rmtree(tmp, ignore_errors=True)


def generate_notices_html(template, data_rows, manual_fields=None,
                          filename_fields=None, group_by_field=None,
                          task_id=None, render_workers=2,
                          dpi=300, title=PDF_TITLE_DEFAULT,
                          profile_label=""):
    """Render every row into a locked non-extractable PDF and pack into
    per-group zips (or a single output.zip if no group_by_field)."""
    total = len(data_rows)

    if group_by_field:
        sorted_rows = sorted(
            data_rows, key=lambda r: _format_value(r.get(group_by_field, "")))
        buckets = defaultdict(list)
        for r in sorted_rows:
            g = _safe_name(r.get(group_by_field, "") or "unassigned")
            buckets[g].append(r)
        groups_ordered = list(buckets.items())
    else:
        groups_ordered = [("output", list(data_rows))]

    # Cache base HTML text once to avoid re-reading per row.
    with open(TEMPLATE_BASE_HTML_PATH, "r", encoding="utf-8") as f:
        base_html = f.read()

    start_msg = (f"Processing {len(groups_ordered)} group(s), {total} row(s) "
                 f"· profile={profile_label or '?'} · "
                 f"workers={render_workers} · dpi={dpi}")
    if task_id:
        _update_task(task_id, status="running", stage="queued",
                     total=total, progress=0,
                     groups_total=len(groups_ordered), groups_done=0,
                     ready_parts=[], current_group="",
                     group_total=0, group_progress=0, message=start_msg)

    # Try to spin up a ProcessPoolExecutor with per-process Chromium.
    pool = None
    try:
        pool = concurrent.futures.ProcessPoolExecutor(
            max_workers=render_workers,
            initializer=_html_worker_init)
    except Exception as e:
        print(f"[generate_notices_html] ProcessPool unavailable ({e}); "
              f"falling back to single-process serial rendering")

    cumulative = 0
    try:
        for idx, (group_name, rows) in enumerate(groups_ordered):
            if task_id:
                _update_task(task_id, current_group=group_name)

            if pool is not None:
                zip_path = _process_group_html(
                    template, rows, group_name, manual_fields,
                    filename_fields, task_id, cumulative, total,
                    pool, base_html, TEMPLATE_STATIC_DIR, dpi, title)
            else:
                # Serial fallback: render every row in the main process.
                zip_path = _process_group_html_serial(
                    template, rows, group_name, manual_fields,
                    filename_fields, task_id, cumulative, total,
                    base_html, TEMPLATE_STATIC_DIR, dpi, title)

            cumulative += len(rows)
            with TASKS_LOCK:
                t = TASKS.get(task_id) if task_id else None
                if t is not None:
                    parts = t.get("ready_parts") or []
                    parts.append({
                        "index": len(parts),
                        "name": f"{_safe_name(group_name)}.zip",
                        "group": group_name,
                        "path": zip_path,
                    })
                    t["ready_parts"] = parts
                    t["groups_done"] = idx + 1
                    t["progress"] = cumulative
    finally:
        if pool is not None:
            pool.shutdown(wait=True)

    if task_id:
        _update_task(task_id, status="done", stage="done",
                     progress=total,
                     message=f"All {len(groups_ordered)} group(s) ready")


def _process_group_html_serial(template, rows, group_name, manual_fields,
                               filename_fields, task_id, cumulative_before,
                               grand_total, base_html, static_dir,
                               dpi, title):
    """Same as _process_group_html but without a process pool — renders
    each row sequentially in-process. Used as fallback when
    multiprocessing can't be initialized (some VPS setups)."""
    tmp = tempfile.mkdtemp(prefix=f"notice_{_safe_name(group_name)[:32]}_")
    try:
        taken = set()
        jobs = []
        for i, record in enumerate(rows, start=1):
            merged = {**(manual_fields or {}), **record}
            base = _build_filename(merged, filename_fields, i)
            unique = base; n = 1
            while unique in taken:
                n += 1
                unique = f"{base}_{n}"
            taken.add(unique)
            output = os.path.join(tmp, f"{unique}.pdf")
            jobs.append((unique, output, record))

        group_total = len(jobs)
        if task_id:
            _update_task(task_id, stage="rendering",
                         current_group=group_name,
                         group_total=group_total, group_progress=0)

        # Keep one browser open across the whole group.
        pw = _sync_playwright().start()
        browser = pw.chromium.launch()
        try:
            for done, (unique, output, record) in enumerate(jobs, start=1):
                render_notice_row_pdf(
                    template, record, output,
                    manual_fields=manual_fields,
                    base_html=base_html, static_dir=static_dir,
                    browser=browser, rasterize=True, lock=True,
                    dpi=dpi, title=title)
                # Notice-history audit log — same schema as the parallel
                # batch path above and /generate_one.
                merged = {**(manual_fields or {}), **record}
                _log_notice_record(
                    name=str(merged.get("name", "") or ""),
                    principal=str(merged.get("Principal_Amount", "") or ""),
                )
                if task_id and (done % 3 == 0 or done == group_total):
                    _update_task(task_id,
                                 progress=cumulative_before + done,
                                 group_progress=done,
                                 message=f"rendering {group_name}: {done}/{group_total}")
        finally:
            browser.close()
            pw.stop()

        safe = _safe_name(group_name)
        zip_path = os.path.join(
            UPLOAD_DIR, f"part_{uuid.uuid4().hex}_{safe}.zip")
        with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
            for (unique, output, _) in jobs:
                if os.path.exists(output):
                    zf.write(output, f"{unique}.pdf")
        return zip_path
    finally:
        shutil.rmtree(tmp, ignore_errors=True)

# ── routes ───────────────────────────────────────────────────────

@app.before_request
def _require_auth():
    if request.path in _AUTH_EXEMPT_PATHS:
        return None
    # Fonts/images under /static are always public — they're referenced
    # from the preview iframe before any session cookie is available.
    if request.path.startswith("/static/"):
        return None
    if session.get("auth_ok") is True:
        return None
    # API endpoints: JSON 401 so the front-end can react without following
    # a redirect into HTML and breaking its fetch().then(r=>r.json()) chain.
    if request.path.startswith(_AUTH_API_PREFIXES):
        return jsonify(error="auth required"), 401
    # Everything else (HTML pages): bounce to the login form.
    return redirect("/login")


@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        supplied = (request.form.get("password") or "").encode("utf-8")
        expected = APP_PASSWORD.encode("utf-8")
        if _hmac.compare_digest(supplied, expected):
            session["auth_ok"] = True
            session.permanent = True
            return redirect("/")
        return Response(
            LOGIN_TEMPLATE.replace("__ERROR__",
                '<div class="error">Incorrect password.</div>'),
            mimetype="text/html", status=401)
    return Response(
        LOGIN_TEMPLATE.replace("__ERROR__", ""),
        mimetype="text/html")


@app.route("/logout", methods=["GET", "POST"])
def logout():
    session.pop("auth_ok", None)
    session.pop("sid", None)
    session.pop("tpl_path", None)
    session.pop("xls_path", None)
    return redirect("/login")


@app.route("/")
def index():
    # Bypass Jinja: the inline HTML template contains literal `{{...}}`
    # placeholder syntax inside JavaScript strings that would otherwise
    # collide with Jinja's expression delimiters.
    return Response(HTML_TEMPLATE, mimetype="text/html")


@app.route("/verify")
def verify_page():
    """Notice-verification UI: enter a printed 16-char serial, get
    back the original (name, principal, generated_at) the firm
    issued the letter with. Also shows recent history."""
    return Response(VERIFY_TEMPLATE, mimetype="text/html")


@app.route("/inventory")
def inventory_page():
    """Inventory mode: bulk-import a case Excel once, then pick
    subsets (by cnic / name / phone) to generate notices on demand.
    `{{date}}` defaults to today, so the sidebar just needs a
    date override field instead of the full manual placeholder form."""
    return Response(INVENTORY_TEMPLATE, mimetype="text/html")


@app.route("/api/templates", methods=["GET"])
def api_list_templates():
    return jsonify(list_templates())


@app.route("/upload", methods=["POST"])
def upload():
    """Accept an Excel file; analyze its headers against the chosen
    template's placeholders. Stashes Excel path + template slug in
    session so /generate can pick them up later."""
    excel_file = request.files.get("excel")
    template_slug = (request.form.get("template_slug") or "default").strip()
    if not _valid_slug(template_slug):
        return jsonify(error="invalid template_slug"), 400

    if not excel_file:
        return jsonify(error="Please upload an Excel file."), 400

    template = load_template(template_slug)
    if template is None:
        return jsonify(error=f"Template '{template_slug}' not found."), 400

    sid = uuid.uuid4().hex
    work_dir = os.path.join(UPLOAD_DIR, sid)
    os.makedirs(work_dir, exist_ok=True)
    xls_path = os.path.join(work_dir, "data.xlsx")
    excel_file.save(xls_path)

    try:
        placeholders = extract_template_placeholders(template)
        headers, data_rows = read_excel(xls_path)
    except Exception as e:
        shutil.rmtree(work_dir, ignore_errors=True)
        return jsonify(error=f"Failed to read Excel: {e}"), 400

    headers_clean = [h for h in headers if h]
    matched = [p for p in placeholders if p in set(headers_clean)]
    missing = [p for p in placeholders if p not in set(headers_clean)]

    session["sid"] = sid
    session["xls_path"] = xls_path
    session["template_slug"] = template_slug

    return jsonify(
        placeholders=placeholders,
        excel_headers=headers_clean,
        matched=matched,
        missing=missing,
        row_count=len(data_rows),
        preview=data_rows[:5],
        template_name=template.get("name"),
        template_slug=template_slug,
    )


@app.route("/generate", methods=["POST"])
def generate():
    sid = session.get("sid")
    xls_path = session.get("xls_path")
    template_slug = session.get("template_slug", "default")

    if not sid or not xls_path or not os.path.exists(xls_path):
        return jsonify(error="Please upload an Excel file first."), 400

    template = load_template(template_slug)
    if template is None:
        return jsonify(error=f"Template '{template_slug}' not found."), 400

    payload = request.get_json() or {}
    manual_fields = payload.get("manual_fields", {}) or {}
    filename_fields = payload.get("filename_fields") or []
    group_by_field = payload.get("group_by_field") or None

    profile = _get_machine_profile(payload.get("machine"))

    _, data_rows = read_excel(xls_path)
    if not data_rows:
        return jsonify(error="No data rows found in Excel."), 400

    # Date picker override — webpage sends ISO YYYY-MM-DD, we want DD/MM/YYYY.
    date_value = (payload.get("date_value") or "").strip()
    if date_value:
        try:
            d = datetime.date.fromisoformat(date_value)
            formatted = d.strftime("%d/%m/%Y")
            for r in data_rows:
                r["date"] = formatted
        except ValueError:
            pass

    task_id = _new_task()

    def worker():
        try:
            generate_notices_html(
                template, data_rows,
                manual_fields=manual_fields,
                filename_fields=filename_fields,
                group_by_field=group_by_field,
                task_id=task_id,
                render_workers=profile["render_workers"],
                profile_label=profile["label"],
            )
        except Exception as e:
            traceback.print_exc()
            _update_task(task_id, status="error", stage="error",
                         error=f"{e.__class__.__name__}: {e}",
                         message=f"Generation failed: {e}")
        finally:
            shutil.rmtree(os.path.join(UPLOAD_DIR, sid), ignore_errors=True)

    threading.Thread(target=worker, daemon=True).start()
    return jsonify(task_id=task_id)


@app.route("/status/<task_id>")
def status(task_id):
    task = _get_task(task_id)
    if task is None:
        return jsonify(error="Unknown task."), 404
    # Strip disk paths from the public response.
    parts_public = [
        {"index": p["index"], "name": p["name"], "group": p.get("group", "")}
        for p in (task.get("ready_parts") or [])
    ]
    return jsonify(
        status=task["status"],
        stage=task["stage"],
        progress=task["progress"],
        total=task["total"],
        message=task["message"],
        error=task["error"],
        groups_total=task.get("groups_total", 0),
        groups_done=task.get("groups_done", 0),
        current_group=task.get("current_group", ""),
        group_total=task.get("group_total", 0),
        group_progress=task.get("group_progress", 0),
        ready_parts=parts_public,
    )


@app.route("/download/<task_id>/<int:part_index>")
def download_part(task_id, part_index):
    task = _get_task(task_id)
    if task is None:
        return jsonify(error="Unknown task."), 404
    parts = task.get("ready_parts") or []
    if part_index < 0 or part_index >= len(parts):
        return jsonify(error="Part not ready."), 404

    part = parts[part_index]
    zip_path = part.get("path")
    if not zip_path or not os.path.exists(zip_path):
        return jsonify(error="Part file missing."), 404

    # Clean up this part's zip after a generous delay so retries work.
    # Persistent parts (inventory-mode batch zips) are kept on disk so
    # the user can re-download days later; the batch dir holds history.
    if not part.get("persistent"):
        def delayed_cleanup():
            try:
                if os.path.exists(zip_path):
                    os.remove(zip_path)
            except OSError:
                pass
        threading.Timer(600.0, delayed_cleanup).start()

    return send_file(zip_path, mimetype="application/zip",
                     as_attachment=True, download_name=part["name"])


# ── template editor / preview (task 15) ─────────────────────────

PREVIEW_SAMPLE_DATA = {
    "date": "15/04/2026", "name": "Ali Hassan",
    "cnic": "4210112345678", "phone": "0300-1234567",
    "disb_date": "01/12/2025", "Due_date": "15/02/2026",
    "Principal_Amount": "50,000.00", "Interest": "2,500.00",
    "Penalty": "500.00", "Payable": "53,000.00",
    "Transaction_id": "TX100001", "easypaisa_account": "EP-0001",
}


def _build_preview_html(sample=None, security_config=None, template=None,
                        fill_placeholders=False, strip_unfilled=None):
    """Render the base template with the given template's blocks, assets,
    and security overlay. Preview matches the final PDF because we reuse
    build_notice_html — the same function the rendering pipeline uses.

    `fill_placeholders` defaults to False so the iframe preview shows the
    bare template with literal `{{placeholder}}` tokens. Set True to get
    a sample-data or manual-form partial-fill render.
    `strip_unfilled` defaults to `fill_placeholders`; pass False for a
    partial fill that keeps un-typed `{{foo}}` visible."""
    sample = sample if sample is not None else PREVIEW_SAMPLE_DATA
    if template is None:
        template = load_template("default")
    else:
        template = dict(template)  # don't mutate caller's dict
    if security_config is not None:
        template["security"] = security_config
    return build_notice_html(template, sample, base_html=None,
                             static_dir=TEMPLATE_STATIC_DIR,
                             manual_fields=None,
                             fill_placeholders=fill_placeholders,
                             strip_unfilled=strip_unfilled)


@app.route("/preview.html")
def preview_html():
    """Live preview iframe (GET). Takes ?slug=<template_slug> (defaults
    to 'default') and renders the template WITHOUT filling placeholders
    — you see literal `{{name}}` etc. so the iframe shows the template,
    not a sample. Pass ?fill=1 to substitute sample data instead."""
    slug = (request.args.get("slug") or "default").strip()
    if not _valid_slug(slug):
        return jsonify(error="invalid slug"), 400
    fill = request.args.get("fill", "0").lower() in ("1", "true", "yes")
    tpl = load_template(slug) or load_template("default")
    cfg = tpl.get("security") or load_default_security()
    html = _build_preview_html(
        security_config=cfg, template=tpl, fill_placeholders=fill)
    resp = Response(html, mimetype="text/html")
    resp.headers["Cache-Control"] = "no-store"
    return resp


@app.route("/api/preview_html", methods=["POST"])
def api_preview_html():
    """Live-preview endpoint for unsaved sidebar state.

    Body JSON: {slug, blocks?, security?, assets_config?, fields?}.
      - `slug` picks the on-disk template as base (for uploaded PNGs).
      - `blocks` / `security` / `assets_config` override currently
        persisted values for the in-memory preview.
      - `fields` is an optional dict of placeholder values from the
        Manual-single form. When present, the preview switches to
        partial-fill mode (substitute those keys, leave others as
        literal `{{placeholders}}`)."""
    payload = request.get_json(silent=True) or {}
    slug = (payload.get("slug") or "default").strip()
    if not _valid_slug(slug):
        return jsonify(error="invalid slug"), 400
    tpl = load_template(slug) or load_template("default")
    if payload.get("blocks") is not None:
        incoming = payload["blocks"]
        merged_blocks = dict(tpl.get("blocks") or {})
        for p in EDITABLE_PURPOSES:
            if p in incoming:
                merged_blocks[p] = incoming[p]
        tpl["blocks"] = merged_blocks
    if payload.get("security") is not None:
        tpl["security"] = _merge_security_config(payload["security"])
    if payload.get("assets_config") is not None:
        tpl["assets_config"] = _merge_assets_config(payload["assets_config"])

    fields = payload.get("fields")
    if fields:
        # Partial-fill: substitute provided keys, keep others literal.
        html = _build_preview_html(
            sample=fields, template=tpl,
            fill_placeholders=True, strip_unfilled=False)
    else:
        # Bare template preview — keep all {{...}} literal.
        html = _build_preview_html(template=tpl, fill_placeholders=False)
    resp = Response(html, mimetype="text/html")
    resp.headers["Cache-Control"] = "no-store"
    return resp


@app.route("/editor", methods=["GET"])
def editor():
    """Deprecated — the old standalone editor. Redirect to the unified
    workstation, which now contains all editing controls."""
    return redirect("/")


@app.route("/api/security", methods=["GET", "POST"])
def api_security():
    """Shortcut for default-template security (back-compat)."""
    if request.method == "GET":
        return jsonify(load_default_security())
    payload = request.get_json(silent=True) or {}
    merged = save_default_security(payload)
    save_template("S&S Law Firm — Default", security=merged,
                  overwrite_slug="default")
    return jsonify(ok=True, config=merged)


# ── template CRUD API (task 17) ─────────────────────────────────

@app.route("/api/templates/<slug>", methods=["GET"])
def api_template_get(slug):
    if not _valid_slug(slug):
        return jsonify(error="invalid slug"), 400
    tpl = load_template(slug)
    if tpl is None:
        return jsonify(error=f"template '{slug}' not found"), 404
    return jsonify(_template_public(tpl))


@app.route("/api/templates/<slug>", methods=["PUT"])
def api_template_put(slug):
    """Update blocks / security / assets_config / name of an existing
    template (or the built-in default). Body JSON: {name?, blocks?,
    security?, assets_config?}."""
    if not _valid_slug(slug):
        return jsonify(error="invalid slug"), 400
    tpl = load_template(slug)
    if tpl is None:
        return jsonify(error=f"template '{slug}' not found"), 404
    payload = request.get_json(silent=True) or {}
    name = payload.get("name") or tpl.get("name")
    blocks = payload.get("blocks")
    security = payload.get("security")
    assets_config = payload.get("assets_config")
    try:
        save_template(name, blocks=blocks, security=security,
                      assets_config=assets_config, overwrite_slug=slug)
    except ValueError as e:
        return jsonify(error=str(e)), 400
    return jsonify(ok=True, template=_template_public(load_template(slug)))


@app.route("/api/templates", methods=["POST"])
def api_template_create():
    """Create a new template. Body JSON: {name, base_slug?, blocks?,
    security?, assets_config?}. If base_slug is given, the new template
    starts as a copy of that one (including its assets)."""
    payload = request.get_json(silent=True) or {}
    name = (payload.get("name") or "").strip()
    if not name:
        return jsonify(error="name is required"), 400
    base_slug = (payload.get("base_slug") or "default").strip()
    if not _valid_slug(base_slug):
        return jsonify(error="invalid base_slug"), 400
    base = load_template(base_slug)
    if base is None:
        return jsonify(error=f"base template '{base_slug}' not found"), 400

    # Log incoming clone payload so we can diagnose "edits didn't save"
    # reports. Blocks are truncated to keep the log readable.
    pb = payload.get("blocks") or {}
    print(f"[clone] name={name!r} base={base_slug!r} "
          f"blocks_keys={sorted(pb.keys())} "
          f"sample notice-subject={str(pb.get('notice-subject',''))[:40]!r}")

    blocks = payload.get("blocks") or base.get("blocks")
    security = payload.get("security") or base.get("security")
    assets_config = payload.get("assets_config") or base.get("assets_config")
    try:
        new_slug = save_template(name, blocks=blocks, security=security,
                                 assets_config=assets_config)
    except ValueError as e:
        return jsonify(error=str(e)), 400

    # Copy assets from base to new template.
    for kind, src in (base.get("assets") or {}).items():
        if not src or not os.path.isfile(src):
            continue
        dst_dir = os.path.join(TEMPLATES_ROOT, new_slug, "assets")
        os.makedirs(dst_dir, exist_ok=True)
        shutil.copy(src, os.path.join(dst_dir, f"{kind}.png"))

    return jsonify(ok=True,
                   template=_template_public(load_template(new_slug)))


@app.route("/api/templates/<slug>", methods=["DELETE"])
def api_template_delete(slug):
    if not _valid_slug(slug):
        return jsonify(error="invalid slug"), 400
    try:
        ok = delete_template(slug)
    except ValueError as e:
        return jsonify(error=str(e)), 400
    return jsonify(ok=ok)


@app.route("/api/templates/<slug>/assets/<kind>", methods=["POST", "DELETE"])
def api_template_asset(slug, kind):
    if not _valid_slug(slug):
        return jsonify(error="invalid slug"), 400
    if kind not in ASSET_KINDS:
        return jsonify(error=f"unknown asset kind: {kind}"), 400
    if load_template(slug) is None:
        return jsonify(error=f"template '{slug}' not found"), 404
    if request.method == "DELETE":
        delete_template_asset(slug, kind)
        return jsonify(ok=True)
    f = request.files.get("file")
    if not f:
        return jsonify(error="no file uploaded (expected field 'file')"), 400
    # Bump meta.updated so the frontend can invalidate caches.
    try:
        save_template_asset(slug, kind, f)
    except ValueError as e:
        return jsonify(error=str(e)), 400
    save_template(load_template(slug).get("name"), overwrite_slug=slug)
    return jsonify(ok=True, kind=kind)


def _template_public(tpl):
    """Strip filesystem paths from a template's assets dict before
    sending to the client. The UI only needs to know WHICH slots have
    overrides — it renders images via /api/templates/<slug>/assets/<kind>."""
    if tpl is None:
        return None
    return {
        "slug": tpl["slug"],
        "name": tpl["name"],
        "builtin": tpl.get("builtin", False),
        "blocks": tpl.get("blocks") or {},
        "security": tpl.get("security") or {},
        "assets": {k: bool(tpl.get("assets", {}).get(k)) for k in ASSET_KINDS},
        "assets_config": tpl.get("assets_config") or DEFAULT_ASSETS_CONFIG,
    }


@app.route("/api/templates/<slug>/assets/<kind>", methods=["GET"])
def api_template_asset_image(slug, kind):
    """Serve a template's asset image (with fallback to the default's
    shipped asset if this template hasn't overridden it)."""
    if not _valid_slug(slug):
        return jsonify(error="invalid slug"), 400
    if kind not in ASSET_KINDS:
        return jsonify(error="unknown kind"), 400
    # per-template override first
    override = os.path.join(TEMPLATES_ROOT, slug, "assets", f"{kind}.png")
    if os.path.isfile(override):
        return send_file(override, mimetype="image/png")
    # fallback: shipped default
    defaults = {
        "logo": "logo.png",
        "seal": "seal_default.png",
        "signature_seal": "signature_seal_default.png",
    }
    shipped = os.path.join(TEMPLATE_STATIC_DIR, "images", defaults[kind])
    if os.path.isfile(shipped):
        return send_file(shipped, mimetype="image/png")
    return jsonify(error="asset not found"), 404


@app.route("/api/templates/<slug>/placeholders")
def api_template_placeholders(slug):
    """List every {{placeholder}} referenced by the base HTML + the
    template's block texts. Used by the manual-single-notice form
    to render one input per placeholder.

    Also flags which placeholders are MONEY fields (header matches
    MONEY_KEYWORDS) so the frontend can attach a blur-time auto-format
    handler that turns `50000` into `50,000.00`."""
    if not _valid_slug(slug):
        return jsonify(error="invalid slug"), 400
    tpl = load_template(slug)
    if tpl is None:
        return jsonify(error=f"template '{slug}' not found"), 404
    placeholders = extract_template_placeholders(tpl)
    money = [p for p in placeholders if _is_money_header(p)]
    return jsonify(placeholders=placeholders, money=money)


@app.route("/api/verify")
def api_verify():
    """Reverse-lookup a printed 16-char serial (`XXXX-XXXX-XXXX-XXXX`)
    into the history record: `{name, principal, generated_at}`. Answers
    'was this letter actually issued, and to whom?' without trusting the
    paper."""
    serial = (request.args.get("serial") or "").strip().upper()
    # Accept either with dashes or without; normalize.
    serial = re.sub(r"\s+", "", serial)
    if not re.fullmatch(r"[A-Z0-9-]{1,40}", serial):
        return jsonify(ok=False, error="invalid serial format"), 400
    row = lookup_notice_by_serial(serial)
    if row is None:
        return jsonify(ok=False, error="not found"), 404
    return jsonify(ok=True, **row)


@app.route("/api/history")
def api_history():
    """Paginated history list (most recent first). Query params:
      ?limit=N (default 50, max 500)
      ?offset=N (default 0)
      ?q=text (optional substring match on name)"""
    try:
        limit = min(500, max(1, int(request.args.get("limit", 50))))
        offset = max(0, int(request.args.get("offset", 0)))
    except ValueError:
        return jsonify(error="limit/offset must be integers"), 400
    q = (request.args.get("q") or "").strip()
    with sqlite3.connect(HISTORY_DB_PATH) as conn:
        conn.row_factory = sqlite3.Row
        if q:
            total = conn.execute(
                "SELECT COUNT(*) FROM notice_history WHERE name LIKE ?",
                (f"%{q}%",)).fetchone()[0]
            rows = conn.execute(
                "SELECT serial, name, principal, generated_at "
                "FROM notice_history WHERE name LIKE ? "
                "ORDER BY id DESC LIMIT ? OFFSET ?",
                (f"%{q}%", limit, offset)).fetchall()
        else:
            total = conn.execute(
                "SELECT COUNT(*) FROM notice_history").fetchone()[0]
            rows = conn.execute(
                "SELECT serial, name, principal, generated_at "
                "FROM notice_history ORDER BY id DESC LIMIT ? OFFSET ?",
                (limit, offset)).fetchall()
    return jsonify(
        total=total, limit=limit, offset=offset,
        rows=[dict(r) for r in rows])


# ── inventory mode: /api/cases/* ──────────────────────────────────
#
# These endpoints back the /inventory page. Flow:
#   1. Import Excel → preview (scan for 订单编号 conflicts)
#   2. Commit with policy=skip|overwrite → async task
#   3. Search by cnic/name/phone, paginate, accumulate selected order_ids
#   4. Generate notices for the selection → feeds the existing batch
#      render pipeline (generate_notices_html) unchanged.
#
# /status/<task_id> and /download/<task_id>/<part_index> are shared with
# the Excel-batch flow — the TASKS dict is generic, import tasks just
# never populate ready_parts.

_CASES_TOKEN_RE = re.compile(r"[0-9a-f]{32}")


@app.route("/api/cases/stats", methods=["GET"])
def api_cases_stats():
    return jsonify(count=_cases_count())


@app.route("/api/cases/import/preview", methods=["POST"])
def api_cases_import_preview():
    """Stage 1 of import. Accepts the Excel, stashes it on disk,
    streams through once to count rows + collect 订单编号 conflicts.
    Returns a token the client passes back to /commit. No DB writes yet.

    For a 40-万-row sheet this is the slow step (~20-40 s); keep it
    synchronous because the preview numbers drive the confirm dialog."""
    excel_file = request.files.get("excel")
    if not excel_file:
        return jsonify(error="Please upload an Excel file."), 400
    token = uuid.uuid4().hex
    stash_path = os.path.join(CASES_IMPORTS_DIR, f"{token}.xlsx")
    excel_file.save(stash_path)
    try:
        wb = openpyxl.load_workbook(stash_path, read_only=True, data_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header_row = next(rows_iter)
        except StopIteration:
            wb.close()
            try: os.remove(stash_path)
            except OSError: pass
            return jsonify(error="Excel is empty."), 400
        headers = [str(h).strip() if h is not None else "" for h in header_row]
        oid_col = _find_col(headers, CASE_COL_ALIASES["order_id"])
        if not oid_col:
            wb.close()
            try: os.remove(stash_path)
            except OSError: pass
            return jsonify(
                error=("No order-id column found. Expected one of: "
                       + ", ".join(CASE_COL_ALIASES["order_id"]))), 400
        oid_idx = headers.index(oid_col)

        # Stream order_ids only — we don't need the full row dict here.
        all_ids = []
        seen = set()
        dup_in_file = set()
        for row in rows_iter:
            if oid_idx >= len(row):
                continue
            v = _normalize_order_id(row[oid_idx])
            if not v:
                continue
            if v in seen:
                dup_in_file.add(v)
            else:
                seen.add(v)
                all_ids.append(v)
        wb.close()

        existing = _cases_existing_ids(all_ids)
        return jsonify(
            token=token,
            total=len(all_ids),
            new_count=len(all_ids) - len(existing),
            conflict_count=len(existing),
            conflicts_sample=list(existing)[:100],
            dup_in_file_count=len(dup_in_file),
            dup_in_file_sample=list(dup_in_file)[:20],
            headers=[h for h in headers if h],
            order_id_col=oid_col,
            cnic_col=_find_col(headers, CASE_COL_ALIASES["cnic"]),
            name_col=_find_col(headers, CASE_COL_ALIASES["name"]),
            phone_col=_find_col(headers, CASE_COL_ALIASES["phone"]),
        )
    except Exception as e:
        try: os.remove(stash_path)
        except OSError: pass
        traceback.print_exc()
        return jsonify(error=f"Preview failed: {e}"), 400


def _cases_import_worker(task_id, stash_path, policy):
    """Async committer. Reads the stashed Excel and upserts rows per
    policy; updates TASKS[task_id] progress as it goes."""
    try:
        _update_task(task_id, status="running", stage="reading",
                     message="Reading Excel…")
        headers, data_rows = read_excel(stash_path)
        oid_col  = _find_col(headers, CASE_COL_ALIASES["order_id"])
        cnic_col = _find_col(headers, CASE_COL_ALIASES["cnic"])
        name_col = _find_col(headers, CASE_COL_ALIASES["name"])
        phone_col = _find_col(headers, CASE_COL_ALIASES["phone"])
        if not oid_col:
            raise ValueError("order-id column missing (file changed?)")
        total = len(data_rows)
        _update_task(task_id, stage="inserting",
                     total=total, progress=0,
                     message=f"Importing {total} rows…")
        now = datetime.datetime.now().isoformat(timespec="seconds")
        inserted = updated = skipped_existing = skipped_blank = dups_in_file = 0
        BATCH = 1000
        seen_this_run = set()

        # One long-lived writer holds _CASES_LOCK; WAL lets concurrent
        # readers (search endpoints) keep working.
        with _CASES_LOCK, sqlite3.connect(CASES_DB_PATH) as conn:
            conn.execute("BEGIN")
            try:
                for i, row in enumerate(data_rows, 1):
                    oid = _normalize_order_id(row.get(oid_col))
                    if not oid:
                        skipped_blank += 1
                        continue
                    if oid in seen_this_run:
                        dups_in_file += 1
                        # In-file duplicates: let the last occurrence win
                        # under overwrite; drop duplicates under skip.
                        if policy != "overwrite":
                            continue
                    seen_this_run.add(oid)

                    # Normalize every cell so numbers match what the
                    # existing batch pipeline would produce on its own
                    # read — JSON round-trip stays value-preserving for
                    # strings / ints / floats.
                    row_out = {k: v for k, v in row.items() if k}
                    payload = json.dumps(row_out, ensure_ascii=False,
                                         default=str)
                    cnic_v  = str(row.get(cnic_col,  "") or "") if cnic_col  else ""
                    name_v  = str(row.get(name_col,  "") or "") if name_col  else ""
                    phone_v = str(row.get(phone_col, "") or "") if phone_col else ""

                    if policy == "overwrite":
                        # Overwrite explicitly replaces the row's view of
                        # "truth", so original_name resets to the newly
                        # imported name — later inline edits are still
                        # reversible back to "what this import said".
                        cur = conn.execute(
                            "UPDATE cases SET cnic=?, name=?, "
                            "original_name=?, phone=?, row_json=?, "
                            "updated_at=? WHERE order_id=?",
                            (cnic_v, name_v, name_v, phone_v, payload,
                             now, oid))
                        if cur.rowcount == 0:
                            conn.execute(
                                "INSERT INTO cases(order_id, cnic, name, "
                                "original_name, phone, row_json, "
                                "created_at, updated_at) "
                                "VALUES (?,?,?,?,?,?,?,?)",
                                (oid, cnic_v, name_v, name_v, phone_v,
                                 payload, now, now))
                            inserted += 1
                        else:
                            updated += 1
                    else:  # skip
                        cur = conn.execute(
                            "INSERT OR IGNORE INTO cases(order_id, cnic, "
                            "name, original_name, phone, row_json, "
                            "created_at, updated_at) "
                            "VALUES (?,?,?,?,?,?,?,?)",
                            (oid, cnic_v, name_v, name_v, phone_v,
                             payload, now, now))
                        if cur.rowcount == 1:
                            inserted += 1
                        else:
                            skipped_existing += 1

                    if i % BATCH == 0:
                        conn.commit()
                        conn.execute("BEGIN")
                        _update_task(task_id, progress=i,
                                     message=f"Importing… {i}/{total}")
                conn.commit()
            except Exception:
                conn.execute("ROLLBACK")
                raise

        summary = (f"inserted={inserted}, updated={updated}, "
                   f"skipped_existing={skipped_existing}, "
                   f"skipped_blank={skipped_blank}, "
                   f"dups_in_file={dups_in_file}")
        _update_task(task_id, status="done", stage="done",
                     progress=total,
                     message="Import done. " + summary,
                     cases_inserted=inserted,
                     cases_updated=updated,
                     cases_skipped_existing=skipped_existing,
                     cases_skipped_blank=skipped_blank,
                     cases_dups_in_file=dups_in_file)
    except Exception as e:
        traceback.print_exc()
        _update_task(task_id, status="error", stage="error",
                     error=f"{e.__class__.__name__}: {e}",
                     message=f"Import failed: {e}")
    finally:
        try: os.remove(stash_path)
        except OSError: pass


@app.route("/api/cases/import/commit", methods=["POST"])
def api_cases_import_commit():
    payload = request.get_json() or {}
    token = (payload.get("token") or "").strip()
    policy = (payload.get("policy") or "skip").strip().lower()
    if policy not in ("skip", "overwrite"):
        return jsonify(error="policy must be 'skip' or 'overwrite'"), 400
    if not _CASES_TOKEN_RE.fullmatch(token):
        return jsonify(error="invalid token"), 400
    stash_path = os.path.join(CASES_IMPORTS_DIR, f"{token}.xlsx")
    if not os.path.exists(stash_path):
        return jsonify(error="import session expired — re-upload."), 400
    task_id = _new_task()
    threading.Thread(
        target=_cases_import_worker,
        args=(task_id, stash_path, policy),
        daemon=True,
    ).start()
    return jsonify(task_id=task_id)


@app.route("/api/cases/search", methods=["GET", "POST"])
def api_cases_search():
    """Paginated listing. Three input modes on the same endpoint:
    - empty `q` → full listing
    - single-token `q` → prefix LIKE on the indexed column (fast)
    - multi-token `q` (comma / whitespace / newline-separated) →
      exact-match IN via a TEMP table join (bypasses the 999-bind cap)

    POST with JSON `{field, q, page, limit}` is accepted for inputs
    too long to fit in a URL (an Excel column of 20 k CNICs pasted
    into the search box)."""
    if request.method == "POST":
        payload = request.get_json(silent=True) or {}
        field = (payload.get("field") or "cnic").strip().lower()
        q_raw = payload.get("q") or ""
        try:
            page = max(1, int(payload.get("page") or 1))
            limit = min(500, max(1, int(payload.get("limit") or 100)))
        except (TypeError, ValueError):
            return jsonify(error="bad page/limit"), 400
    else:
        field = (request.args.get("field") or "cnic").strip().lower()
        q_raw = request.args.get("q") or ""
        try:
            page = max(1, int(request.args.get("page") or 1))
            limit = min(500, max(1, int(request.args.get("limit") or 100)))
        except ValueError:
            return jsonify(error="bad page/limit"), 400
    if field not in ("cnic", "name", "phone"):
        return jsonify(error="field must be cnic / name / phone"), 400

    values = _parse_multi_values(q_raw)
    offset = (page - 1) * limit

    with sqlite3.connect(CASES_DB_PATH) as conn:
        conn.row_factory = sqlite3.Row
        if not values:
            total = conn.execute("SELECT COUNT(*) FROM cases").fetchone()[0]
            rows = conn.execute(
                "SELECT order_id, cnic, name, original_name, phone, created_at "
                "FROM cases ORDER BY order_id LIMIT ? OFFSET ?",
                (limit, offset)).fetchall()
            mode = "all"
        elif len(values) == 1:
            v = values[0]
            total = conn.execute(
                f"SELECT COUNT(*) FROM cases WHERE {field} LIKE ?",
                (v + "%",)).fetchone()[0]
            rows = conn.execute(
                f"SELECT order_id, cnic, name, original_name, phone, created_at "
                f"FROM cases WHERE {field} LIKE ? "
                f"ORDER BY order_id LIMIT ? OFFSET ?",
                (v + "%", limit, offset)).fetchall()
            mode = "prefix"
        else:
            _cases_temp_values_table(conn, values)
            total = conn.execute(
                f"SELECT COUNT(DISTINCT c.order_id) FROM cases c "
                f"JOIN _srch_q q ON c.{field} = q.v").fetchone()[0]
            rows = conn.execute(
                f"SELECT DISTINCT c.order_id, c.cnic, c.name, "
                f"c.original_name, c.phone, c.created_at FROM cases c "
                f"JOIN _srch_q q ON c.{field} = q.v "
                f"ORDER BY c.order_id LIMIT ? OFFSET ?",
                (limit, offset)).fetchall()
            mode = "exact"
    return jsonify(
        total=total, page=page, limit=limit,
        mode=mode, values_parsed=len(values),
        rows=[dict(r) for r in rows],
    )


@app.route("/api/cases/search/ids", methods=["GET", "POST"])
def api_cases_search_ids():
    """Return every order_id matching the filter, for 'select all
    matching' in the UI. Same three modes as /api/cases/search.
    Capped so the client can't force a multi-million-id payload;
    narrow the filter or paginate if you need the rest."""
    if request.method == "POST":
        payload = request.get_json(silent=True) or {}
        field = (payload.get("field") or "cnic").strip().lower()
        q_raw = payload.get("q") or ""
        try:
            cap = min(100_000, max(1, int(payload.get("cap") or 50_000)))
        except (TypeError, ValueError):
            return jsonify(error="bad cap"), 400
    else:
        field = (request.args.get("field") or "cnic").strip().lower()
        q_raw = request.args.get("q") or ""
        try:
            cap = min(100_000, max(1, int(request.args.get("cap") or 50_000)))
        except ValueError:
            return jsonify(error="bad cap"), 400
    if field not in ("cnic", "name", "phone"):
        return jsonify(error="field must be cnic / name / phone"), 400

    values = _parse_multi_values(q_raw)
    with sqlite3.connect(CASES_DB_PATH) as conn:
        if not values:
            total = conn.execute("SELECT COUNT(*) FROM cases").fetchone()[0]
            rows = conn.execute(
                "SELECT order_id FROM cases ORDER BY order_id LIMIT ?",
                (cap,)).fetchall()
        elif len(values) == 1:
            v = values[0]
            total = conn.execute(
                f"SELECT COUNT(*) FROM cases WHERE {field} LIKE ?",
                (v + "%",)).fetchone()[0]
            rows = conn.execute(
                f"SELECT order_id FROM cases WHERE {field} LIKE ? "
                f"ORDER BY order_id LIMIT ?", (v + "%", cap)).fetchall()
        else:
            _cases_temp_values_table(conn, values)
            total = conn.execute(
                f"SELECT COUNT(DISTINCT c.order_id) FROM cases c "
                f"JOIN _srch_q q ON c.{field} = q.v").fetchone()[0]
            rows = conn.execute(
                f"SELECT DISTINCT c.order_id FROM cases c "
                f"JOIN _srch_q q ON c.{field} = q.v "
                f"ORDER BY c.order_id LIMIT ?", (cap,)).fetchall()
    return jsonify(total=total, returned=len(rows),
                   capped=(total > len(rows)),
                   ids=[r[0] for r in rows])


@app.route("/api/cases/by_ids", methods=["POST"])
def api_cases_by_ids():
    """Fetch display metadata (order_id / name / cnic / phone) for a
    list of order_ids. Feeds the "Selected cases" panel so rows added
    via 'select all matching' or bulk-match (which don't flow through
    the paginated table) still render name / cnic / phone."""
    payload = request.get_json(silent=True) or {}
    raw = payload.get("order_ids") or []
    if not isinstance(raw, list) or not raw:
        return jsonify(rows=[])
    ids = []
    seen = set()
    for x in raw:
        s = str(x).strip()
        if s and s not in seen:
            seen.add(s)
            ids.append(s)
        if len(ids) >= 50_000:
            break
    out = []
    with sqlite3.connect(CASES_DB_PATH) as conn:
        conn.row_factory = sqlite3.Row
        for i in range(0, len(ids), 500):
            chunk = ids[i:i + 500]
            qs = ",".join("?" * len(chunk))
            for r in conn.execute(
                    f"SELECT order_id, name, original_name, cnic, phone "
                    f"FROM cases WHERE order_id IN ({qs})",
                    tuple(chunk)):
                out.append(dict(r))
    return jsonify(rows=out)


@app.route("/api/cases/update_name", methods=["POST"])
def api_cases_update_name():
    """Inline-edit the `name` of a single case from the search-result
    table. Updates BOTH the indexed `cases.name` column (so future
    searches/bulk_match find the corrected spelling) AND the `name`
    key inside `row_json` (so the render pipeline pulls the new
    value at generate time). Other fields are untouched.

    Body JSON: { order_id, name }"""
    payload = request.get_json(silent=True) or {}
    oid = str(payload.get("order_id") or "").strip()
    new_name = str(payload.get("name") or "").strip()
    if not oid:
        return jsonify(error="order_id required"), 400
    if len(new_name) > 200:
        return jsonify(error="name too long (max 200 chars)"), 400
    now = datetime.datetime.now().isoformat(timespec="seconds")
    with _CASES_LOCK, sqlite3.connect(CASES_DB_PATH) as conn:
        conn.row_factory = sqlite3.Row
        row = conn.execute(
            "SELECT row_json FROM cases WHERE order_id = ?",
            (oid,)).fetchone()
        if row is None:
            return jsonify(error="not found"), 404
        try:
            data = json.loads(row["row_json"])
        except ValueError:
            data = {}
        data["name"] = new_name
        conn.execute(
            "UPDATE cases SET name = ?, row_json = ?, updated_at = ? "
            "WHERE order_id = ?",
            (new_name, json.dumps(data, ensure_ascii=False), now, oid))
        conn.commit()
    return jsonify(ok=True, order_id=oid, name=new_name, updated_at=now)


@app.route("/api/cases/bulk_match", methods=["POST"])
def api_cases_bulk_match():
    """Bulk-match an explicit list of cnic / name / phone values and
    return the order_ids that exist in the inventory — the "upload a
    spreadsheet of CNICs and pick the ones we have" case.

    Two input shapes share this endpoint:

    - multipart/form-data: { field, excel: <file> }
        The Excel's first header row is scanned for a column matching
        the chosen field (via CASE_COL_ALIASES); if none match, the
        first column is used. Every cell in that column becomes a
        match value. Caps at 100 k values.

    - application/json: { field, values: [...] }
        For typed / pasted input that's too long to fit the search
        box. Same 100 k cap.

    Response: { field, total_values, matched_count, missing_count,
                missing_sample:[...], matched_ids:[...],
                matched_meta:[{order_id,name,cnic,phone}, ...up to 200] }
    `matched_meta` lets the client show a preview table without a
    follow-up /by_ids round-trip; pull /by_ids if it needs more.
    """
    field = None
    values = []
    src_col = None
    is_mp = (request.content_type or "").startswith("multipart/form-data")
    if is_mp:
        field = (request.form.get("field") or "cnic").strip().lower()
        excel = request.files.get("excel")
        if not excel:
            return jsonify(error="excel file required"), 400
        stash = os.path.join(CASES_IMPORTS_DIR,
                             f"bulk_{uuid.uuid4().hex}.xlsx")
        excel.save(stash)
        try:
            wb = openpyxl.load_workbook(stash, read_only=True,
                                        data_only=True)
            ws = wb.active
            rows_iter = ws.iter_rows(values_only=True)
            try:
                header_row = next(rows_iter)
            except StopIteration:
                return jsonify(error="empty excel"), 400
            headers = [str(h).strip() if h is not None else ""
                       for h in header_row]
            aliases = CASE_COL_ALIASES.get(field, ())
            match = _find_col(headers, aliases)
            col_idx = headers.index(match) if match else 0
            src_col = match or (headers[0] if headers else "")
            seen = set()
            for row in rows_iter:
                if col_idx >= len(row):
                    continue
                v = row[col_idx]
                if v is None:
                    continue
                if isinstance(v, float) and v.is_integer():
                    v = int(v)
                s = str(v).strip()
                if s and s not in seen:
                    seen.add(s)
                    values.append(s)
                    if len(values) >= 100_000:
                        break
            wb.close()
        except Exception as e:
            traceback.print_exc()
            return jsonify(error=f"excel read failed: {e}"), 400
        finally:
            try: os.remove(stash)
            except OSError: pass
    else:
        payload = request.get_json(silent=True) or {}
        field = (payload.get("field") or "cnic").strip().lower()
        raw = payload.get("values") or []
        if not isinstance(raw, list):
            return jsonify(error="values must be a list"), 400
        seen = set()
        for v in raw:
            if v is None:
                continue
            s = str(v).strip()
            if s and s not in seen:
                seen.add(s)
                values.append(s)
                if len(values) >= 100_000:
                    break

    if field not in ("cnic", "name", "phone"):
        return jsonify(error="field must be cnic / name / phone"), 400
    if not values:
        return jsonify(error="no values to match"), 400

    matched_ids = []
    matched_meta = []
    matched_value_set = set()
    with sqlite3.connect(CASES_DB_PATH) as conn:
        conn.row_factory = sqlite3.Row
        _cases_temp_values_table(conn, values)
        for r in conn.execute(
                f"SELECT c.order_id, c.name, c.original_name, c.cnic, "
                f"c.phone, c.{field} AS matched_v "
                f"FROM cases c JOIN _srch_q q ON c.{field} = q.v "
                f"ORDER BY c.order_id"):
            matched_ids.append(r["order_id"])
            matched_value_set.add(r["matched_v"])
            if len(matched_meta) < 200:
                matched_meta.append({
                    "order_id": r["order_id"], "name": r["name"],
                    "original_name": r["original_name"],
                    "cnic": r["cnic"], "phone": r["phone"],
                })
    missing = [v for v in values if v not in matched_value_set]
    return jsonify(
        field=field,
        total_values=len(values),
        matched_count=len(matched_ids),
        missing_count=len(missing),
        missing_sample=missing[:200],
        matched_ids=matched_ids,
        matched_meta=matched_meta,
        source_column=src_col,
    )


@app.route("/api/cases/delete", methods=["POST"])
def api_cases_delete():
    """Delete by order_ids, or `all=true` to wipe the store.
    Kept minimal — the UI exposes this as a tiny safety-wrapped button
    so a misclick on a 400 k-row import is still recoverable (user
    re-uploads)."""
    payload = request.get_json() or {}
    if payload.get("all") is True:
        with _CASES_LOCK, sqlite3.connect(CASES_DB_PATH) as conn:
            conn.execute("DELETE FROM cases")
            conn.commit()
        return jsonify(ok=True, deleted="all")
    ids = payload.get("order_ids") or []
    if not isinstance(ids, list) or not ids:
        return jsonify(error="order_ids must be a non-empty list "
                             "(or pass {\"all\": true})"), 400
    ids = [str(x).strip() for x in ids if str(x).strip()]
    deleted = 0
    with _CASES_LOCK, sqlite3.connect(CASES_DB_PATH) as conn:
        for i in range(0, len(ids), 500):
            chunk = ids[i:i + 500]
            qs = ",".join("?" * len(chunk))
            cur = conn.execute(
                f"DELETE FROM cases WHERE order_id IN ({qs})",
                tuple(chunk))
            deleted += cur.rowcount
        conn.commit()
    return jsonify(ok=True, deleted=deleted)


@app.route("/api/cases/generate", methods=["POST"])
def api_cases_generate():
    """Kick off a batch render for the currently-selected order_ids.
    Body JSON: {
      template_slug, order_ids:[...],
      date_value?:YYYY-MM-DD (default today),
      manual_fields?:{}, filename_fields?:[...],
      group_by_field?:str, machine?:vps|mac }
    Returns {task_id} — client reuses /status and /download."""
    payload = request.get_json() or {}
    template_slug = (payload.get("template_slug") or "default").strip()
    if not _valid_slug(template_slug):
        return jsonify(error="invalid template_slug"), 400
    template = load_template(template_slug)
    if template is None:
        return jsonify(error=f"Template '{template_slug}' not found."), 400

    raw_ids = payload.get("order_ids") or []
    if not isinstance(raw_ids, list) or not raw_ids:
        return jsonify(error="No order_ids selected."), 400
    if len(raw_ids) > 50_000:
        return jsonify(error="Too many order_ids (max 50,000 per batch)."), 400
    # Coerce + dedupe while preserving user's selection order.
    seen, order_ids = set(), []
    for oid in raw_ids:
        s = str(oid).strip()
        if s and s not in seen:
            seen.add(s)
            order_ids.append(s)

    data_rows, missing = _cases_fetch_rows(order_ids)
    if missing:
        return jsonify(
            error=f"{len(missing)} order_id(s) not found in inventory",
            sample=missing[:10]), 400
    if not data_rows:
        return jsonify(error="No matching rows."), 400

    # Date override. Inventory mode's whole point is "use today's date by
    # default"; empty means today, explicit ISO overrides.
    date_str = (payload.get("date_value") or "").strip()
    try:
        d = (datetime.date.fromisoformat(date_str) if date_str
             else datetime.date.today())
    except ValueError:
        return jsonify(error="date_value must be YYYY-MM-DD"), 400
    formatted_date = d.strftime("%d/%m/%Y")
    for r in data_rows:
        r["date"] = formatted_date

    manual_fields   = payload.get("manual_fields") or {}
    filename_fields = payload.get("filename_fields") or ["name"]
    # Inventory mode's standard output shape is "one zip per 负责人,
    # all bundled into a dated batch zip". Caller can override with an
    # explicit column name or pass "__none__" to skip grouping.
    group_by_field  = payload.get("group_by_field")
    if group_by_field is None or group_by_field == "":
        group_by_field = "负责人"
    elif group_by_field == "__none__":
        group_by_field = None
    profile         = _get_machine_profile(payload.get("machine"))

    task_id = _new_task()

    def worker():
        try:
            generate_notices_html(
                template, data_rows,
                manual_fields=manual_fields,
                filename_fields=filename_fields,
                group_by_field=group_by_field,
                task_id=task_id,
                render_workers=profile["render_workers"],
                profile_label=profile["label"],
            )
            # Bundle per-group zips into one dated batch zip so the user
            # downloads a single file instead of N per-group links.
            _wrap_inventory_batch(task_id)
        except Exception as e:
            traceback.print_exc()
            _update_task(task_id, status="error", stage="error",
                         error=f"{e.__class__.__name__}: {e}",
                         message=f"Generation failed: {e}")

    threading.Thread(target=worker, daemon=True).start()
    return jsonify(task_id=task_id, row_count=len(data_rows))


@app.route("/generate_one", methods=["POST"])
def generate_one():
    """Render ONE notice from user-typed field values and stream the
    PDF back immediately. Body JSON:
      {template_slug, fields, date_value?}
    `fields` maps placeholder-name → user value. `date_value` (ISO
    YYYY-MM-DD from the date picker) overrides `fields["date"]` with
    DD/MM/YYYY formatting. Response is a `application/pdf` attachment."""
    payload = request.get_json(silent=True) or {}
    slug = (payload.get("template_slug") or "default").strip()
    if not _valid_slug(slug):
        return jsonify(error="invalid template_slug"), 400
    tpl = load_template(slug)
    if tpl is None:
        return jsonify(error=f"template '{slug}' not found"), 400
    fields = dict(payload.get("fields") or {})
    date_value = (payload.get("date_value") or "").strip()
    if date_value:
        try:
            d = datetime.date.fromisoformat(date_value)
            fields["date"] = d.strftime("%d/%m/%Y")
        except ValueError:
            pass

    # Per-request concurrency gate: each Chromium render takes ~0.3-1s
    # and ~300 MB RSS. Without this, a user script (or a double-clicked
    # button) can line up 200 requests and OOM the VPS.
    if not _SINGLE_GEN_SEM.acquire(blocking=False):
        return jsonify(error="renderer busy, retry in a moment"), 429

    safe_name = _safe_name(fields.get("name") or "notice") or "notice"
    tmp = tempfile.mkdtemp(prefix="notice_one_")
    try:
        out_pdf = os.path.join(tmp, "notice.pdf")
        # Pooled Chromium: cold-start is ~1s, keeping one browser across
        # single-notice requests drops the per-request tax to render-only.
        render_notice_row_pdf(tpl, fields, out_pdf,
                              browser=_get_shared_single_browser(),
                              rasterize=True, lock=True)
        with open(out_pdf, "rb") as f:
            data = f.read()
        # Notice-history audit: log every successful single render so the
        # serial printed under the QR can be reverse-looked-up later.
        _log_notice_record(
            name=str(fields.get("name", "") or ""),
            principal=str(fields.get("Principal_Amount", "") or ""),
        )
    except Exception as e:
        traceback.print_exc()
        shutil.rmtree(tmp, ignore_errors=True)
        _SINGLE_GEN_SEM.release()
        return jsonify(error=f"{e.__class__.__name__}: {e}"), 500
    shutil.rmtree(tmp, ignore_errors=True)
    _SINGLE_GEN_SEM.release()
    # HTTP headers are latin-1 only — a Chinese (or any non-ASCII) name
    # in safe_name kills the response with a UnicodeEncodeError. Follow
    # RFC 5987: give both an ASCII fallback filename= and a UTF-8
    # filename*= that modern browsers prefer.
    from urllib.parse import quote
    ascii_fallback = safe_name.encode("ascii", "ignore").decode() or "notice"
    disp = (f'attachment; filename="{ascii_fallback}.pdf"; '
            f"filename*=UTF-8''{quote(safe_name + '.pdf')}")
    return Response(
        data, mimetype="application/pdf",
        headers={
            "Content-Disposition": disp,
            "Cache-Control": "no-store",
        })


# ── HTML template (unified single-page workstation) ─────────────

HTML_TEMPLATE = r"""
<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>S&amp;S Law Firm — Legal Notice Workstation</title>
<link rel="icon" type="image/png" href="__LOGO_DATA_URI__">
<style>
  :root { --primary: #0b1220; --accent: #c9a14a;
          --danger: #dc2626; --success: #16a34a;
          --bg: #f4f1ea; --panel: #fff; --border: #e0dcd2; }
  * { box-sizing: border-box; margin: 0; padding: 0; }
  html, body { height: 100vh; overflow: hidden; }
  body { font-family: -apple-system, "Segoe UI", Roboto, sans-serif;
         background: var(--bg); color: #111; line-height: 1.45;
         font-size: 13px; }

  /* ── Top brand bar ── */
  .brand { background: linear-gradient(180deg, #0b1220 0%, #111a2e 100%);
           border-bottom: 2px solid var(--accent);
           padding: 10px 20px; display: flex; align-items: center;
           gap: 14px; height: 56px; }
  .brand .logo { height: 32px; width: auto;
                 filter: drop-shadow(0 1px 3px rgba(0,0,0,.3)); }
  .brand .name { color: #fff; font-size: 1rem; font-weight: 700;
                 letter-spacing: .02em; }
  .brand .sub { color: #cbd5e1; font-size: .72rem;
                text-transform: uppercase; letter-spacing: .14em;
                margin-left: 4px; }
  .brand .spacer { flex: 1; }
  .brand a { color: #cbd5e1; text-decoration: none; font-size: .72rem;
             letter-spacing: .12em; text-transform: uppercase;
             padding: 6px 10px; border: 1px solid rgba(201,161,74,.3);
             border-radius: 4px; transition: .15s; }
  .brand a:hover { color: var(--accent); border-color: var(--accent);
                    background: rgba(201,161,74,.08); }

  /* ── 2-column layout ── */
  .layout { display: grid; grid-template-columns: 420px 1fr;
            height: calc(100vh - 56px); }
  .sidebar { background: var(--panel); border-right: 1px solid var(--border);
             overflow-y: auto; padding: 12px 16px 40px; }
  .preview { background: #e8e4d9; overflow: auto; }
  .preview iframe { width: 100%; height: 100%; border: 0; background: #fff; }

  /* ── Collapsible sections ── */
  details { border: 1px solid var(--border); border-radius: 6px;
            margin-bottom: 10px; background: #fcfaf5; overflow: hidden; }
  details[open] { background: #fff; }
  summary { list-style: none; cursor: pointer; padding: 10px 12px;
            font-weight: 700; font-size: .84rem; color: #222;
            display: flex; align-items: center; gap: 8px;
            letter-spacing: .01em; user-select: none; }
  summary::-webkit-details-marker { display: none; }
  summary::before { content: "▶"; font-size: .7rem; color: #888;
                    transition: transform .15s; }
  details[open] summary::before { transform: rotate(90deg); }
  summary .hint { margin-left: auto; font-weight: 400; font-size: .72rem;
                  color: #888; }
  .section-body { padding: 0 12px 12px; }

  /* ── Form controls ── */
  .field { margin-bottom: 10px; }
  .field label { display: block; font-weight: 600; font-size: .76rem;
                 color: #333; margin-bottom: 3px; }
  .field input[type=text], .field input[type=date],
  .field input[type=number], .field input[type=password],
  .field select, .field textarea {
    width: 100%; padding: 6px 9px; border: 1px solid #cbd5e1;
    border-radius: 5px; font-size: .84rem; font-family: inherit; }
  .field textarea { min-height: 64px; resize: vertical; line-height: 1.5; }
  .field textarea.big { min-height: 110px; }
  .field textarea.huge { min-height: 180px; }
  .field input:focus, .field select:focus, .field textarea:focus {
    outline: none; border-color: var(--accent);
    box-shadow: 0 0 0 2px rgba(201,161,74,.18); }
  .field .hint { font-size: .7rem; color: #888; margin-top: 3px;
                 line-height: 1.35; }
  .field .row { display: grid; grid-template-columns: 1fr auto;
                align-items: center; gap: 8px; }
  .field .out { font-family: ui-monospace, Menlo, monospace;
                font-size: .72rem; color: #555; min-width: 42px;
                text-align: right; }
  .field input[type=range] { width: 100%; accent-color: var(--accent); }
  .field input[type=color] { width: 100%; height: 30px; padding: 0;
                             border: 1px solid #cbd5e1; border-radius: 5px; }
  .toggle { display: flex; align-items: center; gap: 8px;
            font-size: .82rem; font-weight: 600; color: #111;
            cursor: pointer; margin-bottom: 8px; }
  .toggle input { width: 14px; height: 14px; accent-color: var(--accent); }
  .row2 { display: grid; grid-template-columns: 1fr 1fr; gap: 8px; }

  /* ── Image upload slot ── */
  .img-slot { display: grid; grid-template-columns: 64px 1fr;
              gap: 10px; padding: 8px; border: 1px solid var(--border);
              border-radius: 6px; margin-bottom: 8px; align-items: center;
              background: #fafaf7; }
  .img-slot img { width: 64px; height: 64px; object-fit: contain;
                  background: #fff; border: 1px solid #e5e5e5;
                  border-radius: 4px; }
  .img-slot .meta .kind { font-weight: 700; font-size: .8rem; }
  .img-slot .meta .sub { font-size: .68rem; color: #888; line-height: 1.3; }
  .img-slot .meta .btns { margin-top: 5px; display: flex; gap: 6px; }
  .img-slot .meta label.btn-file {
    padding: 4px 10px; font-size: .72rem; background: #f1efe9;
    border: 1px solid #d6d2c9; border-radius: 4px; cursor: pointer; }
  .img-slot .meta label.btn-file:hover { background: #e8e4d9; }
  .img-slot .meta .reset { padding: 4px 8px; font-size: .72rem;
                           background: transparent; border: 1px solid #e5e5e5;
                           color: #888; border-radius: 4px; cursor: pointer; }
  .img-slot .meta .placement-grid {
    display: grid; grid-template-columns: 1fr 1fr; gap: 4px 8px;
    margin-top: 7px;
  }
  .img-slot .meta .placement-grid label {
    font-size: .66rem; color: #666; display: flex; flex-direction: column;
    gap: 2px;
  }
  .img-slot .meta .placement-grid input[type="number"] {
    padding: 3px 5px; font-size: .78rem; border: 1px solid var(--border);
    border-radius: 3px; background: #fff; width: 100%;
  }

  /* ── Buttons ── */
  .btn { display: inline-block; padding: 7px 13px; border: none;
         border-radius: 5px; font-size: .82rem; font-weight: 600;
         cursor: pointer; transition: .12s; line-height: 1.2; }
  .btn-primary { background: var(--primary); color: #fff; }
  .btn-primary:hover:not(:disabled) { background: #1a2541; }
  .btn-primary:disabled { background: #9ca3af; cursor: not-allowed; }
  .btn-secondary { background: #f1efe9; color: #333;
                   border: 1px solid #d6d2c9; }
  .btn-secondary:hover { background: #e8e4d9; }
  .btn-danger:disabled,
  .btn-secondary:disabled { opacity: .4; cursor: not-allowed; }

  /* ── Generate-mode radio toggle (Excel vs Manual) ── */
  .mode-toggle {
    display: flex; gap: 4px; margin-bottom: 12px;
    background: #f1efe9; border: 1px solid var(--border);
    border-radius: 5px; padding: 3px;
  }
  .mode-toggle label {
    flex: 1; text-align: center; font-size: .78rem;
    padding: 5px 8px; border-radius: 4px; cursor: pointer;
    font-weight: 600; color: #555; user-select: none;
  }
  .mode-toggle input[type="radio"] { display: none; }
  .mode-toggle input[type="radio"]:checked + span,
  .mode-toggle label:has(input:checked) {
    background: var(--primary); color: #fff;
  }

  /* ── Manual-single placeholder grid ── */
  #manualFieldGrid {
    display: grid; grid-template-columns: 1fr 1fr; gap: 6px 8px;
    margin-bottom: 12px;
  }
  #manualFieldGrid .field { margin: 0; }
  #manualFieldGrid .field.full { grid-column: 1 / -1; }
  #manualFieldGrid .field label {
    font-family: ui-monospace, "SF Mono", "Menlo", "Consolas", monospace;
    font-size: .7rem; color: #555; margin-bottom: 2px;
  }
  #manualFieldGrid .field input {
    padding: 5px 7px; font-size: .82rem;
    border: 1px solid var(--border); border-radius: 4px;
    background: #fff; width: 100%;
  }
  #manualFieldGrid .field input[readonly] {
    background: #f5f1e6; color: #555; cursor: not-allowed;
    font-weight: 700;
  }
  #manualFieldGrid .field .auto-tag {
    font-family: -apple-system, "Segoe UI", Roboto, sans-serif;
    font-size: .62rem; font-weight: 500;
    color: #8a7a3a; margin-left: 4px;
    text-transform: none; letter-spacing: 0;
  }

  .btn-danger { background: #fef2f2; color: var(--danger);
                border: 1px solid #fecaca; }
  .btn-danger:hover { background: #fee2e2; }
  .btn-block { width: 100%; text-align: center; }
  .btn-row { display: flex; gap: 6px; flex-wrap: wrap; }

  /* ── Tags & info ── */
  .tag { display: inline-block; padding: 2px 8px; border-radius: 999px;
         font-size: .72rem; margin: 2px 2px; }
  .tag-ok { background: #dcfce7; color: #166534; }
  .tag-miss { background: #fee2e2; color: #991b1b; }
  .info { padding: 8px 12px; border-radius: 5px; font-size: .78rem;
          margin: 8px 0; line-height: 1.4; }
  .info-ok { background: #f0fdf4; border: 1px solid #bbf7d0; color: #166534; }
  .info-warn { background: #fffbeb; border: 1px solid #fde68a; color: #92400e; }
  .info-err { background: #fef2f2; border: 1px solid #fecaca; color: #991b1b; }

  .check-grid { display: flex; flex-wrap: wrap; gap: 4px 12px;
                padding: 8px 10px; border: 1px solid #cbd5e1;
                border-radius: 5px; max-height: 100px; overflow-y: auto;
                background: #fff; }
  .check-grid label { font-size: .76rem; font-weight: 400; margin: 0;
                      display: inline-flex; align-items: center; gap: 4px;
                      cursor: pointer; }

  table.preview { width: 100%; border-collapse: collapse;
                  font-size: .72rem; margin-top: 4px; }
  table.preview th, table.preview td { padding: 3px 6px;
                  border: 1px solid #e2e8f0; text-align: left;
                  white-space: nowrap; max-width: 120px;
                  overflow: hidden; text-overflow: ellipsis; }
  table.preview th { background: #f1efe9; font-weight: 600; }

  /* ── Progress ── */
  .progress-bar { width: 100%; height: 8px; background: #e2e8f0;
                  border-radius: 999px; overflow: hidden; margin-top: 8px; }
  .progress-fill { height: 100%; background: var(--primary);
                   width: 0%; transition: width .3s ease; }
  .progress-text { font-size: .74rem; color: #666; margin-top: 4px;
                   line-height: 1.3; }

  .spinner { display: inline-block; width: 14px; height: 14px;
             border: 2px solid #fff; border-top-color: transparent;
             border-radius: 50%; animation: spin .6s linear infinite;
             vertical-align: middle; margin-right: 6px; }
  @keyframes spin { to { transform: rotate(360deg); } }
  .status-msg { font-size: .72rem; color: #666; margin-top: 5px;
                line-height: 1.35; }
  .status-msg.ok { color: var(--success); }
  .status-msg.err { color: var(--danger); }

  code { background: #f1efe9; padding: 0 3px; border-radius: 3px;
         font-family: ui-monospace, Menlo, monospace; font-size: .82em; }
</style>
</head>
<body>

<header class="brand">
  <img class="logo" src="__LOGO_DATA_URI__" alt="S&amp;S Law Firm">
  <span class="name">S&amp;S LAW FIRM</span>
  <span class="sub">Legal Notice Workstation</span>
  <span class="spacer"></span>
  <a href="/inventory" style="margin-right:10px;">📦 Inventory</a>
  <a href="/verify" style="margin-right:10px;">🔍 Verify</a>
  <a href="/logout">Sign Out</a>
</header>

<div class="layout">

<aside class="sidebar">

  <!-- ═══ Section 1 · Template ═══ -->
  <details open>
    <summary>Template <span class="hint" id="templateHint"></span></summary>
    <div class="section-body">
      <div class="field">
        <label>Active template</label>
        <div class="row" style="grid-template-columns: 1fr auto;">
          <select id="templateSlug"></select>
          <button class="btn btn-danger" id="tplDeleteBtn" onclick="deleteTemplate()" title="Delete this template">Delete</button>
        </div>
      </div>
      <div class="field">
        <label>Template name</label>
        <input type="text" id="templateName">
      </div>
      <div class="btn-row">
        <button class="btn btn-secondary" onclick="renameTemplate()">Rename</button>
        <button class="btn btn-secondary" onclick="cloneAsNew()">Save as new…</button>
        <button class="btn btn-primary" onclick="reloadTemplate()">Reload</button>
      </div>
      <div class="status-msg" id="tplStatus"></div>
    </div>
  </details>

  <!-- ═══ Section 2 · Body content ═══ -->
  <details>
    <summary>Body Content <span class="hint">7 editable blocks</span></summary>
    <div class="section-body">
      <div class="field">
        <label><code>letterhead-firm</code> — firm name (line 1) + tag (line 2)</label>
        <textarea id="bk-letterhead-firm" class="big"></textarea>
      </div>
      <div class="field">
        <label><code>letterhead-partners</code> — one partner per line, format: <code>Name | Role</code></label>
        <textarea id="bk-letterhead-partners" class="big"></textarea>
      </div>
      <div class="field">
        <label><code>notice-subject</code> — subject line</label>
        <textarea id="bk-notice-subject" class="big"></textarea>
      </div>
      <div class="field">
        <label><code>notice-body-text</code> — main narrative (paragraphs split by blank line; <code>[[AMOUNTS_TABLE]]</code> and <code>[[CALLOUT]]</code> are magic markers)</label>
        <textarea id="bk-notice-body-text" class="huge"></textarea>
      </div>
      <div class="field">
        <label><code>legal-consequences</code> — one consequence per line</label>
        <textarea id="bk-legal-consequences" class="huge"></textarea>
      </div>
      <div class="field">
        <label><code>payment-instructions</code> — paragraphs split by blank line</label>
        <textarea id="bk-payment-instructions" class="big"></textarea>
      </div>
      <div class="field">
        <label><code>page-footer</code> — one line per <code>&lt;p&gt;</code>; <code>Office:</code> / <code>Email:</code> / <code>Phone:</code> auto-bolded</label>
        <textarea id="bk-page-footer" class="big"></textarea>
      </div>
      <div class="btn-row">
        <button class="btn btn-secondary" onclick="resetBlocks()">Reset to defaults</button>
        <button class="btn btn-primary" onclick="saveBlocks()">Save body</button>
      </div>
      <div class="status-msg" id="blocksStatus"></div>
    </div>
  </details>

  <!-- ═══ Section 3 · Images ═══ -->
  <details>
    <summary>Images <span class="hint">upload · size · offset · rotation</span></summary>
    <div class="section-body">
      <div id="imageSlots"></div>
      <div class="btn-row">
        <button class="btn btn-secondary" onclick="resetAssetsConfig()">Reset placement</button>
        <button class="btn btn-primary" onclick="saveAssetsConfig()">Save placement</button>
      </div>
      <div class="status-msg" id="assetsStatus"></div>
    </div>
  </details>

  <!-- ═══ Section 4 · Watermark & pattern ═══ -->
  <details>
    <summary>Watermark &amp; Pattern <span class="hint">anti-counterfeit</span></summary>
    <div class="section-body">
      <label class="toggle">
        <input type="checkbox" id="wm-enabled"> Enable watermark
      </label>
      <div class="field">
        <label>English text (use <code>{name}</code>)</label>
        <textarea id="wm-en"></textarea>
      </div>
      <div class="field">
        <label>Urdu text (use <code>{name}</code>)</label>
        <textarea id="wm-ur" dir="auto"></textarea>
      </div>
      <div class="row2">
        <div class="field">
          <label>Font size (px)</label>
          <div class="row">
            <input type="range" id="wm-size" min="10" max="64" step="1">
            <span class="out" id="wm-size-out"></span>
          </div>
        </div>
        <div class="field">
          <label>Opacity (%)</label>
          <div class="row">
            <input type="range" id="wm-op" min="0" max="100" step="1">
            <span class="out" id="wm-op-out"></span>
          </div>
        </div>
      </div>
      <div class="row2">
        <div class="field">
          <label>Count / page</label>
          <div class="row">
            <input type="range" id="wm-count" min="1" max="100" step="1">
            <span class="out" id="wm-count-out"></span>
          </div>
        </div>
        <div class="field">
          <label>Ink color</label>
          <input type="color" id="wm-color">
        </div>
      </div>

      <label class="toggle" style="margin-top:14px;">
        <input type="checkbox" id="pt-enabled"> Enable pattern
      </label>
      <div class="row2">
        <div class="field">
          <label>Pattern opacity (%)</label>
          <div class="row">
            <input type="range" id="pt-op" min="0" max="100" step="1">
            <span class="out" id="pt-op-out"></span>
          </div>
        </div>
        <div class="field">
          <label>Density</label>
          <select id="pt-density">
            <option value="low">Low</option>
            <option value="medium">Medium</option>
            <option value="high">High</option>
            <option value="ultra">Ultra</option>
          </select>
        </div>
      </div>
      <div class="btn-row">
        <button class="btn btn-secondary" onclick="resetSecurity()">Reset</button>
        <button class="btn btn-primary" onclick="saveSecurity()">Save security</button>
      </div>
      <div class="status-msg" id="secStatus"></div>
    </div>
  </details>

  <!-- ═══ Section 5 · Generate ═══ -->
  <details open>
    <summary>Generate <span class="hint">Excel batch · manual single</span></summary>
    <div class="section-body">

      <!-- Mode toggle — Excel batch vs manual single-notice -->
      <div class="mode-toggle">
        <label><input type="radio" name="gen-mode" value="excel" checked> Excel batch</label>
        <label><input type="radio" name="gen-mode" value="manual"> Manual single</label>
      </div>

      <!-- ── EXCEL PANEL ────────────────────────────── -->
      <div id="excelPanel">
        <div class="field">
          <label>Excel data file (.xlsx)</label>
          <input type="file" id="xlsFile" accept=".xlsx,.xls">
        </div>
        <button class="btn btn-primary btn-block" id="analyzeBtn" onclick="doUpload()" disabled>
          Analyze Excel
        </button>

        <div id="analysisArea" style="display:none; margin-top: 12px;">
          <div id="matchInfo"></div>
          <div id="placeholderTags" style="margin-bottom: 8px;"></div>

          <div id="manualSection" style="display:none;">
            <div class="info info-warn" style="margin-bottom: 6px;">
              Fields below are not in the Excel — enter values shared across all rows.
            </div>
            <div id="manualFields"></div>
          </div>

          <div id="previewSection" style="display:none; margin-top: 10px;">
            <label style="font-size:.72rem; font-weight:600; color:#555;">Data preview (first 5 rows)</label>
            <div style="overflow-x:auto; max-height: 150px; overflow-y: auto;" id="previewTable"></div>
          </div>

          <div class="field" id="dateSection" style="display:none; margin-top:10px;">
            <label>Notice date &mdash; overrides any <code>date</code> column</label>
            <input type="date" id="dateValue">
          </div>

          <div class="field">
            <label>Filename columns (joined with <code>_</code>)</label>
            <div class="check-grid" id="filenameFields"></div>
          </div>

          <div class="row2">
            <div class="field">
              <label>Group by</label>
              <select id="groupByField">
                <option value="">No grouping</option>
              </select>
            </div>
            <div class="field">
              <label>Machine profile</label>
              <select id="machineProfile">
                <option value="vps" selected>VPS — 1 worker</option>
                <option value="mac">Mac — up to 6 workers</option>
              </select>
            </div>
          </div>

          <button class="btn btn-primary btn-block" id="genBtn" onclick="doGenerate()">
            Generate Notices
          </button>
          <div id="progressSection" style="display:none">
            <div class="progress-bar"><div class="progress-fill" id="progressFill"></div></div>
            <div class="progress-text" id="progressText">Starting…</div>
          </div>
          <div class="status-msg" id="genStatus"></div>
        </div>
      </div>

      <!-- ── MANUAL PANEL — single notice ──────────── -->
      <div id="manualPanel" style="display:none;">
        <div class="info" style="margin-bottom:10px;">
          Fill every placeholder below, then generate one PDF.
          Uses the current template's saved blocks / images / watermark.
        </div>
        <div id="manualFieldGrid"></div>
        <button class="btn btn-primary btn-block" id="genOneBtn" onclick="doGenerateOne()">
          Generate single PDF
        </button>
        <div class="status-msg" id="genOneStatus"></div>
      </div>

    </div>
  </details>

</aside>

<main class="preview">
  <iframe id="preview" src="about:blank"></iframe>
</main>

</div>

<script>
let activeSlug  = 'default';
let activeTpl   = null;
let excelAnalysis = null;
let previewDebounce = null;

// ── helpers ────────────────────────────────────────
const $ = id => document.getElementById(id);
function flash(el, msg, kind = '') {
  el.textContent = msg;
  el.className = 'status-msg ' + kind;
  if (!kind || kind === 'ok')
    setTimeout(() => { if (el.textContent === msg) el.textContent = ''; }, 2500);
}
function bindRange(range, out, suffix = '') {
  const up = () => out.textContent = range.value + suffix;
  range.addEventListener('input', up); up();
}
// Live preview: debounce → POST current sidebar state → srcdoc into iframe.
// Lets the user see edits without having to press any Save button first.
// The actual PDF render still uses saved state, so Save before Generate.
function reloadPreview() {
  clearTimeout(previewDebounce);
  previewDebounce = setTimeout(async () => {
    if (!activeTpl) return;
    try {
      const body = {
        slug: activeSlug,
        blocks: currentBlocks(),
        security: currentSecurity(),
        assets_config: currentAssetsConfig(),
      };
      // In Manual mode, also push typed placeholder values so the
      // preview reflects them. Missing keys stay as literal {{foo}}.
      const mp = $('manualPanel');
      if (mp && mp.style.display !== 'none') {
        body.fields = collectManualFields();
      }
      const r = await fetch('/api/preview_html', {
        method: 'POST', headers: {'Content-Type': 'application/json'},
        body: JSON.stringify(body),
      });
      if (!r.ok) return;
      $('preview').srcdoc = await r.text();
    } catch (e) { /* ignore transient preview failures */ }
  }, 300);
}
function wireStaticLivePreview() {
  // Static elements exist at boot. Bind each once.
  const sel = 'textarea[id^="bk-"], #wm-enabled, #wm-en, #wm-ur, #wm-size,'
            + ' #wm-op, #wm-count, #wm-color, #pt-enabled, #pt-op, #pt-density';
  document.querySelectorAll(sel).forEach(el => {
    const ev = (el.type === 'checkbox' || el.tagName === 'SELECT') ? 'change' : 'input';
    el.addEventListener(ev, reloadPreview);
  });
}
function wireDynamicLivePreview() {
  // Placement number inputs are re-created every renderImageSlots();
  // the old DOM nodes (with their listeners) are garbage-collected, so
  // just bind fresh each time.
  document.querySelectorAll('[data-ac]').forEach(el => {
    el.addEventListener('input', reloadPreview);
  });
}

// ── 1. template list / active template ────────────
async function loadTemplateList() {
  const r = await fetch('/api/templates');
  const items = await r.json();
  const sel = $('templateSlug');
  sel.innerHTML = items.map(t =>
    `<option value="${t.slug}">${t.name}${t.builtin ? ' (built-in)' : ''}</option>`
  ).join('');
  if (!items.find(t => t.slug === activeSlug)) activeSlug = 'default';
  sel.value = activeSlug;
}
const BLOCK_IDS = [
  'letterhead-firm', 'letterhead-partners',
  'notice-subject', 'notice-body-text',
  'legal-consequences', 'payment-instructions',
  'page-footer',
];
async function loadActiveTemplate() {
  const r = await fetch('/api/templates/' + encodeURIComponent(activeSlug));
  if (!r.ok) { alert('Failed to load template'); return; }
  activeTpl = await r.json();
  $('templateName').value = activeTpl.name || '';
  $('templateHint').textContent = activeTpl.builtin ? 'built-in' : 'custom';
  // Built-in template can't be deleted — hide the button entirely so
  // the user never sees "Delete" next to a protected template.
  $('tplDeleteBtn').disabled = activeTpl.builtin;
  $('tplDeleteBtn').style.display = activeTpl.builtin ? 'none' : '';
  // blocks
  for (const p of BLOCK_IDS) {
    const el = $('bk-' + p);
    if (el) el.value = (activeTpl.blocks || {})[p] || '';
  }
  // security
  const s = activeTpl.security || {};
  const wm = s.watermark || {};
  $('wm-enabled').checked = !!wm.enabled;
  $('wm-en').value = wm.english_template || '';
  $('wm-ur').value = wm.urdu_template || '';
  $('wm-size').value = wm.font_size || 22;
  $('wm-op').value   = wm.opacity   || 67;
  $('wm-count').value = wm.count    || 25;
  $('wm-color').value = wm.color   || '#323255';
  const pt = s.pattern || {};
  $('pt-enabled').checked = !!pt.enabled;
  $('pt-op').value      = pt.opacity || 22;
  $('pt-density').value = pt.density || 'medium';
  bindRange($('wm-size'),  $('wm-size-out'),  'px');
  bindRange($('wm-op'),    $('wm-op-out'),    '%');
  bindRange($('wm-count'), $('wm-count-out'), '');
  bindRange($('pt-op'),    $('pt-op-out'),    '%');
  renderImageSlots();
  // Refresh the manual-single placeholder grid if it's the visible mode
  // — different templates may expose different {{placeholders}}.
  if ($('manualPanel') && $('manualPanel').style.display !== 'none') {
    loadManualFields();
  }
  reloadPreview();
}
$('templateSlug').addEventListener('change', async e => {
  activeSlug = e.target.value;
  await loadActiveTemplate();
});

async function reloadTemplate() {
  await loadTemplateList();
  await loadActiveTemplate();
}
async function renameTemplate() {
  const name = $('templateName').value.trim();
  if (!name) return;
  const r = await fetch('/api/templates/' + encodeURIComponent(activeSlug), {
    method: 'PUT', headers: {'Content-Type': 'application/json'},
    body: JSON.stringify({name})
  });
  const d = await r.json();
  if (!r.ok) return flash($('tplStatus'), d.error || 'failed', 'err');
  activeTpl = d.template;
  await loadTemplateList();
  $('templateSlug').value = activeSlug;
  flash($('tplStatus'), 'Renamed.', 'ok');
}
async function cloneAsNew() {
  const name = prompt('New template name:', ($('templateName').value || 'Template') + ' Copy');
  if (!name) return;
  const r = await fetch('/api/templates', {
    method: 'POST', headers: {'Content-Type': 'application/json'},
    body: JSON.stringify({name, base_slug: activeSlug,
                          blocks: currentBlocks(),
                          security: currentSecurity(),
                          assets_config: currentAssetsConfig()}),
  });
  const d = await r.json();
  if (!r.ok) return flash($('tplStatus'), d.error || 'failed', 'err');
  activeSlug = d.template.slug;
  await loadTemplateList();
  $('templateSlug').value = activeSlug;
  await loadActiveTemplate();
  flash($('tplStatus'), 'Created "' + d.template.name + '".', 'ok');
}
async function deleteTemplate() {
  if (activeTpl?.builtin) return;
  if (!confirm('Delete template "' + activeTpl.name + '"? This cannot be undone.')) return;
  const r = await fetch('/api/templates/' + encodeURIComponent(activeSlug), {method: 'DELETE'});
  if (r.ok) {
    activeSlug = 'default';
    await loadTemplateList();
    $('templateSlug').value = 'default';
    await loadActiveTemplate();
    flash($('tplStatus'), 'Deleted.', 'ok');
  }
}

// ── 2. body blocks ────────────────────────────────
function currentBlocks() {
  const out = {};
  for (const p of BLOCK_IDS) {
    const el = $('bk-' + p);
    if (el) out[p] = el.value;
  }
  return out;
}
async function saveBlocks() {
  const r = await fetch('/api/templates/' + encodeURIComponent(activeSlug), {
    method: 'PUT', headers: {'Content-Type': 'application/json'},
    body: JSON.stringify({blocks: currentBlocks()}),
  });
  const d = await r.json();
  if (!r.ok) return flash($('blocksStatus'), d.error || 'failed', 'err');
  activeTpl = d.template;
  flash($('blocksStatus'), 'Saved body content.', 'ok');
  reloadPreview();
}
async function resetBlocks() {
  if (!confirm('Discard your block edits and reload?')) return;
  await loadActiveTemplate();
}

// ── 3. images ─────────────────────────────────────
// Per-kind configurable placement ranges. The rotation is a BASE angle —
// the anti-counterfeit overlay adds a per-document random ±15° on top for
// seal and signature_seal, so even with rot=0 each PDF still gets a
// unique jitter. Logo has no random jitter; its rot is the final angle.
const IMG_SLOTS = [
  {kind:'logo',           label:'Logo (top of page)',      note:'1000×1000 transparent PNG',
   size:{min:12, max:30, step:0.5}, off:{min:-20, max:20, step:0.5}, rot:{min:-30, max:30, step:1}},
  {kind:'seal',           label:'Firm seal (red round)',   note:'1000×1000 transparent PNG',
   size:{min:16, max:36, step:0.5}, off:{min:-20, max:20, step:0.5}, rot:{min:-30, max:30, step:1}},
  {kind:'signature_seal', label:'Signature seal (blue)',   note:'2000×600 transparent PNG',
   size:{min:40, max:80, step:0.5}, off:{min:-20, max:20, step:0.5}, rot:{min:-30, max:30, step:1}},
];
function _acVal(kind, key, fallback) {
  const c = (activeTpl.assets_config || {})[kind] || {};
  const v = c[key];
  return (v === undefined || v === null || isNaN(+v)) ? fallback : +v;
}
function renderImageSlots() {
  $('imageSlots').innerHTML = IMG_SLOTS.map(s => {
    const has = !!(activeTpl.assets || {})[s.kind];
    const sz  = _acVal(s.kind, 'size', {logo:22, seal:22, signature_seal:44}[s.kind]);
    const dx  = _acVal(s.kind, 'dx',  0);
    const dy  = _acVal(s.kind, 'dy',  0);
    const rot = _acVal(s.kind, 'rot', 0);
    return `
      <div class="img-slot">
        <img src="/api/templates/${activeSlug}/assets/${s.kind}?t=${Date.now()}" alt="${s.kind}"
             onerror="this.style.opacity=.2">
        <div class="meta">
          <div class="kind">${s.label}</div>
          <div class="sub">${s.note} · ${has ? 'custom upload' : 'shipped default'}</div>
          <div class="btns">
            <label class="btn-file">
              Upload
              <input type="file" accept="image/png" style="display:none"
                     onchange="uploadAsset('${s.kind}', this.files[0])">
            </label>
            ${has ? `<button class="reset" onclick="resetAsset('${s.kind}')">Reset image</button>` : ''}
          </div>
          <div class="placement-grid">
            <label>size (mm)<input type="number" data-ac="${s.kind}.size"
                   value="${sz}" min="${s.size.min}" max="${s.size.max}" step="${s.size.step}"></label>
            <label>rotation (°)<input type="number" data-ac="${s.kind}.rot"
                   value="${rot}" min="${s.rot.min}" max="${s.rot.max}" step="${s.rot.step}"></label>
            <label>offset X (mm)<input type="number" data-ac="${s.kind}.dx"
                   value="${dx}" min="${s.off.min}" max="${s.off.max}" step="${s.off.step}"></label>
            <label>offset Y (mm)<input type="number" data-ac="${s.kind}.dy"
                   value="${dy}" min="${s.off.min}" max="${s.off.max}" step="${s.off.step}"></label>
          </div>
        </div>
      </div>`;
  }).join('');
  wireDynamicLivePreview();
}
function currentAssetsConfig() {
  const out = {};
  for (const s of IMG_SLOTS) out[s.kind] = {};
  for (const inp of document.querySelectorAll('[data-ac]')) {
    const [kind, key] = inp.dataset.ac.split('.');
    out[kind][key] = +inp.value;
  }
  return out;
}
async function saveAssetsConfig() {
  const r = await fetch('/api/templates/' + encodeURIComponent(activeSlug), {
    method: 'PUT', headers: {'Content-Type': 'application/json'},
    body: JSON.stringify({assets_config: currentAssetsConfig()}),
  });
  const d = await r.json();
  if (!r.ok) return flash($('assetsStatus'), d.error || 'failed', 'err');
  activeTpl = d.template;
  flash($('assetsStatus'), 'Saved placement.', 'ok');
  reloadPreview();
}
async function resetAssetsConfig() {
  if (!confirm('Discard placement edits and reload?')) return;
  await loadActiveTemplate();
}
async function uploadAsset(kind, file) {
  if (!file) return;
  const fd = new FormData();
  fd.append('file', file);
  const r = await fetch('/api/templates/' + encodeURIComponent(activeSlug) + '/assets/' + kind,
                        {method: 'POST', body: fd});
  const d = await r.json();
  if (!r.ok) return flash($('assetsStatus'), d.error || 'upload failed', 'err');
  await loadActiveTemplate();
  flash($('assetsStatus'), 'Updated ' + kind + '.', 'ok');
}
async function resetAsset(kind) {
  if (!confirm('Revert ' + kind + ' image to the shipped default?')) return;
  const r = await fetch('/api/templates/' + encodeURIComponent(activeSlug) + '/assets/' + kind,
                        {method: 'DELETE'});
  if (r.ok) { await loadActiveTemplate(); flash($('assetsStatus'), 'Reset.', 'ok'); }
}

// ── 4. watermark / pattern ───────────────────────
function currentSecurity() {
  return {
    watermark: {
      enabled: $('wm-enabled').checked,
      english_template: $('wm-en').value,
      urdu_template:    $('wm-ur').value,
      font_size: +$('wm-size').value, opacity: +$('wm-op').value,
      count:     +$('wm-count').value, color: $('wm-color').value,
    },
    pattern: {
      enabled: $('pt-enabled').checked,
      opacity: +$('pt-op').value, density: $('pt-density').value,
    },
  };
}
async function saveSecurity() {
  const r = await fetch('/api/templates/' + encodeURIComponent(activeSlug), {
    method: 'PUT', headers: {'Content-Type': 'application/json'},
    body: JSON.stringify({security: currentSecurity()}),
  });
  const d = await r.json();
  if (!r.ok) return flash($('secStatus'), d.error || 'failed', 'err');
  activeTpl = d.template;
  flash($('secStatus'), 'Saved security overlay.', 'ok');
  reloadPreview();
}
async function resetSecurity() {
  if (!confirm('Discard security edits and reload?')) return;
  await loadActiveTemplate();
}

// ── 5. Excel upload + generate ───────────────────
$('xlsFile').addEventListener('change', function() {
  $('analyzeBtn').disabled = !this.files.length;
});
async function doUpload() {
  const btn = $('analyzeBtn');
  btn.disabled = true;
  btn.innerHTML = '<span class="spinner"></span>Analyzing…';
  const fd = new FormData();
  fd.append('excel', $('xlsFile').files[0]);
  fd.append('template_slug', activeSlug);
  try {
    const r = await fetch('/upload', {method: 'POST', body: fd});
    const d = await r.json();
    if (d.error) { alert(d.error); return; }
    excelAnalysis = d;
    renderAnalysis(d);
  } catch (e) { alert('Upload failed: ' + e.message); }
  finally {
    btn.disabled = false;
    btn.textContent = 'Analyze Excel';
  }
}
function renderAnalysis(d) {
  $('analysisArea').style.display = '';
  const info = $('matchInfo');
  if (d.missing.length === 0) {
    info.innerHTML = `<div class="info info-ok">All ${d.placeholders.length} placeholders matched. ${d.row_count} rows.</div>`;
  } else {
    info.innerHTML = `<div class="info info-warn">${d.matched.length}/${d.placeholders.length} matched · ${d.missing.length} need manual input · ${d.row_count} rows.</div>`;
  }
  $('placeholderTags').innerHTML = d.placeholders.map(p => {
    const ok = d.matched.indexOf(p) !== -1;
    return `<span class="tag ${ok ? 'tag-ok' : 'tag-miss'}">&#123;&#123;${p}&#125;&#125;</span>`;
  }).join('');
  const manualMissing = d.missing.filter(p => p !== 'date');
  const ms = $('manualSection');
  if (manualMissing.length) {
    ms.style.display = '';
    $('manualFields').innerHTML = manualMissing.map(p =>
      `<div class="field"><label>&#123;&#123;${p}&#125;&#125;</label><input type="text" data-field="${p}" placeholder="Value shared across rows"></div>`
    ).join('');
  } else ms.style.display = 'none';
  const dateSec = $('dateSection');
  if (d.placeholders.includes('date')) {
    dateSec.style.display = '';
    if (!$('dateValue').value) {
      const t = new Date();
      $('dateValue').value = t.getFullYear() + '-' +
        String(t.getMonth()+1).padStart(2,'0') + '-' +
        String(t.getDate()).padStart(2,'0');
    }
  } else { dateSec.style.display = 'none'; }
  if (d.preview.length) {
    $('previewSection').style.display = '';
    let html = '<table class="preview"><thead><tr><th>#</th>';
    d.excel_headers.forEach(h => html += `<th>${h}</th>`);
    html += '</tr></thead><tbody>';
    d.preview.forEach((row, i) => {
      html += `<tr><td>${i+1}</td>`;
      d.excel_headers.forEach(h => html += `<td>${row[h] ?? ''}</td>`);
      html += '</tr>';
    });
    html += '</tbody></table>';
    $('previewTable').innerHTML = html;
  }
  $('filenameFields').innerHTML = d.excel_headers.map(h =>
    `<label><input type="checkbox" value="${h}">${h}</label>`).join('');
  $('groupByField').innerHTML = '<option value="">No grouping</option>' +
    d.excel_headers.map(h => `<option value="${h}">${h}</option>`).join('');
}

async function doGenerate() {
  const btn = $('genBtn');
  btn.disabled = true;
  btn.innerHTML = '<span class="spinner"></span>Processing…';
  $('progressSection').style.display = '';
  $('progressFill').style.width = '0%';
  $('progressText').textContent = 'Starting…';
  const gs = $('genStatus');
  gs.textContent = ''; gs.className = 'status-msg';

  const manualFields = {};
  document.querySelectorAll('#manualFields input[data-field]').forEach(i =>
    manualFields[i.dataset.field] = i.value);
  const filenameFields = Array.from(document.querySelectorAll('#filenameFields input:checked'))
    .map(cb => cb.value);

  const downloaded = [];
  const triggered = new Set();
  function trigger(tid, part) {
    const a = document.createElement('a');
    a.href = '/download/' + tid + '/' + part.index;
    a.download = part.name;
    document.body.appendChild(a); a.click(); a.remove();
    downloaded.push(part.name);
    gs.innerHTML = 'Downloaded (' + downloaded.length + '): <b>' + downloaded.join('</b>, <b>') + '</b>';
    gs.className = 'status-msg ok';
  }

  let tid;
  try {
    const r = await fetch('/generate', {
      method: 'POST', headers: {'Content-Type': 'application/json'},
      body: JSON.stringify({
        manual_fields: manualFields, filename_fields: filenameFields,
        group_by_field: $('groupByField').value,
        date_value: $('dateValue').value,
        machine: $('machineProfile').value,
      })
    });
    const d = await r.json();
    if (d.error) { alert(d.error); return; }
    tid = d.task_id;

    while (true) {
      await new Promise(r => setTimeout(r, 900));
      const s = await (await fetch('/status/' + tid)).json();
      if (s.error) throw new Error(s.error);
      const pct = s.total > 0 ? Math.round(s.progress / s.total * 100) : 0;
      $('progressFill').style.width = pct + '%';
      let line = `${s.stage} — ${s.progress}/${s.total} (${pct}%)`;
      if (s.groups_total) {
        const gi = s.status === 'done' ? s.groups_total : Math.min(s.groups_done + 1, s.groups_total);
        line += ` · group ${gi}/${s.groups_total}`;
        if (s.current_group && s.status !== 'done') {
          line += ' [' + s.current_group + (s.group_total ? ' ' + s.group_progress + '/' + s.group_total : '') + ']';
        }
      }
      $('progressText').textContent = line;
      (s.ready_parts || []).forEach(p => {
        if (!triggered.has(p.index)) { triggered.add(p.index); trigger(tid, p); }
      });
      if (s.status === 'done') break;
      if (s.status === 'error') throw new Error(s.error || 'Generation failed');
    }
    $('progressText').textContent = 'Done. ' + downloaded.length + ' group(s) downloaded.';
  } catch (e) {
    gs.textContent = 'Failed: ' + e.message; gs.className = 'status-msg err';
    alert('Generation failed: ' + e.message);
  } finally {
    btn.disabled = false;
    btn.textContent = 'Generate Notices';
  }
}

// ── 6. generate-mode toggle + manual single-notice ─
function setGenMode(mode) {
  $('excelPanel').style.display  = mode === 'excel'  ? '' : 'none';
  $('manualPanel').style.display = mode === 'manual' ? '' : 'none';
  if (mode === 'manual') loadManualFields();
  // Refresh preview — switching modes changes whether typed fields
  // feed into the partial-fill preview.
  reloadPreview();
}
document.querySelectorAll('input[name="gen-mode"]').forEach(r => {
  r.addEventListener('change', e => setGenMode(e.target.value));
});
// Event delegation: any typing / date change inside the manual grid
// triggers the (debounced) preview refresh. Added once here — the
// script runs at end-of-body so the grid element already exists, and
// only its children get regenerated on template switches.
(function wireManualGrid() {
  const grid = $('manualFieldGrid');
  if (!grid) return;
  grid.addEventListener('input', reloadPreview);
  grid.addEventListener('change', reloadPreview);
})();
async function loadManualFields() {
  const r = await fetch('/api/templates/' + encodeURIComponent(activeSlug) + '/placeholders');
  if (!r.ok) { flash($('genOneStatus'), 'failed to load placeholders', 'err'); return; }
  const d = await r.json();
  const grid = $('manualFieldGrid');
  const today = new Date().toISOString().slice(0, 10);
  // Preserve any values the user already typed so switching templates or
  // reloading fields doesn't wipe half-filled forms.
  const prev = {};
  for (const inp of document.querySelectorAll('[data-manual]')) {
    prev[inp.dataset.manual] = inp.value;
  }
  // Any placeholder whose name contains "date" (case-insensitive) gets
  // a date picker — catches `date`, `Due_date`, `disb_date`, etc.
  const isDate  = p => /date/i.test(p);
  // Money placeholders come from the backend (matches MONEY_KEYWORDS).
  const money   = new Set(d.money || []);
  // Payable is derived from Principal + Interest + Penalty — make
  // the input read-only so the user can't fight the total.
  grid.innerHTML = d.placeholders.map(p => {
    const cls = (p === 'notice-body-text' || p.length > 18) ? 'field full' : 'field';
    if (isDate(p)) {
      return `<div class="${cls}"><label>&#123;&#123;${p}&#125;&#125;</label>`
           + `<input type="date" data-manual="${p}" value="${prev[p] || today}"></div>`;
    }
    const v = prev[p] || '';
    if (p === 'Payable') {
      return `<div class="${cls}"><label>&#123;&#123;${p}&#125;&#125;`
           + ` <span class="auto-tag">auto · Principal + Interest + Penalty</span></label>`
           + `<input type="text" data-manual="${p}" data-derived="1" readonly`
           + ` value="${v.replace(/"/g,'&quot;')}" placeholder="0.00"></div>`;
    }
    const moneyAttrs = money.has(p)
      ? ' data-money="1" inputmode="decimal" onblur="fmtMoneyInput(this)"' : '';
    return `<div class="${cls}"><label>&#123;&#123;${p}&#125;&#125;</label>`
         + `<input type="text" data-manual="${p}"${moneyAttrs}`
         + ` value="${v.replace(/"/g,'&quot;')}"`
         + ` placeholder="value for ${p}"></div>`;
  }).join('');
  recalcPayable();
}
function _parseMoney(s) {
  if (!s) return 0;
  const num = parseFloat(s.replace(/[^\d.-]/g, ''));
  return isNaN(num) ? 0 : num;
}
function _fmtMoney(num) {
  return num.toLocaleString('en-US',
    {minimumFractionDigits: 2, maximumFractionDigits: 2});
}
function recalcPayable() {
  const principal = _parseMoney(($('manualFieldGrid').querySelector('[data-manual="Principal_Amount"]') || {}).value);
  const interest  = _parseMoney(($('manualFieldGrid').querySelector('[data-manual="Interest"]')         || {}).value);
  const penalty   = _parseMoney(($('manualFieldGrid').querySelector('[data-manual="Penalty"]')          || {}).value);
  const payable   = $('manualFieldGrid').querySelector('[data-manual="Payable"]');
  if (payable) payable.value = _fmtMoney(principal + interest + penalty);
}
function fmtMoneyInput(inp) {
  if (inp.value.trim()) {
    const num = parseFloat(inp.value.replace(/[^\d.-]/g, ''));
    if (!isNaN(num)) inp.value = _fmtMoney(num);
  }
  recalcPayable();
  reloadPreview();
}
function collectManualFields() {
  const out = {};
  for (const inp of document.querySelectorAll('[data-manual]')) {
    const key = inp.dataset.manual;
    let v = inp.value;
    if (inp.type === 'date' && v) {
      const [y, m, d] = v.split('-');
      v = `${d}/${m}/${y}`;
    }
    if (v && v.trim() !== '') out[key] = v;
  }
  return out;
}
async function doGenerateOne() {
  const fields = {};
  for (const inp of document.querySelectorAll('[data-manual]')) {
    const key = inp.dataset.manual;
    let v = inp.value;
    if (inp.type === 'date' && v) {
      // <input type="date"> always yields ISO (YYYY-MM-DD); the notice
      // template wants DD/MM/YYYY. Convert client-side so every date
      // field (date, Due_date, disb_date, ...) gets the right format.
      const [y, m, d] = v.split('-');
      v = `${d}/${m}/${y}`;
    }
    if (v && v.trim() !== '') fields[key] = v;
  }
  const btn = $('genOneBtn');
  btn.disabled = true;
  btn.innerHTML = '<span class="spinner"></span>Rendering…';
  flash($('genOneStatus'), '', '');
  try {
    const r = await fetch('/generate_one', {
      method: 'POST', headers: {'Content-Type': 'application/json'},
      body: JSON.stringify({
        template_slug: activeSlug,
        fields,
      }),
    });
    if (!r.ok) {
      const err = await r.json().catch(() => ({error: 'server error'}));
      throw new Error(err.error || 'failed');
    }
    const blob = await r.blob();
    const url = URL.createObjectURL(blob);
    const a = document.createElement('a');
    // Keep Unicode letters (Chinese etc.) in the download filename;
    // only strip filesystem-unsafe punctuation. \w is ASCII-only in JS.
    const raw = (fields.name || 'notice').replace(/[\\\/:*?"<>|\r\n\t]+/g, '_').slice(0, 60);
    a.href = url; a.download = (raw || 'notice') + '.pdf';
    document.body.appendChild(a); a.click(); a.remove();
    URL.revokeObjectURL(url);
    flash($('genOneStatus'), 'Downloaded ' + a.download + '.', 'ok');
  } catch (e) {
    flash($('genOneStatus'), 'Failed: ' + e.message, 'err');
  } finally {
    btn.disabled = false;
    btn.textContent = 'Generate single PDF';
  }
}

// ── boot ─────────────────────────────────────────
(async () => {
  await loadTemplateList();
  await loadActiveTemplate();
  wireStaticLivePreview();
})();
</script>
</body>
</html>
"""


VERIFY_TEMPLATE = r"""
<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>Notice Verification — S&amp;S Law Firm</title>
<link rel="icon" type="image/png" href="__LOGO_DATA_URI__">
<style>
  :root { --primary:#0b1220; --accent:#c9a14a; --bg:#f4f1ea;
          --panel:#fff; --border:#e0dcd2; --ok:#16a34a; --err:#dc2626; }
  * { box-sizing: border-box; margin:0; padding:0; }
  body { font-family:-apple-system,"Segoe UI",Roboto,"PingFang SC",sans-serif;
         background:var(--bg); color:#111; padding:28px 20px; }
  .wrap { max-width:900px; margin:0 auto; }
  header { display:flex; align-items:center; gap:14px; margin-bottom:22px; }
  header img { width:44px; height:44px; border-radius:50%; }
  header h1 { font-size:1.2rem; font-weight:700; color:var(--primary); }
  header .sub { color:#666; font-size:.8rem; }
  header nav { margin-left:auto; }
  header nav a { color:#555; text-decoration:none; font-size:.85rem;
                 padding:6px 12px; border:1px solid var(--border);
                 border-radius:5px; background:#fff; }
  header nav a:hover { background:#eee; }

  .card { background:var(--panel); border:1px solid var(--border);
          border-radius:8px; padding:20px; margin-bottom:20px; }
  .card h2 { font-size:.9rem; font-weight:700; color:var(--primary);
             text-transform:uppercase; letter-spacing:.08em;
             border-bottom:2px solid var(--accent); padding-bottom:8px;
             margin-bottom:14px; }

  .verify-form { display:flex; gap:10px; }
  .verify-form input {
    flex:1; padding:11px 14px; font-size:1rem;
    font-family:ui-monospace,"SF Mono","Menlo",monospace;
    letter-spacing:.08em; text-transform:uppercase;
    border:1px solid var(--border); border-radius:5px; background:#fcfaf4;
  }
  .verify-form button {
    padding:11px 22px; background:var(--primary); color:#fff;
    border:none; border-radius:5px; font-size:.9rem; font-weight:600;
    cursor:pointer;
  }
  .verify-form button:hover { background:#1a2541; }

  .result { margin-top:14px; padding:14px 16px; border-radius:6px;
            font-size:.88rem; line-height:1.55; }
  .result-ok  { background:#eaf7ee; border:1px solid #b9e1c3; color:#14532d; }
  .result-err { background:#fdecec; border:1px solid #f3b7b7; color:#7f1d1d; }
  .result .label { color:#6b7280; font-size:.7rem; letter-spacing:.1em;
                   text-transform:uppercase; margin-right:8px; }
  .result .kv { margin:3px 0; }

  table.hist { width:100%; border-collapse:collapse; font-size:.82rem; }
  table.hist th { text-align:left; padding:8px 10px;
                  background:#f8f6f0; color:#555; font-weight:600;
                  font-size:.72rem; text-transform:uppercase;
                  letter-spacing:.06em; border-bottom:2px solid var(--border); }
  table.hist td { padding:8px 10px; border-bottom:1px solid #f0ede5;
                  vertical-align:top; }
  table.hist td.serial { font-family:ui-monospace,"SF Mono","Menlo",monospace;
                         font-size:.78rem; letter-spacing:.5pt; }
  table.hist td.money { font-variant-numeric:tabular-nums; text-align:right; }
  table.hist tr:hover td { background:#faf8f3; }
  .search-row { display:flex; gap:10px; margin-bottom:10px; }
  .search-row input {
    flex:1; padding:8px 12px; font-size:.85rem;
    border:1px solid var(--border); border-radius:5px; background:#fff;
  }
  .pager { margin-top:10px; display:flex; justify-content:space-between;
           align-items:center; font-size:.8rem; color:#666; }
  .pager button {
    padding:5px 12px; border:1px solid var(--border); background:#fff;
    border-radius:4px; cursor:pointer; font-size:.8rem;
  }
  .pager button:disabled { opacity:.4; cursor:not-allowed; }
  .muted { color:#888; font-style:italic; }
</style>
</head>
<body>
<div class="wrap">
  <header>
    <img src="__LOGO_DATA_URI__" alt="logo">
    <div>
      <h1>Notice Verification</h1>
      <div class="sub">S&amp;S Law Firm — verify authenticity by serial</div>
    </div>
    <nav><a href="/">← Workstation</a></nav>
  </header>

  <div class="card">
    <h2>🔍 Verify by Serial</h2>
    <div class="verify-form">
      <input id="serial" placeholder="XXXX-XXXX-XXXX-XXXX"
             autocomplete="off" autofocus>
      <button onclick="doVerify()">Verify</button>
    </div>
    <div id="result"></div>
  </div>

  <div class="card">
    <h2>📋 Recent History</h2>
    <div class="search-row">
      <input id="q" placeholder="Filter by name...">
      <button onclick="loadHistory(0)" style="padding:8px 16px; background:var(--primary); color:#fff; border:none; border-radius:5px; cursor:pointer;">Search</button>
    </div>
    <table class="hist">
      <thead>
        <tr><th>Date</th><th>Time</th><th>Name</th>
            <th style="text-align:right;">Principal</th>
            <th>Serial</th></tr>
      </thead>
      <tbody id="histBody"></tbody>
    </table>
    <div class="pager">
      <div id="histTotal"></div>
      <div>
        <button id="prevBtn" onclick="loadHistory(offset - limit)">← Prev</button>
        <button id="nextBtn" onclick="loadHistory(offset + limit)">Next →</button>
      </div>
    </div>
  </div>
</div>

<script>
const $ = id => document.getElementById(id);
let limit = 50, offset = 0, total = 0;

async function doVerify() {
  const raw = $('serial').value.trim().toUpperCase();
  const out = $('result');
  if (!raw) { out.innerHTML = ''; return; }
  try {
    const r = await fetch('/api/verify?serial=' + encodeURIComponent(raw));
    const d = await r.json();
    if (r.ok && d.ok) {
      out.className = 'result result-ok';
      out.innerHTML = `
        <div class="kv"><span class="label">AUTHENTIC</span>
          This notice was issued by S&amp;S Law Firm.</div>
        <div class="kv"><span class="label">Respondent</span><strong>${escape(d.name)}</strong></div>
        <div class="kv"><span class="label">Principal</span>${escape(d.principal) || '—'}</div>
        <div class="kv"><span class="label">Issued</span>${escape(d.generated_at)}</div>`;
    } else {
      out.className = 'result result-err';
      out.innerHTML = `<strong>Not found.</strong> No notice in our records matches this serial.
        The letter may be forged, or was issued outside this system.`;
    }
  } catch (e) {
    out.className = 'result result-err';
    out.innerHTML = 'Verification failed: ' + escape(e.message);
  }
}
function escape(s) {
  return String(s == null ? '' : s)
    .replace(/&/g, '&amp;').replace(/</g, '&lt;')
    .replace(/>/g, '&gt;').replace(/"/g, '&quot;');
}

async function loadHistory(off) {
  offset = Math.max(0, off);
  const q = $('q').value.trim();
  const qs = new URLSearchParams({limit, offset});
  if (q) qs.set('q', q);
  const r = await fetch('/api/history?' + qs.toString());
  const d = await r.json();
  total = d.total || 0;
  const body = $('histBody');
  if (!d.rows || d.rows.length === 0) {
    body.innerHTML = '<tr><td colspan="5" class="muted">No records.</td></tr>';
  } else {
    body.innerHTML = d.rows.map(r => {
      const [date, time] = (r.generated_at || '').split('T');
      return `<tr>
        <td>${escape(date)}</td>
        <td>${escape((time || '').slice(0, 8))}</td>
        <td>${escape(r.name)}</td>
        <td class="money">${escape(r.principal)}</td>
        <td class="serial">${escape(r.serial)}</td>
      </tr>`;
    }).join('');
  }
  $('histTotal').textContent = `${total} total · showing ${offset + 1}–${Math.min(offset + limit, total)}`;
  $('prevBtn').disabled = offset <= 0;
  $('nextBtn').disabled = offset + limit >= total;
}

$('serial').addEventListener('keydown', e => { if (e.key === 'Enter') doVerify(); });
$('q').addEventListener('keydown', e => { if (e.key === 'Enter') loadHistory(0); });
loadHistory(0);
</script>
</body>
</html>
"""


INVENTORY_TEMPLATE = r"""
<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>Inventory — S&amp;S Law Firm</title>
<link rel="icon" type="image/png" href="__LOGO_DATA_URI__">
<style>
  :root { --primary:#0b1220; --accent:#c9a14a; --bg:#f4f1ea;
          --panel:#fff; --border:#e0dcd2; --ok:#16a34a; --err:#dc2626;
          --muted:#6b7280; }
  * { box-sizing: border-box; margin:0; padding:0; }
  body { font-family:-apple-system,"Segoe UI",Roboto,"PingFang SC",sans-serif;
         background:var(--bg); color:#111; padding:20px; }
  .wrap { max-width:1200px; margin:0 auto; padding-bottom:120px; }
  header { display:flex; align-items:center; gap:14px; margin-bottom:18px; }
  header img { width:44px; height:44px; border-radius:50%; }
  header h1 { font-size:1.2rem; font-weight:700; color:var(--primary); }
  header .sub { color:#666; font-size:.8rem; }
  header nav { margin-left:auto; display:flex; gap:8px; }
  header nav a { color:#555; text-decoration:none; font-size:.85rem;
                 padding:6px 12px; border:1px solid var(--border);
                 border-radius:5px; background:#fff; }
  header nav a:hover { background:#eee; }

  .card { background:var(--panel); border:1px solid var(--border);
          border-radius:8px; padding:18px 20px; margin-bottom:16px; }
  .card h2 { font-size:.85rem; font-weight:700; color:var(--primary);
             text-transform:uppercase; letter-spacing:.08em;
             border-bottom:2px solid var(--accent); padding-bottom:7px;
             margin-bottom:14px; display:flex; align-items:center; gap:8px; }
  .card h2 .count { margin-left:auto; color:var(--muted);
                    font-size:.72rem; font-weight:500;
                    letter-spacing:.02em; text-transform:none; }

  .btn { padding:8px 16px; border:1px solid var(--border); background:#fff;
         border-radius:5px; cursor:pointer; font-size:.85rem; color:#333; }
  .btn:hover:not(:disabled) { background:#f0ede5; }
  .btn:disabled { opacity:.45; cursor:not-allowed; }
  .btn-primary { background:var(--primary); color:#fff; border-color:var(--primary); }
  .btn-primary:hover:not(:disabled) { background:#1a2541; }
  .btn-accent { background:var(--accent); color:#0b1220; border-color:#b08b39;
                font-weight:600; }
  .btn-accent:hover:not(:disabled) { background:#d4af5f; }
  .btn-danger { color:#7f1d1d; border-color:#f3b7b7; background:#fdecec; }
  .btn-danger:hover:not(:disabled) { background:#fbd8d8; }
  .btn-sm { padding:4px 10px; font-size:.78rem; }

  input[type=text], input[type=date], input[type=number], select {
    padding:7px 10px; border:1px solid var(--border); border-radius:5px;
    background:#fff; font-size:.88rem; color:#111;
  }
  input[type=file] { font-size:.85rem; }

  .row { display:flex; gap:10px; align-items:center; flex-wrap:wrap; }
  .grow { flex:1; min-width:160px; }

  .import-summary { display:grid; grid-template-columns:repeat(4, 1fr);
                    gap:10px; margin-top:12px; }
  .stat { background:#faf8f3; border:1px solid var(--border);
          border-radius:6px; padding:10px 12px; }
  .stat .k { font-size:.68rem; text-transform:uppercase; color:var(--muted);
             letter-spacing:.08em; }
  .stat .v { font-size:1.3rem; font-weight:700; color:var(--primary);
             margin-top:3px; font-variant-numeric:tabular-nums; }
  .stat.warn { background:#fff8e6; border-color:#f5d58c; }
  .stat.warn .v { color:#8b6914; }

  .policy-row { display:flex; gap:18px; align-items:center;
                margin-top:14px; padding:12px 14px;
                background:#faf8f3; border-radius:6px;
                border:1px solid var(--border); }
  .policy-row label { display:flex; align-items:center; gap:6px;
                      font-size:.88rem; cursor:pointer; }

  .progress { margin-top:10px; height:8px; background:#eee;
              border-radius:4px; overflow:hidden; }
  .progress-bar { height:100%; background:var(--accent); width:0%;
                  transition:width .25s; }
  .progress-msg { font-size:.78rem; color:#555; margin-top:6px; }
  .progress-msg.err { color:var(--err); }
  .progress-msg.ok { color:var(--ok); }

  table.data { width:100%; border-collapse:collapse; font-size:.84rem;
               margin-top:8px; }
  table.data th { text-align:left; padding:8px 10px; background:#f8f6f0;
                  color:#555; font-weight:600; font-size:.72rem;
                  text-transform:uppercase; letter-spacing:.06em;
                  border-bottom:2px solid var(--border); }
  table.data td { padding:7px 10px; border-bottom:1px solid #f0ede5;
                  vertical-align:middle; }
  table.data tr:hover td { background:#faf8f3; }
  table.data td.mono { font-family:ui-monospace,"SF Mono",monospace;
                       font-size:.78rem; letter-spacing:.2pt; }
  table.data .check { width:30px; text-align:center; }
  table.data tr.selected td { background:#fffbe8; }

  .pager { margin-top:10px; display:flex; justify-content:space-between;
           align-items:center; font-size:.8rem; color:#666; }
  .pager .group { display:flex; gap:6px; align-items:center; }
  .muted { color:var(--muted); font-style:italic; }

  /* Sticky bottom action bar */
  .action-bar { position:fixed; bottom:0; left:0; right:0;
                background:rgba(11,18,32,.98); color:#fff;
                padding:12px 24px; display:flex; gap:14px;
                align-items:center; box-shadow:0 -4px 18px rgba(0,0,0,.2);
                z-index:50; }
  .action-bar .selcount { font-size:.95rem; font-weight:600; }
  .action-bar .selcount .n { color:var(--accent); font-size:1.1rem;
                             margin:0 4px; }
  .action-bar .spacer { flex:1; }
  .action-bar input, .action-bar select {
    background:rgba(255,255,255,.12); color:#fff;
    border-color:rgba(255,255,255,.2);
  }
  .action-bar input::placeholder { color:rgba(255,255,255,.5); }
  .action-bar label { font-size:.78rem; color:#cbd5e1; }

  .downloads { margin-top:10px; display:flex; flex-wrap:wrap; gap:6px; }
  .downloads a { display:inline-block; padding:5px 12px;
                 background:#eaf7ee; border:1px solid #b9e1c3;
                 color:#14532d; border-radius:4px; font-size:.82rem;
                 text-decoration:none; }
  .downloads a:hover { background:#d6f0dd; }

  .conflicts-sample { font-family:ui-monospace,monospace; font-size:.78rem;
                      color:#7f1d1d; background:#fdecec; padding:8px 10px;
                      border-radius:5px; margin-top:8px;
                      max-height:120px; overflow-y:auto; }
  .conflicts-sample code { display:inline-block; margin-right:6px; }

  .hidden { display:none !important; }

  details.advanced summary { cursor:pointer; font-size:.82rem;
                             color:var(--muted); margin-top:8px; }
  details.advanced[open] summary { margin-bottom:8px; }

  /* Search card additions */
  .search-hint { font-size:.75rem; color:var(--muted);
                 margin-top:6px; padding-left:2px; }
  details.bulk-box { margin-top:12px; border:1px dashed var(--border);
                     border-radius:6px; padding:8px 12px;
                     background:#fbfaf5; }
  details.bulk-box summary { cursor:pointer; font-size:.85rem;
                             color:var(--primary); font-weight:600;
                             padding:4px 2px; }
  details.bulk-box[open] summary { margin-bottom:6px; }
  .bulk-body { padding:6px 2px 4px; }
  .bulk-body .bulk-lbl { display:block; font-size:.74rem;
                         color:#555; margin-bottom:4px;
                         text-transform:uppercase; letter-spacing:.06em; }
  .bulk-body textarea { width:100%; padding:8px 10px; resize:vertical;
                        border:1px solid var(--border); border-radius:5px;
                        background:#fff; font-size:.82rem;
                        font-family:ui-monospace,"SF Mono",monospace; }
  .bulk-body .bulk-muted { font-size:.72rem; color:var(--muted);
                           margin-top:4px; line-height:1.35; }
  .bulk-status { font-size:.82rem; color:var(--muted); }
  .bulk-status.ok  { color:var(--ok); }
  .bulk-status.err { color:var(--err); }
  .bulk-missing { margin-top:10px; padding:10px 12px;
                  background:#fff8e6; border:1px solid #f5d58c;
                  border-radius:5px; font-size:.78rem;
                  line-height:1.5; color:#8b6914; }
  .bulk-missing .ids { font-family:ui-monospace,monospace;
                       font-size:.74rem; max-height:120px;
                       overflow-y:auto; display:block;
                       margin-top:6px; word-break:break-all; }

  /* Selected card */
  .selected-scroll { max-height:420px; overflow-y:auto;
                     border:1px solid var(--border); border-radius:5px; }
  .selected-scroll table { margin-top:0; }
  .selected-scroll thead th { position:sticky; top:0; z-index:1; }
  #selectedTable td.del { text-align:center; }
  #selectedTable td.del button {
    border:none; background:transparent; cursor:pointer;
    font-size:.9rem; color:var(--muted); padding:2px 6px;
    border-radius:3px;
  }
  #selectedTable td.del button:hover {
    background:#fdecec; color:var(--err);
  }

  /* Inline-editable name cell in search results */
  td.name-cell { padding:6px 10px; transition:background .25s; }
  td.name-cell .name-edit {
    cursor:text; padding:2px 6px; border-radius:3px;
    transition:background .15s, outline .15s; min-height:1.2em;
    display:inline-block;
  }
  td.name-cell .name-edit:hover { background:#fcf8e8; }
  td.name-cell .name-edit:focus { outline:2px solid var(--accent);
                                  outline-offset:-2px;
                                  background:#fffbe8; }
  td.name-cell .name-edit::after { content:"✎"; color:var(--muted);
                                   margin-left:6px; font-size:.75em;
                                   opacity:0; }
  tr:hover td.name-cell .name-edit::after { opacity:.5; }
  td.name-cell .name-edit:focus::after { opacity:0; }
  td.name-cell.saving  .name-edit { background:#fff8e6; }
  td.name-cell.saved   .name-edit { background:#eaf7ee; }
  td.name-cell.save-err .name-edit { background:#fdecec; color:var(--err); }
  .name-orig-hint { display:none; font-size:.7rem; color:var(--muted);
                    font-style:italic; margin-top:2px; line-height:1.3; }
  td.name-cell.edited .name-orig-hint { display:block; }
  td.name-cell.edited .name-edit { border-bottom:1.5px dotted var(--accent); }
</style>
</head>
<body>
<div class="wrap">
  <header>
    <img src="__LOGO_DATA_URI__" alt="logo">
    <div>
      <h1>Inventory Mode</h1>
      <div class="sub">Import cases once · select · generate on demand</div>
    </div>
    <nav>
      <a href="/">← Workstation</a>
      <a href="/verify">🔍 Verify</a>
      <a href="/logout">Sign Out</a>
    </nav>
  </header>

  <!-- ═══ 1 · IMPORT ═══ -->
  <div class="card">
    <h2>📥 Import Case Excel
      <span class="count" id="totalCount">—</span>
    </h2>
    <div class="row">
      <input type="file" id="excelFile" accept=".xlsx">
      <button class="btn btn-primary" id="previewBtn">Preview Import</button>
      <span class="spacer grow"></span>
      <button class="btn btn-danger btn-sm" id="wipeBtn">Clear Inventory</button>
    </div>

    <!-- Preview result (shown after /preview returns) -->
    <div id="previewBox" class="hidden" style="margin-top:14px;">
      <div class="import-summary">
        <div class="stat">
          <div class="k">Rows in file</div>
          <div class="v" id="pvTotal">0</div>
        </div>
        <div class="stat">
          <div class="k">New (will insert)</div>
          <div class="v" id="pvNew">0</div>
        </div>
        <div class="stat warn">
          <div class="k">Conflicts (order_id exists)</div>
          <div class="v" id="pvConflict">0</div>
        </div>
        <div class="stat warn">
          <div class="k">Dup within file</div>
          <div class="v" id="pvDupFile">0</div>
        </div>
      </div>
      <div id="pvDetected" class="muted" style="margin-top:8px; font-size:.8rem;"></div>
      <div id="pvConflictSample" class="conflicts-sample hidden"></div>

      <div class="policy-row" id="policyRow">
        <strong style="font-size:.85rem;">If order_id conflicts:</strong>
        <label><input type="radio" name="policy" value="skip" checked>
               Skip (keep existing)</label>
        <label><input type="radio" name="policy" value="overwrite">
               Overwrite with new row</label>
        <span class="spacer grow"></span>
        <button class="btn btn-accent" id="commitBtn">Commit Import</button>
      </div>

      <div id="importProgressBox" class="hidden" style="margin-top:12px;">
        <div class="progress"><div class="progress-bar" id="importBar"></div></div>
        <div class="progress-msg" id="importMsg">…</div>
      </div>
    </div>
  </div>

  <!-- ═══ 2 · SEARCH ═══ -->
  <div class="card">
    <h2>🔍 Search &amp; Select
      <span class="count" id="searchModeBadge"></span>
    </h2>
    <div class="row">
      <label style="font-size:.82rem; color:#555;">Field</label>
      <select id="searchField">
        <option value="cnic">CNIC</option>
        <option value="name">Name</option>
        <option value="phone">Phone</option>
      </select>
      <input type="text" id="searchQ" class="grow"
             placeholder="single prefix (e.g. '42301')  OR  many values separated by comma / space / newline">
      <button class="btn btn-primary" id="searchBtn">Search</button>
      <span class="spacer grow"></span>
      <button class="btn btn-sm" id="selectPageBtn">Select this page</button>
      <button class="btn btn-sm" id="selectAllBtn">Select all matching</button>
    </div>
    <div class="search-hint">
      One value → prefix match · Many values → exact match (paste a list of CNICs / phones)
    </div>

    <!-- Bulk select via textarea or Excel file upload -->
    <details class="bulk-box">
      <summary>📎 Bulk select from file or big paste</summary>
      <div class="bulk-body">
        <div class="row" style="align-items:flex-start;">
          <div class="grow" style="min-width:280px;">
            <label class="bulk-lbl">Paste values (one per line, or comma/space separated)</label>
            <textarea id="bulkText" rows="4"
              placeholder="4230113751365&#10;4420551500363&#10;4220169696965&#10;&#10;(uses the Field selector above)"></textarea>
          </div>
          <div style="min-width:220px;">
            <label class="bulk-lbl">…or upload an Excel</label>
            <input type="file" id="bulkFile" accept=".xlsx">
            <div class="bulk-muted">
              Looks for a column matching the chosen Field; falls back
              to the first column.
            </div>
          </div>
        </div>
        <div class="row" style="margin-top:10px;">
          <button class="btn btn-accent" id="bulkMatchBtn">Match &amp; add to selection</button>
          <span class="bulk-status" id="bulkStatus"></span>
        </div>
        <div class="bulk-missing hidden" id="bulkMissing"></div>
      </div>
    </details>

    <table class="data" id="resultTable">
      <thead>
        <tr>
          <th class="check"><input type="checkbox" id="headerCheck"
                                   title="Toggle this page"></th>
          <th>Order ID</th>
          <th>Name</th>
          <th>CNIC</th>
          <th>Phone</th>
          <th>Imported</th>
        </tr>
      </thead>
      <tbody id="resultBody">
        <tr><td colspan="6" class="muted">Import an Excel above, then search.</td></tr>
      </tbody>
    </table>

    <div class="pager">
      <div id="pagerInfo">—</div>
      <div class="group">
        <button class="btn btn-sm" id="prevPage">← Prev</button>
        <span id="pageLabel">1</span>
        <button class="btn btn-sm" id="nextPage">Next →</button>
      </div>
    </div>
  </div>

  <!-- ═══ 2.5 · SELECTED CASES ═══ -->
  <div class="card" id="selectedCard">
    <h2>🧺 Selected Cases
      <span class="count" id="selectedSummary">0 selected</span>
    </h2>
    <div class="row" style="margin-bottom:8px;">
      <button class="btn btn-sm" id="selClearBtn2">Clear all</button>
      <button class="btn btn-sm" id="selDedupeNameBtn" title="Keep one row per name">
        Dedupe by name
      </button>
      <button class="btn btn-sm" id="selDedupeCnicBtn" title="Keep one row per CNIC">
        Dedupe by CNIC
      </button>
      <span class="spacer grow"></span>
      <span class="muted" id="selTruncateNote"></span>
    </div>
    <div class="selected-scroll">
      <table class="data" id="selectedTable">
        <thead>
          <tr>
            <th style="width:34px;"></th>
            <th>Order ID</th>
            <th>Name</th>
            <th>CNIC</th>
            <th>Phone</th>
          </tr>
        </thead>
        <tbody id="selectedBody">
          <tr><td colspan="5" class="muted">Nothing selected yet.</td></tr>
        </tbody>
      </table>
    </div>
  </div>

  <!-- ═══ 3 · GENERATE ═══ -->
  <div class="card">
    <h2>📝 Generate Notices</h2>
    <div class="row">
      <label style="font-size:.82rem; color:#555;">Template</label>
      <select id="tplSelect" class="grow" style="max-width:260px;"></select>

      <label style="font-size:.82rem; color:#555;">Date</label>
      <input type="date" id="dateInput">

      <label style="font-size:.82rem; color:#555;">Profile</label>
      <select id="profileSelect">
        <option value="vps">VPS (conservative)</option>
        <option value="mac">Mac (max)</option>
      </select>
    </div>

    <details class="advanced">
      <summary>Advanced</summary>
      <div class="row" style="margin-top:6px;">
        <label style="font-size:.82rem; color:#555;">Filename fields</label>
        <input type="text" id="fnFields" class="grow"
               placeholder="comma-separated, e.g. name,cnic"
               value="name,cnic">
        <label style="font-size:.82rem; color:#555;">Group by</label>
        <input type="text" id="groupBy" placeholder="column (default: 负责人 · type __none__ to skip)">
      </div>
    </details>

    <div id="genProgressBox" class="hidden" style="margin-top:14px;">
      <div class="progress"><div class="progress-bar" id="genBar"></div></div>
      <div class="progress-msg" id="genMsg">…</div>
      <div class="downloads" id="genDownloads"></div>
    </div>
  </div>
</div>

<!-- Sticky selection / action bar -->
<div class="action-bar">
  <div class="selcount">
    <span class="n" id="selN">0</span> selected
  </div>
  <button class="btn btn-sm" id="clearSelBtn">Clear</button>
  <span class="spacer"></span>
  <button class="btn btn-accent" id="generateBtn" disabled>Generate Notices</button>
</div>

<script>
const $ = id => document.getElementById(id);
const H = s => String(s == null ? '' : s)
    .replace(/&/g,'&amp;').replace(/</g,'&lt;')
    .replace(/>/g,'&gt;').replace(/"/g,'&quot;');

// Upper bound on rows rendered in the Selected panel — we keep the
// authoritative Set small in memory but never DOM-render more than
// this (huge <tbody>s pin the tab for seconds on older laptops).
const MAX_SEL_DISPLAY = 2000;

const state = {
  selected: new Set(),              // order_id (the authoritative set)
  meta:     new Map(),              // order_id → {order_id,name,cnic,phone}
  search:   { field:'cnic', q:'', page:1, limit:100, total:0, mode:'all' },
  lastRows: [],
  importToken: null,
  importTaskId: null,
  genTaskId: null,
};

// ── Bootstrap ───────────────────────────────────────────────
async function boot() {
  await Promise.all([refreshCount(), loadTemplates()]);
  $('dateInput').value = new Date().toISOString().slice(0,10);
  renderSelectedPanel();
}

async function refreshCount() {
  try {
    const r = await fetch('/api/cases/stats');
    const d = await r.json();
    $('totalCount').textContent =
      `${(d.count || 0).toLocaleString()} case(s) in inventory`;
  } catch (e) { /* quiet */ }
}

async function loadTemplates() {
  const r = await fetch('/api/templates');
  const list = await r.json();
  const sel = $('tplSelect');
  sel.innerHTML = list.map(t =>
    `<option value="${H(t.slug)}">${H(t.name)}${t.builtin?' (default)':''}</option>`
  ).join('');
}

// ── IMPORT ──────────────────────────────────────────────────
$('previewBtn').onclick = async () => {
  const f = $('excelFile').files[0];
  if (!f) { alert('Pick an Excel file first.'); return; }
  const fd = new FormData();
  fd.append('excel', f);
  $('previewBtn').disabled = true;
  $('previewBtn').textContent = 'Analyzing…';
  try {
    const r = await fetch('/api/cases/import/preview',
                          { method:'POST', body: fd });
    const d = await r.json();
    if (!r.ok) throw new Error(d.error || 'preview failed');
    state.importToken = d.token;
    $('pvTotal').textContent    = (d.total || 0).toLocaleString();
    $('pvNew').textContent      = (d.new_count || 0).toLocaleString();
    $('pvConflict').textContent = (d.conflict_count || 0).toLocaleString();
    $('pvDupFile').textContent  = (d.dup_in_file_count || 0).toLocaleString();
    $('pvDetected').innerHTML = 'Detected columns — '
      + `order_id: <code>${H(d.order_id_col)}</code>, `
      + `name: <code>${H(d.name_col || '—')}</code>, `
      + `cnic: <code>${H(d.cnic_col || '—')}</code>, `
      + `phone: <code>${H(d.phone_col || '—')}</code>`;
    const cs = $('pvConflictSample');
    if ((d.conflicts_sample || []).length) {
      cs.innerHTML = '<strong>Conflicting order_ids (first '
        + d.conflicts_sample.length + '):</strong> '
        + d.conflicts_sample.map(x => `<code>${H(x)}</code>`).join(' ');
      cs.classList.remove('hidden');
    } else { cs.classList.add('hidden'); }
    // If no conflicts, the policy radio is moot — still show it for
    // consistency, preselected to "skip" (which acts as a no-op).
    $('previewBox').classList.remove('hidden');
    $('importProgressBox').classList.add('hidden');
  } catch (e) {
    alert('Preview failed: ' + e.message);
  } finally {
    $('previewBtn').disabled = false;
    $('previewBtn').textContent = 'Preview Import';
  }
};

$('commitBtn').onclick = async () => {
  if (!state.importToken) { alert('No pending import.'); return; }
  const policy = document.querySelector('input[name=policy]:checked').value;
  $('commitBtn').disabled = true;
  try {
    const r = await fetch('/api/cases/import/commit', {
      method:'POST',
      headers:{'Content-Type':'application/json'},
      body: JSON.stringify({ token: state.importToken, policy }),
    });
    const d = await r.json();
    if (!r.ok) throw new Error(d.error || 'commit failed');
    state.importTaskId = d.task_id;
    $('importProgressBox').classList.remove('hidden');
    pollImport();
  } catch (e) {
    alert('Commit failed: ' + e.message);
    $('commitBtn').disabled = false;
  }
};

async function pollImport() {
  const tid = state.importTaskId;
  if (!tid) return;
  try {
    const r = await fetch('/status/' + tid);
    const d = await r.json();
    const pct = d.total ? Math.round((d.progress / d.total) * 100) : 0;
    $('importBar').style.width = pct + '%';
    $('importMsg').textContent = d.message || d.stage || '';
    $('importMsg').className = 'progress-msg';
    if (d.status === 'done') {
      $('importMsg').className = 'progress-msg ok';
      $('commitBtn').disabled = false;
      state.importToken = null;
      await refreshCount();
      return;
    }
    if (d.status === 'error') {
      $('importMsg').className = 'progress-msg err';
      $('importMsg').textContent = d.error || d.message || 'failed';
      $('commitBtn').disabled = false;
      return;
    }
    setTimeout(pollImport, 800);
  } catch (e) {
    $('importMsg').className = 'progress-msg err';
    $('importMsg').textContent = 'Polling failed: ' + e.message;
  }
}

$('wipeBtn').onclick = async () => {
  if (!confirm('Delete ALL cases from inventory? '
             + 'You will need to re-import. This cannot be undone.'))
    return;
  const r = await fetch('/api/cases/delete', {
    method:'POST',
    headers:{'Content-Type':'application/json'},
    body: JSON.stringify({ all: true })
  });
  if (r.ok) {
    state.selected.clear();
    state.meta.clear();
    state.lastRows = [];
    $('resultBody').innerHTML =
      '<tr><td colspan="6" class="muted">Inventory cleared.</td></tr>';
    await refreshCount();
    onSelectionChanged();
  } else { alert('Delete failed.'); }
};

// ── SEARCH ──────────────────────────────────────────────────
async function runSearch(page) {
  state.search.field = $('searchField').value;
  state.search.q     = $('searchQ').value.trim();
  state.search.page  = Math.max(1, page || 1);
  const body = {
    field: state.search.field, q: state.search.q,
    page: state.search.page, limit: state.search.limit,
  };
  let r;
  // Multi-line / very-long paste → must go via POST (URL limit ~4 KB).
  if (state.search.q.length > 3000) {
    r = await fetch('/api/cases/search', {
      method:'POST',
      headers:{'Content-Type':'application/json'},
      body: JSON.stringify(body),
    });
  } else {
    const qs = new URLSearchParams({
      field: body.field, page: body.page, limit: body.limit,
    });
    if (body.q) qs.set('q', body.q);
    r = await fetch('/api/cases/search?' + qs.toString());
  }
  const d = await r.json();
  if (!r.ok) { alert('Search failed: ' + (d.error||'?')); return; }
  state.search.total = d.total;
  state.search.mode  = d.mode || 'all';
  state.lastRows = d.rows || [];
  // Opportunistic meta enrichment — every row the user sees is cached
  // so the Selected panel can draw without a follow-up fetch.
  state.lastRows.forEach(r => state.meta.set(r.order_id, r));
  renderResults();
  renderPager();
  renderSearchBadge(d);
}

function renderSearchBadge(d) {
  const el = $('searchModeBadge');
  if (!el) return;
  if (d.mode === 'exact' && d.values_parsed > 1) {
    el.textContent = `exact match · ${d.values_parsed} values`;
  } else if (d.mode === 'prefix') {
    el.textContent = `prefix match`;
  } else {
    el.textContent = '';
  }
}

function renderResults() {
  const rows = state.lastRows;
  if (!rows.length) {
    $('resultBody').innerHTML =
      '<tr><td colspan="6" class="muted">No matches.</td></tr>';
    $('headerCheck').checked = false;
    return;
  }
  $('resultBody').innerHTML = rows.map(r => {
    const checked = state.selected.has(r.order_id) ? 'checked' : '';
    const cls = checked ? ' class="selected"' : '';
    const curr = r.name || '';
    const orig = (r.original_name == null) ? curr : r.original_name;
    const editedCls = (curr !== orig) ? ' edited' : '';
    return `<tr${cls} data-oid="${H(r.order_id)}">
      <td class="check">
        <input type="checkbox" class="rowck" ${checked}
               data-oid="${H(r.order_id)}"></td>
      <td class="mono">${H(r.order_id)}</td>
      <td class="name-cell${editedCls}" data-oid="${H(r.order_id)}"
          data-orig-name="${H(orig)}">
        <span class="name-edit" contenteditable="true" spellcheck="false"
              data-last="${H(curr)}"
              title="Click to edit · Enter to save · Esc to cancel"
              >${H(curr)}</span>
        <div class="name-orig-hint">orig: ${H(orig)}</div>
      </td>
      <td class="mono">${H(r.cnic)}</td>
      <td class="mono">${H(r.phone)}</td>
      <td class="mono" style="color:#888; font-size:.75rem;">${H(r.created_at||'').slice(0,10)}</td>
    </tr>`;
  }).join('');
  document.querySelectorAll('.rowck').forEach(cb => {
    cb.addEventListener('change', () => {
      const oid = cb.dataset.oid;
      if (cb.checked) state.selected.add(oid);
      else state.selected.delete(oid);
      cb.closest('tr').classList.toggle('selected', cb.checked);
      updateHeaderCheck();
      onSelectionChanged();
    });
  });
  // Inline name editor — save on blur, Enter commits, Esc reverts.
  document.querySelectorAll('.name-edit').forEach(edit => {
    edit.addEventListener('keydown', e => {
      if (e.key === 'Enter') { e.preventDefault(); edit.blur(); }
      else if (e.key === 'Escape') {
        edit.textContent = edit.dataset.last;
        edit.blur();
      }
    });
    edit.addEventListener('blur', () => saveNameEdit(edit));
  });
  updateHeaderCheck();
}

async function saveNameEdit(edit) {
  const td   = edit.closest('td.name-cell');
  const oid  = td.dataset.oid;
  const orig = td.dataset.origName;   // import-time name (stable)
  const last = edit.dataset.last;     // last-saved name
  const next = edit.textContent.trim();
  if (next === last) return;
  td.classList.remove('saved', 'save-err');
  td.classList.add('saving');
  try {
    const r = await fetch('/api/cases/update_name', {
      method:'POST',
      headers:{'Content-Type':'application/json'},
      body: JSON.stringify({ order_id: oid, name: next }),
    });
    const d = await r.json();
    if (!r.ok) throw new Error(d.error || 'update failed');
    edit.dataset.last = next;
    td.classList.toggle('edited', next !== orig);
    td.querySelector('.name-orig-hint').textContent = 'orig: ' + orig;
    // Keep local caches in sync so the Selected panel + future
    // checkbox flips see the corrected name without a refetch.
    const meta = state.meta.get(oid);
    if (meta) {
      meta.name = next;
      if (meta.original_name == null) meta.original_name = orig;
    }
    const idx = state.lastRows.findIndex(x => x.order_id === oid);
    if (idx >= 0) {
      state.lastRows[idx].name = next;
      if (state.lastRows[idx].original_name == null) {
        state.lastRows[idx].original_name = orig;
      }
    }
    td.classList.remove('saving');
    td.classList.add('saved');
    setTimeout(() => td.classList.remove('saved'), 900);
    if (state.selected.has(oid)) renderSelectedPanel();
  } catch (e) {
    td.classList.remove('saving');
    td.classList.add('save-err');
    edit.textContent = last;
    setTimeout(() => td.classList.remove('save-err'), 1600);
  }
}

function renderPager() {
  const s = state.search;
  const pages = Math.max(1, Math.ceil(s.total / s.limit));
  const from = s.total ? (s.page-1)*s.limit + 1 : 0;
  const to = Math.min(s.page * s.limit, s.total);
  $('pagerInfo').textContent = s.total
    ? `${from.toLocaleString()}–${to.toLocaleString()} of ${s.total.toLocaleString()}`
    : '0 results';
  $('pageLabel').textContent = `${s.page} / ${pages}`;
  $('prevPage').disabled = s.page <= 1;
  $('nextPage').disabled = s.page >= pages;
}

function updateHeaderCheck() {
  const cks = [...document.querySelectorAll('.rowck')];
  const hc = $('headerCheck');
  if (!cks.length) { hc.checked = false; hc.indeterminate = false; return; }
  const n = cks.filter(c => c.checked).length;
  hc.checked = (n === cks.length);
  hc.indeterminate = n > 0 && n < cks.length;
}

$('headerCheck').onclick = () => {
  const on = $('headerCheck').checked;
  document.querySelectorAll('.rowck').forEach(cb => {
    cb.checked = on;
    const oid = cb.dataset.oid;
    if (on) state.selected.add(oid); else state.selected.delete(oid);
    cb.closest('tr').classList.toggle('selected', on);
  });
  onSelectionChanged();
};

$('searchBtn').onclick = () => runSearch(1);
$('searchQ').addEventListener('keydown', e => {
  // Shift-Enter = newline (for multi-value paste); Enter = search.
  if (e.key === 'Enter' && !e.shiftKey) { e.preventDefault(); runSearch(1); }
});
$('searchField').onchange = () => runSearch(1);
$('prevPage').onclick = () => runSearch(state.search.page - 1);
$('nextPage').onclick = () => runSearch(state.search.page + 1);

$('selectPageBtn').onclick = () => {
  state.lastRows.forEach(r => {
    state.selected.add(r.order_id);
    state.meta.set(r.order_id, r);
  });
  renderResults();
  onSelectionChanged();
};

$('selectAllBtn').onclick = async () => {
  const s = state.search;
  const body = { field: s.field, q: s.q, cap: 50000 };
  $('selectAllBtn').disabled = true;
  try {
    // POST so very-long multi-value `q` always fits.
    const r = await fetch('/api/cases/search/ids', {
      method:'POST',
      headers:{'Content-Type':'application/json'},
      body: JSON.stringify(body),
    });
    const d = await r.json();
    if (!r.ok) throw new Error(d.error || 'failed');
    d.ids.forEach(id => state.selected.add(id));
    if (d.capped) {
      alert('Selected first ' + d.returned.toLocaleString()
        + ' matching rows (of ' + d.total.toLocaleString()
        + '). Narrow the filter to reach the rest.');
    }
    renderResults();
    onSelectionChanged();
  } catch (e) { alert('Select-all failed: ' + e.message); }
  finally { $('selectAllBtn').disabled = false; }
};

$('clearSelBtn').onclick = clearSelection;
$('selClearBtn2').onclick = clearSelection;

function clearSelection() {
  state.selected.clear();
  renderResults();
  onSelectionChanged();
}

// ── BULK MATCH (paste or Excel upload) ──────────────────────
$('bulkMatchBtn').onclick = async () => {
  const text = $('bulkText').value.trim();
  const file = $('bulkFile').files[0];
  const field = $('searchField').value;
  if (!text && !file) {
    alert('Paste values or pick an Excel first.'); return;
  }
  $('bulkMatchBtn').disabled = true;
  $('bulkStatus').className = 'bulk-status';
  $('bulkStatus').textContent = 'Matching…';
  $('bulkMissing').classList.add('hidden');
  try {
    let d;
    if (file) {
      const fd = new FormData();
      fd.append('excel', file);
      fd.append('field', field);
      const r = await fetch('/api/cases/bulk_match',
                            { method:'POST', body: fd });
      d = await r.json();
      if (!r.ok) throw new Error(d.error || 'failed');
    } else {
      const values = text.split(/[\s,]+/).filter(Boolean);
      const r = await fetch('/api/cases/bulk_match', {
        method:'POST',
        headers:{'Content-Type':'application/json'},
        body: JSON.stringify({ field, values }),
      });
      d = await r.json();
      if (!r.ok) throw new Error(d.error || 'failed');
    }
    d.matched_ids.forEach(id => state.selected.add(id));
    (d.matched_meta || []).forEach(m => state.meta.set(m.order_id, m));
    $('bulkStatus').className = 'bulk-status ok';
    const srcStr = d.source_column
      ? ` (from column <code>${H(d.source_column)}</code>)` : '';
    $('bulkStatus').innerHTML =
      `✓ ${d.matched_count.toLocaleString()} of ${d.total_values.toLocaleString()} `
      + `added to selection${srcStr}`;
    if (d.missing_count > 0) {
      const sample = d.missing_sample.map(x => H(x)).join(' ');
      $('bulkMissing').classList.remove('hidden');
      $('bulkMissing').innerHTML =
        `<strong>${d.missing_count.toLocaleString()} value(s) not in inventory</strong>`
        + ` (sample):<span class="ids">${sample}</span>`;
    }
    onSelectionChanged();
  } catch (e) {
    $('bulkStatus').className = 'bulk-status err';
    $('bulkStatus').textContent = 'Match failed: ' + e.message;
  } finally {
    $('bulkMatchBtn').disabled = false;
  }
};

// ── SELECTION PANEL (meta cache + render) ───────────────────
async function onSelectionChanged() {
  updateSelN();
  await ensureMetaLoaded();
  renderSelectedPanel();
}

async function ensureMetaLoaded() {
  const need = [];
  for (const id of state.selected) {
    if (!state.meta.has(id)) need.push(id);
  }
  if (need.length === 0) return;
  // Chunk so a 40 k selection doesn't pin the backend on a single
  // giant JSON blob.
  for (let i = 0; i < need.length; i += 5000) {
    const chunk = need.slice(i, i + 5000);
    try {
      const r = await fetch('/api/cases/by_ids', {
        method:'POST',
        headers:{'Content-Type':'application/json'},
        body: JSON.stringify({ order_ids: chunk }),
      });
      const d = await r.json();
      (d.rows || []).forEach(row => state.meta.set(row.order_id, row));
    } catch (e) { /* silent; panel shows id-only rows */ }
  }
}

function renderSelectedPanel() {
  const ids = [...state.selected];
  const n = ids.length;
  $('selectedSummary').textContent =
    n === 0 ? 'nothing selected'
            : `${n.toLocaleString()} selected`;
  const body = $('selectedBody');
  if (n === 0) {
    body.innerHTML =
      '<tr><td colspan="5" class="muted">Nothing selected yet. '
      + 'Check rows in the search table, click "Select all matching", '
      + 'or use "Bulk select from file" above.</td></tr>';
    $('selTruncateNote').textContent = '';
    return;
  }
  const shown = ids.slice(0, MAX_SEL_DISPLAY);
  body.innerHTML = shown.map(id => {
    const m = state.meta.get(id) || { order_id: id };
    const edited = m.original_name && m.original_name !== m.name;
    const nameTd = edited
      ? `<td title="Original: ${H(m.original_name)}">${H(m.name)} `
        + `<span style="color:var(--muted); font-size:.72rem; font-style:italic;">(edited)</span></td>`
      : `<td>${H(m.name)}</td>`;
    return `<tr data-oid="${H(id)}">
      <td class="del"><button title="Remove" data-oid="${H(id)}">✕</button></td>
      <td class="mono">${H(id)}</td>
      ${nameTd}
      <td class="mono">${H(m.cnic)}</td>
      <td class="mono">${H(m.phone)}</td>
    </tr>`;
  }).join('');
  body.querySelectorAll('button[data-oid]').forEach(btn => {
    btn.onclick = () => {
      const oid = btn.dataset.oid;
      state.selected.delete(oid);
      renderResults();
      onSelectionChanged();
    };
  });
  $('selTruncateNote').textContent =
    n > MAX_SEL_DISPLAY
      ? `showing first ${MAX_SEL_DISPLAY.toLocaleString()} of ${n.toLocaleString()}`
      : '';
}

// Keep one order_id per distinct `field` value (name / cnic). Rows
// that haven't been meta-loaded yet are left alone — the user can
// re-run dedupe after the panel fully renders.
function dedupeBy(field) {
  const firstFor = new Map();
  const withoutMeta = new Set();
  for (const id of state.selected) {
    const m = state.meta.get(id);
    if (!m) { withoutMeta.add(id); continue; }
    const k = (m[field] || '').trim().toLowerCase();
    if (!k) { withoutMeta.add(id); continue; }
    if (!firstFor.has(k)) firstFor.set(k, id);
  }
  const kept = new Set(firstFor.values());
  const before = state.selected.size;
  state.selected = new Set([...kept, ...withoutMeta]);
  const removed = before - state.selected.size;
  if (removed > 0) {
    $('bulkStatus').className = 'bulk-status ok';
    $('bulkStatus').textContent =
      `Dedupe by ${field}: removed ${removed.toLocaleString()} duplicate row(s).`;
  }
  renderResults();
  onSelectionChanged();
}
$('selDedupeNameBtn').onclick = () => dedupeBy('name');
$('selDedupeCnicBtn').onclick = () => dedupeBy('cnic');

function updateSelN() {
  const n = state.selected.size;
  $('selN').textContent = n.toLocaleString();
  $('generateBtn').disabled = (n === 0);
}

// ── GENERATE ────────────────────────────────────────────────
$('generateBtn').onclick = async () => {
  if (state.selected.size === 0) return;
  const order_ids = [...state.selected];
  const body = {
    template_slug: $('tplSelect').value,
    order_ids,
    date_value: $('dateInput').value || '',
    machine: $('profileSelect').value,
    filename_fields: ($('fnFields').value || 'name')
                     .split(',').map(s => s.trim()).filter(Boolean),
    group_by_field: ($('groupBy').value || '').trim() || null,
  };
  $('generateBtn').disabled = true;
  $('genDownloads').innerHTML = '';
  $('genProgressBox').classList.remove('hidden');
  $('genBar').style.width = '0%';
  $('genMsg').className = 'progress-msg';
  $('genMsg').textContent = 'Queuing…';
  try {
    const r = await fetch('/api/cases/generate', {
      method:'POST',
      headers:{'Content-Type':'application/json'},
      body: JSON.stringify(body),
    });
    const d = await r.json();
    if (!r.ok) throw new Error(d.error || 'failed');
    state.genTaskId = d.task_id;
    pollGenerate();
  } catch (e) {
    $('genMsg').className = 'progress-msg err';
    $('genMsg').textContent = 'Error: ' + e.message;
    $('generateBtn').disabled = false;
  }
};

async function pollGenerate() {
  const tid = state.genTaskId;
  if (!tid) return;
  try {
    const r = await fetch('/status/' + tid);
    const d = await r.json();
    const pct = d.total ? Math.round((d.progress / d.total) * 100) : 0;
    $('genBar').style.width = pct + '%';
    const prog = d.total ? ` (${d.progress}/${d.total})` : '';
    const groupInfo = (d.groups_total > 1)
      ? `  ·  group ${d.groups_done || 0}/${d.groups_total}`
      : '';
    $('genMsg').textContent = (d.message || '') + prog + groupInfo;
    if (d.status === 'done') {
      // Wait for the post-render wrap step to replace per-group parts
      // with the single batch zip. /status also reflects the updated
      // message, but ready_parts is the authoritative signal.
      renderBatchDownload(d.ready_parts || []);
      $('genMsg').className = 'progress-msg ok';
      $('generateBtn').disabled = false;
      return;
    }
    if (d.status === 'error') {
      $('genMsg').className = 'progress-msg err';
      $('genMsg').textContent = d.error || d.message || 'failed';
      $('generateBtn').disabled = false;
      return;
    }
    setTimeout(pollGenerate, 800);
  } catch (e) {
    $('genMsg').className = 'progress-msg err';
    $('genMsg').textContent = 'Polling failed: ' + e.message;
  }
}

function renderBatchDownload(parts) {
  const box = $('genDownloads');
  box.innerHTML = '';
  if (!parts.length) return;
  parts.forEach(p => {
    const a = document.createElement('a');
    a.href = `/download/${state.genTaskId}/${p.index}`;
    a.textContent = `⬇ ${p.name}`;
    a.setAttribute('download', p.name);
    a.style.fontWeight = '600';
    a.style.fontSize = '.88rem';
    box.appendChild(a);
  });
}

boot();
</script>
</body>
</html>
"""


LOGIN_TEMPLATE = r"""
<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>S&amp;S Law Firm — Sign In</title>
<link rel="icon" type="image/png" href="__LOGO_DATA_URI__">
<style>
  * { box-sizing: border-box; margin: 0; padding: 0; }
  body { font-family: -apple-system, "Segoe UI", Roboto, "PingFang SC", sans-serif;
         min-height: 100vh; display: flex; align-items: center; justify-content: center;
         background: radial-gradient(ellipse at top, #1a2541 0%, #0b1220 70%);
         color: #e2e8f0; padding: 20px; }
  .login-card { background: rgba(17, 26, 46, 0.92);
                border: 1px solid rgba(201,161,74,.28);
                border-top: 3px solid #c9a14a;
                border-radius: 12px;
                padding: 38px 34px 34px;
                width: 100%; max-width: 400px;
                box-shadow: 0 24px 60px rgba(0,0,0,.55); }
  .login-logo { display: block; max-width: 260px; width: 86%; height: auto;
                margin: 0 auto 14px;
                filter: drop-shadow(0 2px 10px rgba(0,0,0,.5)); }
  .login-tagline { text-align: center; color: #cbd5e1; font-size: .78rem;
                   letter-spacing: .18em; text-transform: uppercase;
                   font-weight: 500;
                   border-top: 1px solid rgba(201,161,74,.32);
                   padding-top: 12px; margin-bottom: 26px; }
  .login-tagline span { color: #c9a14a; }
  .field { margin-bottom: 20px; }
  .field label { display: block; font-size: .74rem; color: #94a3b8;
                 letter-spacing: .14em; text-transform: uppercase;
                 margin-bottom: 8px; font-weight: 600; }
  .field input { width: 100%; padding: 12px 14px; border-radius: 6px;
                 background: #0b1220; color: #e2e8f0; font-size: 1rem;
                 border: 1px solid rgba(255,255,255,.09); outline: none;
                 transition: border-color .2s, box-shadow .2s; }
  .field input:focus { border-color: #c9a14a;
                       box-shadow: 0 0 0 3px rgba(201,161,74,.18); }
  button { width: 100%; padding: 13px; border: none; border-radius: 6px;
           background: #c9a14a; color: #0b1220; font-weight: 700;
           font-size: .92rem; cursor: pointer;
           letter-spacing: .08em; text-transform: uppercase;
           transition: background .2s, transform .05s; }
  button:hover { background: #d4af5f; }
  button:active { transform: translateY(1px); }
  .error { background: rgba(220,38,38,.14);
           border: 1px solid rgba(220,38,38,.4);
           color: #fca5a5; padding: 10px 14px; border-radius: 6px;
           font-size: .86rem; margin-bottom: 18px; text-align: center; }
</style>
</head>
<body>
<form class="login-card" method="POST" action="/login">
  <img class="login-logo" src="__LOGO_DATA_URI__" alt="S&amp;S Law Firm">
  <div class="login-tagline">Batch <span>Legal Notice</span> Generator</div>
  __ERROR__
  <div class="field">
    <label for="password">Access Password</label>
    <input type="password" id="password" name="password" autofocus required>
  </div>
  <button type="submit">Sign In</button>
</form>
</body>
</html>
"""

# Inlined S&S Law Firm logo (base64 PNG) — single-file delivery constraint.
LOGO_DATA_URI = "data:image/png;base64,iVBORw0KGgoAAAANSUhEUgAAAqwAAAE6CAYAAADTFQDzAAAAAXNSR0IArs4c6QAAAGxlWElmTU0AKgAAAAgABAEaAAUAAAABAAAAPgEbAAUAAAABAAAARgEoAAMAAAABAAIAAIdpAAQAAAABAAAATgAAAAAAAACQAAAAAQAAAJAAAAABAAKgAgAEAAAAAQAAAqygAwAEAAAAAQAAAToAAAAADeAfSQAAAAlwSFlzAAAWJQAAFiUBSVIk8AAAQABJREFUeAHsXQd8FNX2PpJASCMN0hPSIITeOwiiCBZQihV7w65Pnz67Pvt7+vfZBXvDAmIvYAHpvfeShA4JSQgkEPr/fDfMstns7txNdpPd5Bx+y25m7ty580377qlnZKalnyQRQUAQEAQEAUFAEBAEBAFBwEsRaOCl45JhCQKCgCAgCAgCgoAgIAgIAgoBIaxyIQgCgoAgIAgIAoKAICAIeDUCQli9+vTI4AQBQUAQEAQEAUFAEBAEhLDKNSAICAKCgCAgCAgCgoAg4NUICGH16tMjgxMEBAFBQBAQBAQBQUAQEMIq14AgIAgIAoKAICAICAKCgFcjIITVq0+PDE4QEAQEAUFAEBAEBAFBQAirXAOCgCAgCAgCgoAgIAgIAl6NgBBWrz49MjhBQBAQBAQBQUAQEAQEASGscg0IAoKAICAICAKCgCAgCHg1AkJYvfr0yOAEAUFAEBAEBAFBQBAQBISwyjUgCAgCgoAgIAgIAoKAIODVCAhh9erTI4MTBAQBQUAQEAQEAUFAEBDCKteAICAICAKCgCAgCAgCgoBXIyCE1atPjwxOEBAEBAFBQBAQBAQBQUAIq1wDgoAgIAgIAoKAICAICAJejYAQVq8+PTI4QUAQEAQEAUFAEBAEBAEhrHINCAKCgCAgCAgCgoAgIAh4NQJCWL369MjgBAFBQBAQBAQBQUAQEASEsMo1IAgIAoKAICAICAKCgCDg1QgIYfXq0yODEwQEAUFAEBAEBAFBQBAQwirXgCAgCAgCgoAgIAgIAoKAVyMghNWrT48MThAQBAQBQUAQEAQEAUFACKtcA4KAICAICAKCgCAgCAgCXo2AEFavPj0yOEFAEBAEBAFBQBAQBAQBIaxyDQgCgoAgIAgIAoKAICAIeDUCQli9+vTI4AQBQUAQEAQEAUFAEBAEhLDKNSAICAKCgCAgCAgCgoAg4NUICGH16tMjgxMEBAFBQBAQBAQBQUAQEMIq14AgIAgIAoKAICAICAKCgFcjIITVq0+PDE4QEAQEAUFAEBAEBAFBQAirXAOCgCAgCAgCgoAgIAgIAl6NgBBWrz49MjhBQBAQBAQBQUAQEAQEASGscg0IAoKAICAICAKCgCAgCHg1AkJYvfr0yOAEAUFAEBAEBAFBQBAQBISwyjUgCAgCgoAgIAgIAoKAIODVCAhh9erTI4MTBAQBQUAQEAQEAUFAEBDCKteAICAICAKCgCAgCAgCgoBXIyCE1atPjwxOEBAEBAFBQBAQBAQBQUAIq1wDgoAgIAgIAoKAICAICAJejYAQVq8+PTI4QUAQEAQEAUFAEBAEBAEhrHINCAKCgCAgCAgCgoAgIAh4NQJCWL369MjgBAFBQBAQBAQBQUAQEASEsMo1IAgIAoKAICAICAKCgCDg1QgIYfXq0yODEwQEAUFAEBAEBAFBQBAQwirXgCAgCAgCgoAgIAgIAoKAVyMghNWrT48MThAQBAQBQUAQEAQEAUFACKtcA4KAICAICAKCgCAgCAgCXo2AEFavPj0yOEFAEBAEBAFBQBAQBAQBIaxyDQgCgoAgIAgIAoKAICAIeDUCQli9+vTI4AQBQUAQEAQEAUFAEBAEhLDKNSAICAKCgCAgCAgCgoAg4NUICGH16tMjgxMEBAFBQBAQBAQBQUAQEMIq14AgIAgIAoKAICAICAKCgFcjIITVq0+PDE4QEAQEAUFAEBAEBAFBwF8gEAS8BQG/Bn7Uqk1ruva66+jIkSP02Sef0IZ16+n4iePeMkQZhyAgCAgCgoAgIAjUAgJCWGsBdNllZQSCgoKod9++dNHIEdSxQwc6eZIotEkoffvNZJo3Zw4dOnSo8kayRBAQBAQBQUAQEATqBQJCWOvFafbug4xPSKABAwfSoLPPpnbt21GTsDA14O49elDjxo2pWXQzmjFtOu3evdu7D0RGJwgIAoKAICAICAIeQcCvaUTkkx7pWToVBEwQaNiwIbVp15YGnzuEhp53niKrwSEhlq1AVqOjYyg6JoYCGgdQWdlhKioqpBMnTljayA9BQBAQBAQBQUAQqPsICGGt++fY644Qvqrh4eHUum1bGjFyFA0bPoxSUlMJBNZWsCw6OprSMzIoMLAxHThwgEpLSunY0WPsNsB+AyKCgCAgCAgCgoAgUOcREMJa50+xdx2gIqtRkdSzdy+6+557qG+/fhQSGkpnnHGGw4FiXQhrXrNat6YWLTMpPz+fCgoK6MjhI0JaHaImKwQBQUAQEAQEgbqDgBDWunMufeJIWrVqRZddcYXKBJCYlET+/v5Oyar1QYG4RjLZ7dSpMwUFB1Fefh4VFhRaN5HfgoAgIAgIAoKAIFAHERDCWgdPqjcekn9DfzrvggsUWe3bvx8HUkW7RFZxTCCsfn5+hIwCSUmJlJSczMsaUG5Ojvi1euNJlzEJAoKAICAICAJuQkCyBLgJSOnGPgLwQY2Ni6Mh5w1V5v9M1rDCf7U6Aq1sXHw8BTJxDY+IIGQZmPLbb7Rn9y7l21qdvmVbQUAQEAQEAUFAEPA+BETD6n3npE6MCNrQqGbNqGOnTpwBYChdOHw4tWjRgoKDg912fOVZBKIpJaU5BQQ0JuIYrNLSUsnZ6jaEpSNBQBAQBAQBQcA7EBDC6h3noU6NIjgklJKaJ1PPnj1p+EUXMVkdRhGsCYVm1N2CPsM4b2tbzjjQJDyMOetJOlx2hI5ypayjR4+6e3fSnyAgCAgCgoAgIAjUAgJnZKalS26gWgC+Lu4SWtVGAQGEhP+jRo+mHr16Vtv87ypOhYWFNGfWLJo8aTItWriQSatkEnAVQ2kvCAgCgoAgIAh4GwJCWL3tjPjoeAIDAznlVEu6fMyV1LlzZxVUFcDkFUFSNSnHjx+nw4cPU96ePTR//nz6asKXlJO9WdwEavIkyL4EAUFAEBAEBAE3IyCE1c2A1rfuGjUK4AIAbWjgWWcpjWoSp6qCib6q5n9UsQLphLa2QYMG6lMVTI8dO0ZFhUW0bdtWmjd3Lk3/axqtXbNG3ASqAqZsIwgIAoKAICAI1DIC7ncqrOUDkt3XDAIglFmt21DX7l2pS5eu1Kp1FqeaSlJEE+uqIgcPHqStW7bSgvnzmKj6Ubce3Sk5KZmzAQS63B0Ic9NmTVXe1qioKEIlrSWLFtNidhNYu3aty/3JBoKAICAICAKCgCBQewhI0FXtYe9zewYRRUBVBpdJ7cpk8pzBg2nwkCGcCaAjgRRifVXIKrSqeXl5tGL5cvp96hSa9NXXtGzZMk5RdZT8/P24JGug+rjaN9pDS4s0WsnNm1NiYiKF8e8gzlTgz64Kh8oOq3343ImQAQsCgoAgIAgIAvUMASGs9eyEV+VwjUj8xOQkat+hAw0ZOoQuvfxyDq7qTs04dVVV/VRPnjxJRziaP29PHs2cMUMR1V9/+oUOHDigPiuXr6CCwgJOWRWgsgzgGwTUVeKKY8YxRDVtSsgD265dOybeIeUEm/s7yYQZLgQgziKCgCAgCAgCgoAg4H0IiA+r950TrxqRH5vmm0U3o/4DB9CQIUOpY+dOqtJUdQYJoooPfFW3bNlC740fTzP/nkEFe/fa7bYp73/AgIF04803UQJrSUGQQVqrQlytd1BSUkKLFy2iKb/+RrOYMBfsLaDjJ45bN5HfgoAgIAgIAoKAIOAFCAhh9YKT4I1DQIL/DI76H3zuuUqTGsfVqoKDQzhtVaMqa1SN4wRZ3Zu/l/7843eaNHEibcndQgdLSh2SRZDmkNAQgh/q6Esv5QCvgeybWu6CYPRZlW8jo0Ap73vnzp3Kd/b3KVNp08aNBH9aEUFAEBAEBAFBQBDwDgSEsHrHefCKUUBj2Zx9PTuwFrVzly6UmZlJMTExFBEZqczy1dVo4iCh1Vy9ahX98fvvNH/uPMrmlFPHjh7TOn6Uec3galk9uCDBoHPOoTacnSCIy7NWV0CgkQqrsKCA9nA6rPXr1rHmdTEtW7qUtm/bprTB1d2HbC8ICAKCgCAgCAgCVUdACGvVsasTW4LwJSQlUmpqGmswUyglJYVS09JUkBICqdwhIITwD83NzaWli5fQ7FkzaSFH6+/Ny69S9yDRXbt3p779+nLAV2cea3KVfVvtDQCuCRhrTnYOf/Mnh3/n5NDO7dsln6s9wGSZICAICAKCgCDgYQSEsHoYYG/sHmVSm0VHU0xsLMUnxFN6ejq1ZG1qenoGa1PdW0IVQVX5TEzhq7pg3jz6e/p02rhhgwpyqg42/g39qVVWFvU/cwB1695NEW0EVTVq1Kg63VbYFqVdUTlr86bNtGHDesrevJl27tihtLAg20VFRRXayx+CgCAgCAgCgoAg4BkEhLB6Blev6hWmdGhSg0NDKaxJE2rZKpPat+/AEf/tKZ1TVLnDrG57wCB7paWltIO1knPnzKFvJ39L2zjH6pEjh22bVuvvgMaNlRvDxaNGUQ/OWhCfkMC+tsGEY3a34Hg2bdyk0m8hBdcmJt7FnNGglD/wecUxiwgCgoAgIAgIAoKA+xEQwup+TL2qRxC3JPZL7dSpE/Xq3YuT8fdQOVOrWolK9+C2bt3KQVV/0uRJE2nj+g26m1WrXVbr1jSCietZgwZxNoGEavVltjHIaQH7vC6YN5/mzp1Dy9jVYTuTcyGtZsjJekFAEBAEBAFBwHUEhLC6jpnXbgEzeVRkFGtQW1ELjvBv0SKDmrNPKkzlRvJ9fFen5Kmzg0fUfR4HLU2ZMkWliQJRLSra53atqqMxoExsREQ4ZbKrQL/+/emccwdXK0+so/1guVFC9tChQ1RWVqZ8W/fm5yt/V7g8bNy4gdavX0/7uDwscryKCAKCgCAgCAgCgkDVERDCWnXsanVLI9UTCCmqOCUlJ6tKTtEx0dSEzf5NmoTxJ1QlyIdfJ0iqpwR+qrt27VJR//PZT3X9+nW0a8dO5RLgqX066xfVuJJYwwri3rNXT6VVjmV/XU+4CRjjAIE9wpkGkAVh//4DtP/AftpfXKyKIkDzunXrFi47u0UFcx0qOegwhZfRn3wLAoKAICAICAKCwGkEhLCexsIrfxlaQ+QdjYiIpPDIcIoIj1AlRlXwFFeaasofJPfH3yFcwcmT5NQapP379ysf1Q2sSV2zZjUtWbyY1q1Z6zVmcWCX1TpLpehCCqyMFi0pkTMiAKOaEGicS5nAInArn/PO5ufn0V7OQLCPtc77OGCraF+R0sDiN9wL9u0rrjFtdE0cv+xDEBAEBAFBQBBwFwJCWN2FZBX7gaY0oHEABXJQFAKIgoICVYL+kJBg9Q1tKQhpNEf1N23WVJm4YeKP5NyoNUW8jENDeir4aCI6HqZ/pH1C8BGqRcEErptP1eivpr6hWYW2tUvXLqq0LFJ3RXNqrPDwcKV1dUd+WVeOBaVnQWLhQoDqWnl5TGT5N8gsJgElB0pYO82fg6UczHWIDsPlgIO6DpcdFs2sK0BLW0FAEBAEBIE6g4AQVg+eSgQ2NWBC6u/vR378G6Z5fy4r6scEquGp7+DgIKU5bcokFFrSZs2iKTY+juK5shTyjYaz1hSlSGtToCk8cvgIk6dSpSmEJvWvP/+g5cuW0wEmWL4kYWFh1IED0AadPUjlcMUkICgwyC0VvKqLA3AuYiK7e/du5WIBNwtFZJV2Fmm0Cllje5B9Yo+qsrbHePJw7NhxOoq/2U8WkwlMKsRntrpnQrYXBAQBQUAQ8DYEhLB66IzAHI3E+1HNohQJhW9pbEys0uwpYsoa02ZMUqElbWgndyi0fobmz/j20FBNu0XKpmXLltHvv02hGTNm0G72Tz1x8oTPVoACng3OaEDx7Oc6gMu8njN4MKf5ak+NOSCtNgVkE4Jv47f1eDBpOFByQJHYcq3sXs4Ju5v27N6jtLT50Njyp5ADvdydPsx6HPJbEBAEBAFBQBCoaQSEsGogDk1pYzbVw4QMP1Fo6Zrgo4Kbmqi/Q0ObUCgHOTXhb9S9R25TPz9/1qyyRhVaVo7gb8j9QNOK/vCBqVqRJw8GRGkcnt0mZRz9jmCh+Zy2ad7cuariU0F+AZusi+uMBg/nJCwsXE0cUri6lxGglcC5XBuze4a3CQK7QGSPcpDbMdbGQpOqtKz4zeVtj584zppWLD9KB/n8lbDrAdwPyt0MDlAxB4Hhg7+hGd+3b5/K4gCf2jKelGB7EUFAEBAEBAFBwBsREMJqclbOv/BCVQIUBBUmfesPCCf+btiwfHnDRuV/YxnM+LWtGTU5tEqrEfizhUuSorLT5k2baNu2rbRj23batn0HHWSfSntav0qd+OACnKdgnmQkJSZx/tZESkpK4gCtFpTeIoOSOfsC/IV9SYyUW3ARQAaHo0dOfR89QofV30fUcrXuVBv40s6ZPZt+++UXXzpUGasgIAgIAoJAPUHAv54cZ5UPs3Wb1jSQE9FDu1rXBFrUgoJC2rlzpyo5un37NpVHNCcnm4nrFp/zT63q+QERL+FUVGvXrFEfaNCRLiyVta4p/J3IBBYlbOPj45WGvbZdB8yO08iziwmVbhUzlM/FhEUIqxm6sl4QEAQEAUGgNhAQwmqCOoKmaipNlMlQqrUapAyaN0SgI0doMZuDdyPSf/NmWrN6Na1YsZLyONhHKjWRMpsj+wE+8EWOiY2hdlzGtk2btpSamqr+hiuBkecW14evadNtL6YGfg3YKuC5XL22+5O/BQFBQBAQBAQBVxAQwqqDFpM9XxKQU3zg4wgCis9hTmpfWlJK2dkgqGtoCaeiWrd2rfJj9KVjq+mxInhpG5eZxeeXH39SGtas1m2oc9fOnOO1NaWlpVMQZ3oICAhQ7iGGfzIIrE+RWB+7xmv6OpD9CQKCgCAgCNQuAkJYaxd/j+0d/om5Obm0bt1apUGFFhWlUsuYuB4/foxOHj8pQTZVQB85aOdzENqCBfNUUF0gB2ehDG7bdu2odZs2nO81U2lhYY4XEQQEAUFAEBAEBAH3ICCE1T041movpaWlKtXRdg6QysnJUQFT8EEt5oj+Eo4GL+EI8EPcBumpfEFA9jp27swEe71XaoBVNP0JUpH5SOq/etUqysnNoWl//snBW6Hs7xzGvq+plJaershrAlfXasbFH3T9SX3hHMkYBQFBQBAQBASBmkRACGtNol3FfRkm/gMcGITk8Xs5oruAqyIh7yaqI2EZ6tYX79vPiecLKL+Ay39yLk5fTSAPYjf6kkto8qRJXJxgGR3i4DBvFkwE8OEzooYJtwBos5GHN7JplEqBhkCuyMgoQoEIfKJ4Ob7Ducxuk7AmajufciHw5hMiYxMEBAFBQBCocwgIYTU5pSCLNSGocoT8msifebAUBKhcIwq/0xKuR4/8mTBHFzAZRQoiRViZrO7lsp7QsPoqObXFNpCT96dlZFC37t2UtngXZzDYsmWLbTOv/hvnAon98TEEJBZFIppyNbPoZjGKsEZGRSpSGxERqXL5Yn1QSDBX3gpkbSyX5uXfwAMa55qodlZDl7oBiXwLAoKAICAICALaCAhh1Yaq6g0NMnoUid75Y50bE78PHy6jMjYtg5giiXsha0lBSsvrzXM1o7w9tJOT+KNNTRHoqh9t9baExrFHjx4UxmnEUH1q+dJlPkdY7SGA845E/fhs2rDR0gRaVZBSVN2KjY1lAttU5X2FBhb5X4EHClEEBDRWxQzK8/5yaV/O+Qsiqz5MhlEtrSZIrWXg8kMQEAQEAUFAEKhBBISwuhFskE98QE5ALE+w1vTEiZNKW1pcjMpCRYqQwoRfxCZ7ENE9u/aouvHQxpVy2c36Lk2jm1Kffv0UEWuV1UrlPwWpq6tEHccFdwKQWGsia30dBIeEqlRa8fFxFB0do8r7RkZGcMYCJrQR4apaVxj7zQYFBnEKNi47e6poBbS6RqEL6/7ktyAgCAgCgoAg4GsICGF14xlD8M2SxUsom3ObIgm7Yb4/xIQEpTQ52RRH559Q+VBPnCK0ICwnmdRKWUxSmsaEhERqzzlPQbaiY2JU0v5I9gWFC0R9FUxktmQfpK1chQzkHYS0Ab6R/5Vzp3ICLfLnZYHs+1vuG9tMuRqkc7Wuzhy81qFjh/oKnRy3ICAICAKCQB1BQAirxolkTqklKGk68++/aQMH3Bw5xtrWMi6LyeUw66p2UAsUFxolcRnUTp07Ka0gNoOJOy09jVpy2qi59ZiwAgsjMwF+80Wlvmz/g7sBJkrIFNHIvxFr7neyFjZci7DKNWqLpvwtCAgCgoAg4E0ISGkbk7Phyov84KGDbPIvUj6oKPWJpPOubG8ylDq/GuVQO3bqrLSIRsR8Ci+DplDEHAFca0izhWsPftD79hWrv823LG8h16ouUtJOEBAEBAFBoKYREMKqgThM+ToC06xBtHTaS5vTCCDtU2paqvqcXkoUFx+vcpnCj1PENQTYa4D4gtTaCFYEIaxaUEkjQUAQEAQEgVpAQAirG0EXslp1MFPS0igjowU1aVKek9ToKZQT8SdwBH3z5s2NRfKtiQCuR/wTEQQEAUFAEBAEfB0BIaymZ1BPu4puFGHV1GiZ7raeNWjbti21zMy0e9SxsXFafph2N67nC2USVc8vADl8QUAQEATqCAJCWN14IoUcVA1M5CFtmdmSUlLsa1GjY6JV5oCq9V6/t5Jrsn6ffzl6QUAQEATqCgKSJcDkTLri1wfl6hlnyBzABNJKq5Ed4DDnr924sTyhPtI1+ftzdSd/P0se0WbR0Sqxfn5evqQAq4Sg/QXKIUBT46/8tHXTYdjfnSwVBAQBQUAQEAQ8hoAQVg1odUkrtFlncOJ2ET0E/Br4UePAxiqn6KoVK2lvfj5XdArgfKxBqqpTYHAQBTUOVMR1z+49nOIqQ60v5EwMKrctF2gQcYyAuh41L0fda9zx3mSNICAICAKCgCDgOQSEsLoRWzG/OgczMjKKEpISCMUBorgYAMhpQOPGlu/GAY3U70a8vDG7CQRwuVGUHkW1MLgFXHjRMNpfXEwHDhygkgMlqkIUqkQd4nRihw6WlVcM27ObSktLnQ+kHq2Va7IenWw5VEFAEBAE6jACQlhNTq5LmidN86vJLuvMaqVB5epLKCMa1bQpteB8qh07daK27dpRUnISBfG6qspRTp4PYoq8t0iYv69oH61bt45QbWwzVxrL51K3B/bvr2r3dWO7ch8V7WPRTd+m3aE0FAQEAUFAEBAE3ISAEFYTIEFYT5w4YdKqfHW5z6BW0zrdCFo9aEkjudZ9Cw6m6te/P3/6qZyq0Kq6Q6B5DQ8PVx+jvwFnDaTdu3bR/HnzaeqU32jB/AVUdugQHWPXAZcmHkaHPv5d7hKg5xMAfOojRj5+in1++CmpKdSCK9mlpqVTSkoKJTdPprCwcCoqKuRJZz7l5e3hUtfZ9PvUKWpSanvAbXjy+5+X/kvj3xlH33/7re1qt/wdwjmgO3TqSKmpqdScA0OTk5tTbHycKsqxe/du2rlzJ23fuo1+/vFHHneRW/ap24kv4Kd7LNJOEDBDQAirCUIwR/Ob3KRV+WpXCIJWhz7aqFmzZjR46FA6Z/A5lNmqlQqcAlGtCfM0grMGn3su9ejVk7K5VO73331Hs2bOpIJ6XtrV9FJyYWJm2pc0EARMEGiV1Yruf/BB6tuvn0nL8tWPP/UkLVywgH775VdFDA23nxGjRlJ6RgZF833vbmnUKICuuf46uuGmGytMjK33075DB8uf//jn/fTtN9/Qe+PeVWWRLSs88MMX8PPAYUuX9RwBIawmF8BJ1q4e19WwsmaxvgrM/+Fs+h90ztnU/8wzuQhABjVl4hocHFwjRNXA3c/PjwKDAlnD24hCQkIoNi6OunbrRr/+/AstXbKY/V0PGU3r/LeaQGkWDoAVQU3O6jwqcoC1iUB0TAzde999NOyi4YR71VaKCgtpz5496t6NT0ggZAyBwKLSu08f9Xno0UdozuzZ1JAzifTu20etb8T+7u6UEaNH0R133UXxXGlPV+DidOVVV9EFw4bRtWPG0No1a3U31W7nK/hpH5A0FARcQEAIqwlYJ07iRa7pEsCEFf/qk4AUxScmUif2Te3eowe1btOaUth0hoe38bKpDTzwMgRhVZ/QEILWNy09jWaztjUnJ6c2hlQr+9Q9B5iUnThxvFbGKDutHwhksOn/sy8mVNJWHj58mL6ZOIneeestymOyagjyM2e1bk39zuxPV197rZr8Yh2WDzr7bKOZ+sYE1V1y17330G133FGpu6+//JLH+La6Ty4aMYJuv/NORaRtG6LM9Lj336crL7uctm3ZYru6yn/7Cn5VPkDZUBAwQUAIqwlAiqxquwSYdFZHVoOkggAmJCVxxH+CSvqPYKp27dpz1H9ArRJVexDDXBjWJ4xi4mLZjzZO+bZu3bKV/V131mmNq9Kwas6fxIfV3pUjy9yFAO7Bce+9W4msIuPH1VdcYVcbCWvIksWL1eeLzyfQfWxyv3D4cLvPF3dpWEdeMppuvf32Sof9w/ff0+OPPGpZ/s6bb1FOdg69+sbrlmXWP3C8N9x4Iz352GPWi6v821fwq/IByoaCgAYCQlhNQIKZVDcY5Qw2X+FTFwUmudDQJsrsj5RU8KHq0LEjtWvfnuLY7O7shQH8QPyPcnGAI/w5iiAoaPROTQSs8W3AhRcaIJctk2JoB/E3CgjgN7Sm/v7+ysWgnIxpsjE+IfChzcrK4oCJZOrcpYvyh1u+bDnt3L6dCgoKONNAMY/tcJ06dQyhwlHnoIxzpNNW2ggCriAAd6G3331XTW5tt3vogQftklXbdtC8Pnj/P+mrL1jLycS3SZMmFZrg+VRdSUtLpyeeesquC9MXn31eqfspv/5KM6b/Tf0HnFlpHRZ079Hd7nJXF/oKfq4el7QXBFxFQAirCWLKh/W4nksAyBWIVF0RPCjP8DuDiaI/JbI2tQsTPQRJdOnWVWlK7Pmg4dgN8gOSevz4cUVWYfbbuWMH7di+g/ZyAFRZWRkhNdXx4+UR/MDOv6G/pbJVY87PipdQILsWhIaGKtN+Yy4iEIRiArwM/muAGhMEjMMgsMa3o3MAn9pOnTurT0lJCa1auZJmzphJc2bNolx2FUBGgRN8vo/XAfO4wpTPnY7gmMWHVQcpaeMqAgPPHkRt2raptFkp339/TJ1aabmzBdC43n37HaytfU/5qRtt3ZF95KKRIxxOvI/ws8qevPLySw4JK1yj3CG+gp87jlX6EAScIaD3NnPWQx1fp3xY2Y9VRwxNoE5bb24D4hjFSf7TM1pQRssM6tq1G6W3yOB8qpHlSf5ZW+nMNxIkNZtzoS5bukwRwm1bt9IezosKknqMNazKX5LbGJrV/gMGcKBWf0WKrf2FQT7xN0jk+g0baOqvv1FudraCrgn7icVxQERKSkp5qpnmzSmBfWkxRmfaXmvc4QsHV4asrNZ0yWWX0iYuDTt/3jzauH4Dm/uyaW/BXjp21HeraSkyz9ppHcG5wORMRBBwNwIXDLvQbpf+PCFF4ZDD/FxwRebOmUOPPvwwvfDf/1ieQw3dEHQ1ZOgQh8NADunVPLm1FQRWwXUBzxJbwYTYHeIr+LnjWKUPQcAZAkJYnaHD60CYDGJl0lRp+pQ526xhDa5HQFRqagprRCP4WKDtLDfPg3BCgwnNBD6hTUKVmQ05EBHtH87fQRxtH8yBSwgiQPASzPGOBBpU5CRcuWIFrVi+XPl37d69iwoLi+ggP7idRedjHIgIzuCXgjXW5SQK6exPqjGCCO/YsVOZ7kGqoRFdwftrwmMLOhVgFdU0ipJYGwztBshsErsAQCNrTxuMZfhAmxvCgVk4ZowBmh+kzSnmwgNFBYW0n7+Li/fx8lJ1HHBrAInGeECq/fzgxuDHbgVFtHnjJoWDI5xqcrkftM88Lh3BtVEXtMo6xyptag6BAC69jKwh9gTPnXPOHUw/ff+DvdVOl/3A6eratW9HV11zjWrXqGH1gq6a87MimSe9jgT7+m7yZLurd3Ee1rT09ErrkMmguuIr+FX3OGV7QUAHAccMRGfretAGplJdzRPID4iLtwiI5rmck7RXn94qwrbc5Fvuk1tOtEDY/Jm4+itNRyATN5jgoS0AiTMTEDZUlMrm4IONG9bzd7bKfbp502Yq5upTugSoYaNy0z9IsSMBkbY2+0HzWcxlWvHZabURyCmqaiGdVWxsLH/HUkwMf3OwVRz/HR0dQ5FRkUozAwwMwbkLCw9THyzDRAXEVJV+PXiIyg6XKU3QUd4v3Bhw7Aa5PuOU3y0CSJYuWULjOJLYG6QBiDR/dATHKy4BOkhJG1cQSIhLUBNGR9s8/uSTanJrT3vpaBtj+YRPP6MxV1+tJo26VhVjW9tvPDOcydDzzqcXnn1OuTHZtoO1x558N7n6hQx8BT97xy/LBAF3IyCE1QRRkBLdF7k/kz9otbxFojiSH/6m8Dt1ZsLXHa9RDhUkESVR4Yu6edMm1qiuYHPZKo8ny9YZJwjmQXZBgBuCIXFx7DqQyq4D+DRPUcEf4RHlVbLCIyKoCfvINrYx6QEvkHZF3CONnpx/g8TCR/aDd9+z+2JzvrX714JIw9dXRxD/pjsx0+lP2ggCQMCeqdwaGQRPfTrhc3r+mWdp4ldfWa8y/Z3DFha4HXXq3MlueinTDqwawGLiTDDJHT7iYpr01dcVml08ciQ1tUN2v5zwBU3/668Kbavyh6/gV5Vjk20EAVcREMJqgpgrPqyuaLRMduuW1QHs1wUSba1J1OkYxAsaNxBU+J0eYXM/TP5FrDVF4BT8O9esXkNrVq2m/L15Xu/nuYvTV+ED3zdgEczm/8zMTGrVKotatspkl4lUVSkH5BQ+ddDW4FOuMdefgKBvuDdgW2BX26LGr6lhVe4XYK0igoAbEdixc7tpb7CKPP3cs3TeBefTq6+8QsuWLDXdxmiASnaKsFYzD+vWLblGlw6/b7z5Zvrm64kWywp8Sx9+7HSqK2PDX3/+2W3prHwFP+PY5VsQ8CQCQlhN0FU+rEzedKQ85ZI+wdHpszptDrG28VDZIUWeQF4MgfYQ5Kpce3z62AzSApJayBVndmzbTps2baT169YzQV1NWzkJtlES0ejL175xjCX7D9DihYvUB+OHFiMhKZE6dOjI6bqyVKnHhMQEQvouuCEYhB/fjrBDPyD1cAvANeMNgrFan3dnYxIfVmfoyLqqIrCvaJ/yAbdNQ2Wvv169exM+8+bMpQ8/eJ/+njbdXrMKy75nv1L4kBaxxac6gnHCdx0++44khf1coVHFs/GKK6+slB2gkP3dX3juOYJ/rbvEV/Bz1/FKP4KAMwSEsDpDh9edRJCSpuapgfJh9R7Cun3HdprA+QNXsbk+JiaafcmCqX2H9ioXKUzgMO2vWbOGFs5fwES0hM38+/hhXECFBRwoxX8fYTMZSBg0rGUcCWtmNjOB0mtXg6BvyclVwVLTp08naKbxQSBXBLsOwL8N/sAIHmnTti0vi2A8yiiXtTLQMgOjY8eO8vZ7aCUHnDkLMKtJEBAAqE9Y4astGtaaPD/1ZV+zOWXc0PPO0z7cnr17ET4bOFvHhM8+o8mTvnGYIxn32t/Tpmn37azhJnZvQm5pZ/Lciy9UWo2gzG94jOO5Uld1iXOlznmBr+Bnb+yyTBBwJwJCWE3QNMzjJs3Uan8mrLoEQae/6rZBYBJM9zvYjA8tIsgXEv2zylB1DQ0cfD5//OF7Tjd1VBGvchcApJ6qP2U6oXWFCR+fEjpggR0R9ij5CFeBQPZN7dWrdwVfuaOM2beTv1HEH37O0Ggj2MxbBD6szjI7WI8T4/cWzbD1uOS37yPwwXvvu0RYjSNumdmSnnz633TnPXfTZ598Qp98+JFHLTxTfvvNlLAaY8M3FAE///Qjfc3FDBxZnrJaZyk3I+vtnP3evWsXV+DbXaGJr+BXYdDyhyDgAQSEsJqACq3iMU1/RJBVb0trVVpygE1dBxRhRQL+sLAmFtIFEotUTpD8vPx6RVJNTrtaDdIODQ6IbAS7ByBQSwVh8VoQWeR8LdfObnGoAdLZj6faYEKCj44g88FR1hKLCALuRgBWB2hJR4waWaWu4Zpz97330jXXXUdfTphAH77/gUcmhj989z3d/8ADWvfMc08/Q5989JHT4+nDwa7vf/Sh0za2K98dN55e/s9/Kiz2FfwqDFr+EAQ8gIDe28wDO/aVLo8hhZGmT6KfKhvqnZDCpN2la1dCChZDC4wAoWbRnEmAl/tzaikR+wgAp66cbSGSX5yGxhIYInK4Y6fOnOPWflob+73V3FKQVWO8Znstn5g5j5Q260PWCwKOEHji0Udp2bJljlZrLQ8PD6ext91Gf/09ne69/34ObgzQ2s6sEfoZecloeuPtt7TIKvq77c47VKETZ31vyc2l6eyugPR4ZgKt6o8//ECzZ86w29Sb8bM7YFkoCHgAAe9kVx440Kp2CbOvbrUjpLTS1WhVdTxV3S4yqikHNPSqRGBAZPv06UvIwSpiHwEEXiEYBFhZC8hgb14ODZA3CrJWGJMTs/EdhSXBJLWPWR+yXhBwhACsFLfccKMKqHLURnc5AqNuuXUs/fjbLwQtZlUlkqv5PfL4YzRj7mx69vnnVdU73b5Anl994w2Ltcredtu3baOxN95E/fkZMXXKFHtNlBvWk489ToP6n0n/vPcfNG/uPLvtvBE/uwOVhYKABxEQwmoCLl7iR4+az5DRDQgMqh55m2Bc0KR27tKlEmGFmwC0h8H8rVsVyduOz5PjASYgqp06dyZgZS3AFdg1jY72SuwasA8rtP46ArcX3etcpz9pIwjYIgD/7us40T/8UeE3Xl1pzpWp3vvwA7r7H/e61BU0qrfecTtN+esPVSkL5BOCMc2aOVP5/Ot02KZtG3qCfWzNBMGsTzxSOf3VkcNH6O477lRuDjoxA96Cn9nxynpBwFMIeB+78tSRVrFfzGwReKUjyBKA+u3eJqjy1LpNmwopmowxQiMcwkSse4/uqiSrsVy+yxFAmdoevXqpaj222nOkuYK2pw1jG5sQ73WQYby6hSyOHzvOhFVcArzuJNaxAYEUPvPUv+maMWNo/fr11T463IO33n47Pffii1p9devRg379Y6ryibWegO7k1FhXXzmGbrz2Orrq8iscBlHZ7mTU6NE06tJLbBdX+rsRW2ls5fVXX3W5uEBt42d7DPK3IFCTCHgfu6rJo9fYlyu+fahnjUT93ibJSc05x2gHZR42cooaY8TfSHSPCPgIJmciFRGAnypS7FjnYzVaADuY3NtyqrCkxERjsdd8+/lzlgM+tzqCgCvd4EKd/qSNIOAMgQXz5tNwLnf6r38+QLns61ldQUDXBcOHOe3m2huupw8+/khVurNuiAj/a8Zcxen95qvFKI7yzptvWTdx+vuxJ56gMwcOdNrmnMGDK6yH6f/dceMqLHPlj9rAz5XxSVtBwBMICGE1QRU+rLrR0wEcOQ4zsTeJf0N/SkxO5IT4rSoNC9pjpLVCUFFHLm8YExPjdeOvNOgaXADsgAkKCtieV2jd8aJDKqisVlmUzOZJtPcmQVlWEG0dUa4vkiVABypp40YEvuPE/0MGnU333HlXtYOy/vXww5X8zI2hjhg1irAezzpbGf/OO7SNi6JYy3vjx9NqzrGsI7jHXn3jdSatA+w2h1vR1ddeY1mHstb/+uf9lr+r86Om8KvOGGVbQcBdCAhhNUESfn26wSh4cHkbaYmOieXSo2nswxqtyBWSXBsuDvCJ2rF9uwoUS0hIoNS0dHELsLoeIiMiGZM0io2LVRiBoBoRv0h3hRK1IKxYn6LKu8ZYbV37P0Gy7Zki7Y0MgYXiEmAPGVlWEwj89ssvdNnIUTSGzfHT/vqrSjmBm3KBj9vuvLPScNty7unHnnyi0nJjAbSVtgLT+z133aWqdNmus/c30t0hCKsfFxexlXvuv09NaLEc/T72yCOVcq3abuPq357Ez9WxSHtBwFMI1CvCGhwS6lISZ4B+jDVpulkCUIceWi1vklatMrnUaLoiXCDe2UyyjCTXuzhJ9YrlK9TLAf6OcBtISkr2puHX6lgSk5MVJsAGxHQ7k/t9pwoD7OcqYXPnzFEEFuuBcWarylrs2jwATJ50NazIw6o7MTOOCdd7SJOKgWjGOvkWBIAAchd//tWX9OAjD2sBsmjBArr1ppvpwqHnqdytqCLninTs1LFCc0zaoP1EzmlH4ijuAFrXxzlYCiRTR0BaX3vzDerbv7+leW/OwHL9jTdY/v5ywhf0+5Splr/NftQ2fmbjk/WCQE0iUK8I65kDz6RWTCpszbvOAFcaVn6Z6wgeWPZMTjrbeqpNy5ZMWNMzFOGC+X8lE9SDrCmE5OXl0cqVK1RqFTyUjbKtnhqLr/WbyH6pqAwGbPDi3LhhA5etLVSHsX//AZo7ey4XZSh3CwDGmZmZXnOIMEM2ZJ9q3esR2tWjGvkijQNEpHWrrCzqV420QkZf8l13EUjj+wJ5ni+59FIVuKh7pJs3baKHH3yQzhkwkCZNnGixCpltn5GRUaEJAiZhPXImLZ3ct9BcIquBroAYv/7WmwSi2qZdO/rfG68pP3dsv5zz0D7/zLO6Xal2tY2fS4OVxoKAhxGoF4QVL+3+AwbQFRwF2o2j4YO5zKauoGSproYVmjZotbwlPVRTTmUFk3Z0TDSVHCihpYuXKJ9VowQnIsPz9uyhZUuWKu1aXHw8paSkUBSb1uq7AIOU1BQCJtA8Llm8mIoKi5i8nlDQnOAqWEVFhex3t1RhGxMbo7COatbMK6ALaBygAq4QGKYjOEZXCGsoV0xDbtorOFile8+e2sRYZyzSpu4hgGfusIsvdvnAMKl+9F8P0aiLLiLkNTUTZO1IYn9yQ87k576ZnH/B+U6boKrVX3/+6bSN9UqQ1nHvv0tffP0VNWnSRK3awtpa5GQ9csQ1jbHRb23hZ+xfvgUBb0CgzhNWvLAjIiJoNKceadGihaqOYvgh6pyAwwi64uAkHQFhheYJZMEbpD1rB+NZuwDCXryfTdhz5zD5On0sIK578/eeWn5MEZx01lDgU99F4cDaIUTZg8zNnTOXiov3EZ0igNC6QlM9d/YchS0wTmCNbMeOHbwCOrgC6GYIwIBhSTiseZ2j/XFui320aNmCLrnsUlVBTZccY3uR+ofAiJEjqnzQa9es5XRTV1KuRkaBRlaBVZhUmUm37t3V5MtRO9zr9951Ny3lib2u4Hlg3H9FhYU0losmFHGwVXWkNvCrznhlW0HA3QjUecKKpO+d2STVoWNHfsE2okNsFkfCZl05wi9yVxKq4yWOjzdIx06dKC4uTpnTCvmhCS3h8eMnmHOVa91O0kkqZC3hwgULCT6ZCMZKY1/MrNZZ3jD8Wh1DJvv+pqWnKUwQqAbsYP4/g/9B8BIrKyujBQvmU8HeveUaasYaxRm8QVy9DjEpc0X7Az/ogwdLlZm3q7q/Oqh8vt5w7DIG70SgPfvIt6yGn/euXTtpzGWXU35+vtMDhFbWEKSl05EHH37IaXzDYb7Xx950E8FVwVVZv36DFtE267c28DMbk6wXBGoSgTpNWOGrmpCUROcOHaLSnZSxH+KhskOkU1XEOAkgJa44/mNW3bixYwd/o19PfoOQgqgjCAhlQw8cOEC5OTnqoQmuaiGsTLoOHTxEW3hddnaOImDxbAKHP6ZR/cWT4/TWvkPZjAdfOGhMcf4RqLYlJ5fTmx3jwhCnCSuui1zGLZfNfcA4il+OwBzYGRjX1jGCsELLoys4FhyrroDg4tqB9jmcLRjnDBlCiazN9xZ3GN3jkHY1i8BTTz/NmseqT+j3Mln98/c/HA56zerVdIAnmIYY/vrG346+4Y/9n5dfcrRaLS+vNHWNy6S1Z6+e9NU3kyglNcVp//0HnEmXXXmlU+Jc0/g5HbCsFARqGIE6TVhD2J+pZWZL6suBIXh5Q1PmyksZ5+IQa5IOHSrTjhTFfqDJrU0BUUdlK5DPwKBAlUJl+dJldJK1q3BbsDZrH2fCAbIyd85swgM5KCiI87YmUQsngQi1eWw1sW9cM8iWACzw8ps/dy5jVMbYGfpVaFg5gwT7sUIrvXzpUtq9ezc1Zt+1WNaytm3fzqXAPk8cUyP4sGpeh3ANwX1RxgTUFcE2JSUlyvQ5kBOnZ7RsSYEhQa50IW3rGQKdON/zf195uVpHvZAzCTiSbyZNqrAKFax05VyedD3JhNrZpAs+/3BNWLd2rW63qh20o5N/+IFuvPlmu/2fO3QovfH22/Tkv5+iW24d67DvmsbP4UBkhSBQCwjUacKKgJm2bdsRiCs0XiWsBStj8umKIN8mzEFGoJLZttCw6ua+NOurqusbBTTm1Cr9lF8h+tjFD+3Fixap7lBf3pCTTFSO8wc+vTOn/82+mOWaibjYOBXZa7Srb98w68dwblVIcfF+mvn3TEXqgZ215vRU/JXCdteO8hdjWJMwldYmILBxrcKGjBW61+GRU9pVaEtdEdwbmAQCE2j027RrS/Fx8a50IW3rIQIghjDBV1Xgl29PSnny9O2kbyqsQk5XV+SyKy6nCRO/UtYVR9sVFhaoyli6sQ1GP5gA3//gA/Tz71PoH//8p8rZirytjzz+GL30yv+piR8mvh9/8KGxid3vmsTP7gBkoSBQSwicZi+1NABP7jaR3QGQlggvVHyqomGFr6Ir5tKGrNWq7aCrwKDG1K17Dwpj03bxvmLaunULIUqV7dmEcp2GQEt4AnlmmajkcDAD3AaAESLeEbBVHdOdsQ9f+4Z2GtcMKlzBzL9lSy5t3rxJYQTttEFYcV0ggA3fW3LK8UWO1lDOS9qdgziCA4MsbWsDAzVx0nQJOMjE05VARON4oGGFBtq4vzA5TExKNFbLtyDgEIHrbriBbuck/860mfY2DuFc2hePqJxtAAqFp5/6t8qCYr0dCOxe9jF3RRDv8MvUKfQiuwi0Y82otcBN6KaxY+nryd+45HJj3UdKSgrdPPYWeveD99XnqmuuUX3hGB5+4F88SS62bm73d03hZ3fnslAQqCUE6ixhhWleVSBKOZ3iBAQEuUhdFbzMD5bqbdeYfQcbs4aztgTFEVq1ymJ3gDhqyNpekNVNGzcpLTHG5OfnZyFSeEAammNokVeuWKlys8L/MS4hnh/WbNrmNF31RXCsFlcKNu/n5+XTqpUrLdg18GPCCpcKJSfZHaBcI4lgJeRo3cKkH0QxJjaWWrdtS8GhIbUGHVLrILG/juDarhphLdewGvtITUtloh/rMgkxtpfv+oXAnffcTT/+9gsNHDRI68DxTH/j7bc4GDK9UvsXn3+BUKbUVkD+nnzscctzzna9o7/xDBzOqbQmMjFdtnoV/T7tL1rEeVT//Hs63cdlVZtbpc5y1Ieryz/+6GOaM3uW9mY1gZ/2YKShIFADCBhv3xrYVc3uAqVI4+MTKkQuQxtUxtokV0UR1kN6hDWQA66CXMjz6upYzNqjMkp3zjWLMYCcbt68mdavW2fZ7AwrszayBCBrgCHL4IvJ1a+gMUMqMKSECQio3QAyY2w18d2IE+334ACJyMhI5esL85x1KhulYTWyBPCAkMfWkPXr11M2Yw3sgtn0h36iuLRrbUkga3iDNIP/EO0PK4KrAt9u6wAXBJvF8UQpomntHberxyDtaxcBkM+3x4+jDz75WJnI8fyyFWhhr7hqDE1l0tizd68Kq+FD/vqrr7EZ/YMKy63/+GPqVHri0ceUv7n1ct3fcK9J4qp3IXYmoHNmz6ZLR42if9x9j6UKnm6/1u1Q/erlF1+0XqT1uybw0xqINBIEagCBOqs+S0lJVeZJFWR0Cshirk6ELAGuCjRouprZ4JBg5TPr6j7c0R5kCfW0kcgdmj6MGfXus3OyLd0rLaEV6UICfEPWr1tP27ZupUOdO6uE18hP+PUXX7J2uUQ76Mzoyxe/A5nkd+/RQ/n+wj9z27attHbtGsuhNOAXJzCGGwCirtT3qbW5jPOmTZtVEBK0m926daOpv01hDffWCu0snXn4B67DoGC9ACgcqyup3oyhI0UcfHwNwQQJL/YUNnnuZe20iCBgiwAIJq4TW+ndpw9Xh+qjFu/hwKac7Bxl2YjlCVCCUjxUtFbg3vvzjz/pfy+/zBakjbbdVfp74ldf0Xa+F+++7x+EdH/VERwDyjKPf+cdWjBvvuoKQa2Y8D/HpBNZAVyRT1iz+hwHe+lIbeGnMzZpIwh4GoE6S1hhnkxObq5MQQaxKObEzQddjITGCYD2CTk4dQQVSezNxHW2rW4baALgCpDFKVrwUoBP6pbcLVTCRB2CeCtrAg/ShQegIaUlB5T7wPbt26klR3yjZGEcBzggGAsuA3VZ4K8by6Z8YIfgCGhLkXPRwA7HDuyQKQDXk+FKYWCCvKRb2U8Y2yEiGOmtUNp1PUcTY11NC3z9cBw6UlpFlwAQVuTvNfAAPjjm1NQ0WrRgoc6upU09QuCXn36mpx5/XPlowm2m35n9afC5Q6gnJtjs+28I/MfxcSTz5s6j//3fy6pCn6M29pbP5Wwfc0eNpgFnnUW333UntePSqbqC+x2uVb/+8jNNZr/YPWx9sZWdO3bQtWPG0Flnn013/+Ne01LNsOC889bb9OXnn9t2Zffv2sbP7qBkoSBQgwjUWcIaERlBYeFhqkoVErsH8ssbzvcHOZLUVYFLwCFNlwBFWJks1IYkcMBL23YIlip/+K9k/8ttNuUMYV5jJWG5MPEyyLwx3nXr1iqNBQgr/Lj69uurEnXvZBJblyWSr5eeXHcc1wm0qNmbs2kDa5ytBWQV4JUTNNay2gg0OMuXL1eEFRMGENeN7Cqw1sUUODbdVulPZMYIDq6olXLUEVwCkCnAVYFmFj6CwAP+4dgncrLqJmt3dX/S3vcQOMTX1to1ayiX/bsfuO8+FbyIowDhm/TV1+qDiVVPdj+C5hOuXMgdDbccfw4QRVoqWCngH47gRgRA2iOLriAznTMH4NO6dWtqnpqqrAKJiQkqMwBcoYzrGu4uu9hFahlrT5dy4RCdYCiM468//lAf9D9o8GCuHJiuLF9w08G7aAcT2yVLFtMvP/xkmhPcG/FzBWtpKwi4E4E6S1hDQkMVccMsdiWTCGi88Bt5I12Vw2X6GtagoNpzCUDu0PYd2qvDg+Z05fIVygxmfbx+CBxi0lUecFWZdG1mLcLmTZuVVhlBDj179aYZf/9NdZ2wRvALshf7x+GYgR18fzfamBpB9g2xdqUwlm3bsZ1WceAaJjiYNCDbwPx582qFsIayvx3cAnQE1gMU1XBVcC8VFBQovNasXsOa1RR13KF874kIAkAAJVUvvnCYUzDgumSQPKcN3bxyDRNpfDwl7ujfm/HzFG7SryDgCIE6GXSF6OhQ1nIiZRNm44XsCoAUPNCSupo7D8AplwBNsy5Igi5RcHRSqrIcfpNJnPA/o0ULRUZhmkYqq5IDFQm6inK3aAlPB1wZ+4QWQWkzeFtoCaFpRXowaI7rquB6ieesCMgQgGPeumUra4RyaF9hxdrfwM7wYbV1CQA2pYw1NNpwJQDphVtKc85SgcwNNS3QWuGa0JFyDat+uWKjT+Q0LuLSvnvz85X7CSaEfowRKoUhPZiIICAICAKCgCDgLgTqJGGN5OhsmCeh6SosKCQkfC8sLKxSJDSARqDWfjZ56gjM6CFM7qCpq0mBO0BaWppK4H6Uj3sJm7D25u+tYHJCnSZDSwgzrq07AMaLZduZdK1mdwL4JMKtAn6d0ZyuqK5Ks+hm7K/bSpkhQVgRaAUMbEv4wiUA1xIwOsEfW8FykLf58xfQ0SNHVeBaenq6R1Lg2O7b+m9FVpmw6l6DMOfDDOqqAJ/9HHQFky22L+B77BgTddx7TbiAgoggIAgIAoKAIOAuBOokYYV5F8n7kRC9eH+x8oWCmRsasKoIgiOpD7kAAEAASURBVFL2sSZJR0B4lIY3tIlOc7e1yWTClZ6RofqDeReBCYUFlRNmV3QJOB1wZT2Q7exjhZys0CyDhMG0ndw82bpJnfqdnNRcFUrAsaKIAsz627duq3SMIPCGhhVlbu1J0b4iVcr1IGvzoYVF2plWWa3sNfXYsnC+/oPYXw7j1ZEi1iQjE0RVpIQtD9DmY6IGnz8E5yH4L4xTXIkIAoKAICAICALuQkDvjeauvdVQPyCrMEnCNw+1n5FrdM3q1dpaUtthlrIGCtpKEBp8zAQv7HAO4qkpgda0BZvuUzmAAKZoaJWRZgVVq2zFMGuDTDk6lgLWEm7cuEElzke7LA4eSGK3gLpYRADYJSYncqL/Nopggrxt2LBeFVCwxa48rRVXB2NMThh1WW0aIasAgt3y8vKU+0lKSiq1bJVZo9ghcMQ66tpmiJY/jesZwYjQslZFcG9s3LCRMxIE8nVXoHzEMWnTdUeoyj5lG0FAEBAEBIH6h0CdJKzQIsJ0i3RCKE2K3KTLuUpJcfG+SmcY5lOYTqE5cyToB756jgie7XYoHhBhJwG2bTt3/R2XGE8pnPsSmmWQ1EULF1KJg9ypIGjqWJl4nzjhmHwX7C1UAUNwL0AQDTSFSayJrGsSExfL2KWqawTa1YULFyiibu9cG2SfLwRFWh1hAd9OBFvBXI5SrckcDJeRXq79drSNO5fj2sOkSUfgNoOSsgh8cSS4ZpD2y54fM643VANDycoDTNYRiAXNrp9/nXy0OIJIlgsCgoAgIAh4GAGfjYxACqJ2HBFfsLdAVSPKyd5sgQqaMGasdOzoUZWGCMFI8EM9drS8lCY0hSAQI0aPojZcQhP+nr/+/CtrYVdZ+rD+ASKD8pV4sYeFhdlNfG3dHhre8PCa07B27tyFKwzFq3HBLDt71iw2zVaO+m5wBvtgMpmHwAcTmkJHUlRYQLNmzqDBQ86lxhy8g9QsrVq34oTep3F2tK0vLYe5PqNFhsIOgXlzZ89Rvpj2jsGPrymQVoWdA5cAbHfkcBnNmjGD+vbty2l6mnFAVwJ16tKF1llVHLPXv7uWhYUxYdUIuML5h1YUx+1MWrdrQ+edfwF16NhBpfj5dvI3lMNpv3BfYFuUsMX1Dq0zqn9hQqTuQatOM1q2oE6dOvOkKoIWL1pEixcuslorPwUBQUAQEAQEAecI+CxhHXbRcGWqxkszkQOOJn09kZC4GYLIeOSVPMkaxBJOhp/PJu7jp8gq1rdp05YuGnExDeQE0sj7V05Enfucwi8U/YCwmgm0tk2bNTVr5rb1HTp2pFjWFEJbBlP0siVLmDiUk3PbnRh+jeXmYMeEFZqydWvW0i7Og9iItbcp7G4AP9nfuXoTMK8r0rJlJgerpatjgml81YoVtJ8nJvYEZW0NceQSgPXAZ/XKVSrfYkJCIsXExlD7Th0J1XZqAruoplHsw2qeIQDXQB6TTbMcrCDAIPW4zmI4+A75MX/4/nvl64s+kH0D+SVxHx05ekS5pVi7GKD61YXDhqucvrj+YmPjhLAaF5J8CwKCgCAgCGgh4LOEFRrWKDb1I98l/FVB1L78fII66D15e9TLE7lYD3Flq4ULFigyh5WJrG1FhZWBgwapykZYBhO+v0lUP4JJ8nkf8BOFj54zCeIsAbE1EFUPUy00eNB+wm+xiNN3wZ8QpQ0diXIJoPI8rPbM3sZ2IFZ7WfsGc2/TZs1UMu8UJq4JrC1Euqy6INE8WWnOqaeAIfydFUHnROGOUp+BbOED3Jxpp7Ee52IDFw1IZ01+HFcfS+PrJokrr23busXjpBWkEpH6ZoJjyM/PMyWsjRo1ZJ/UIHXdJ3CC9UFnn8NZN4rUB/l5gRcSoWOSg99whYAfNATX24CzBqpPZmamuh8DAvTcFczGL+sFAUFAEBAE6g8Cp1VGPnbM0IwavqfwQTznnMGKgOIFiZclXAVAMpFuZ/aMmSriHabK7t17UN/+/S1kFYcNFwGzvJGHWXsJwuqMqBgQNuE8lKiB7WnxY6LesXNnpfUCcd+9a7eqZ+1sv3AJYBhY+4zAIWct2bTNxzyL3QvgBwySDreDDlyNpq5IGy7NmMgaUES47z+wn+bMme009ZnKsMBkH4TUGdk38FnKmm5UygHJjeTqPT169dBONWX0UZVvaNsxWTOT8jzFe0xdAmDeb8j3iCEgrX3Y3aELuzlADh85TH9Pn67uN7jOwBKBfL64F+O57Vk8OWzevNz/GcFg0THRRlfyLQgIAoKAICAIaCHgs4TVmjQEcoRyMidoh9bUiI5G3tXdTBa2cXoilMaExjCC87N269Fd5RW1Rgc5VlEX3ZnAVw+EUIewQruFWtl4YXtS4CsL4tAkrIkyw+J4lzBJciQg7GpM/M2cS5FWR22xHKbiubNmq2pGOG5oClE+sa4IjgXkDpkV9rFGFL6/Rw47TqBfHnRVnqv2hBMfVgOfZUuWKTcVYBfGk5jeffpoB0MZfbj6jUkcNMf2AqRs+4Ivbh5bI8rs+Dtbt8W1b5txog1nVejWvbtykYFvOKqq7dyxU/WHAEUIrs9zzj2XklmzbB0EpkP2rfcvvwUBQUAQEAQEAZ8lrPtZg2NNHlF/esh5Q5XpEqf1J/ax++zjT5UJ1jjN1990A3Xt1q2SlgtmTRAWZ4KcpLv36BFWaGubsIYLEeieIq0gnyHBIdSla1eVoB6prHK53vZuJg3OBNtBWEfIpNWxD6tqc8q0vY4JP/w7gXEmp2iKjIxymlVB7cDL/0PaJZAukDuQMZRi3bV9Z6ViAdaHgcwTwK9cO+0cO2wH4rZ58yblyxrMk5j2HTpQJLuxeCo9GK675qkpiqxCq2smuH9QnQrXtjOBe4PhH260AwHt0bMHXXP9dcYilQptwqef068//ayWBbIvN4L2UBfeELgMoD8RQUAQEAQEAUHAFQTM32qu9FaDbfdzCh0j6h+7hVk3NS1N5dNEKUzkTUUOViPIBWl5srJaK4Ji+zKHZggaWWdymCO/dTWsIDWI0k5LT7NE5TvruyrrEPzVsXNHNjVHKnM9SonCZ9K2OpNt3+UuAeU+rNaE37ad9d9ICQbtLXADae3VpzdrsgOsm/jUb2ghO3ftogLu8HvP7j20Ytlyc+z4+KFlhXYa5nQzgSZxw/oNKrE+sIPWsztrJcPZ39gTAtN9amoaazMDtSYUOP+7du5iDavzKleKsHLwnbXgeKKjY1Q5W2MShPsR9xzccVCeFVkFkhITLVYPbA9tNjJZiAgCgoAgIAgIAq4gcNoxzZWtvKDtCS4LCS2hIfCxDGWiCjMlyFtEeiprUhtxIM0aFQjSuUtnFVwDX0/DnQAvWry0UZ2olIOznEkZ+8Lu3LlDmcmxjS3ptd22MQeWpDB5WLJoscMgHtttXPkbQVYIPGt8ijiifr1O2iR/P3+LllDXNLuCCevOswepFGCoYNS3X1+Vtgk+wr4ojfjcwDwffqoaEzIhLONCC2ZikH1cd85y2Fr3s2EDCGs2Y9ZPBQj27tuHlnIatb0cne9uQfR+SmoKE1bzyQSII3Kv5jPBhK+yMznIwVTQ4BvXC75x78Dkj4h/+DWvWrlCWTdQwAIa2xImpT169uSCAsGVghRRyMMdAo0y/LjdKciuYUxy3dlvdfrCZPut8e/QnNmz6YN336tOVxW2xWStgUkAaYUNHPxh/RzAWM/g9G/uEuu+bfvENag7cXbHeUUFQ28QAxNPjAd9u/O+Utl6MMP3kFQXAygeHAXZVmfI3jqu6hyTN2zr3qd9DR4RSkIaGkLjRQpTK0zkf/w+VfnNBYcE045t21T0ctfu3djXM6w8/Q67EyAzAPwXESCCl/Hhw84JK7RH0DQVcFtoj6DRdSYgDQg0acAkwt0CNwNE7nfmoJeGTMAR4JKTk0NwbXAmeMAjPy2+gZku6dq+bTvl5uQqLTRIXlsOVoqMasokv7SCltvZvr1lHY4d7hrArgmXz0Wy+23btmrll8W2CkM+GGdprayPNY9N7jnZ2SpgLzomRpW5jeXgNbgguPtBCfKWkpKqlYMVBSHy9uSp6m9mBO0gT+bgEoKcrdAO7+D0cRg7roUw9p/u2q0rbWTtPv5OT09XLhZwNcC92JAzDFgLiENpFcvAWvcTFxdP3//yk3KHsV5e3d8fffghvfDMs9Xtxq3bX3XN1WrC04knBpMnTWL3Jftp11zd6RcTJ/K93NbVzSq0x3MkK6OFZdk333/HVfdO/21ZUYUfuMbatcpyuOWQ886jV1571eF66xWT+Fgf/ddD1otc/j134QKegAW5vJ27N+jaoZMq5jJt1gx3d013jL2VnvvPi269rzA5xrnEO/Qop73bx++rvfzexWc3P4PWs8vZvLlzlAXTlQPCM6C6GBTwM+2sM8+kw1zsxV1y/oUX0sv/e6Va3f35xx90+y1jq9VHXdzYZwkrTI/JnN8RgnQ6JRw4hUpPLflhGcVkCsQCxC6I/TwRNJLZqpUiqSgSkM0EIisrS+VKRQlTaJms3QscneijHJAD8hHLAVVmhBUzrOTmydSQNZrulpDQEE6RlKyqT0HTu2njRtq6ZavKiGC2LxQPADYIuIEvpo7AzQDlSnHs0JohWwCKCKC4AHJv+pLAd7U5B+jBfaRxYGPatGmjSgWGbBJmAh9WCCZKxiTJbBuQwW1bt9Jq1vQDN1w7aUzqoPkHqXOn+LO2DDmJrQOcHPWPvMLZXARCFddw1OjU8iOcBWA3a6FXrVpFvXr3puxNm9UECcQEk5dWfG9BKxPKEwBooRvwpAjV5XCc0N6C7CIQER+8uECUqys83aJvmIRgEtqStbrpGRlVIhPQMm9kLTg04XiGzJ83v7pDc+v2uF6vv+lG1Sf8oG+59VZ68bnn3bKPaX/9pc5V6zZtXO4PeXbnz5/Pz86K5/LHH37gSQu7ZnGu66oQV9wvKCyxY/sOKjZxHUERk2+/maxiBbp26VrB9cT6gPB8hKWruvLYw4+oSRhyEuOa85Rrj844d+3aSU8/9ZS692DVyOBJA94LrgrejZs2blLuZKgECQXOjxz/AcsIFDrt2rWvUr/W44D1U6WCPKWgxnsa6SGtBaR2Ngf4fv3ll/TH1KnWqxz+NjBA6fAWLVpwrugWWsGm1h3Cxe2qq66m98aPt15crd83nLpfXe1kJz9jN7ILWW5uLiFuRKQyAu5nU5X34ZElKJdq+BGu5hfpvDlzVTEAEDncaDD9B3HuSETQI2UR8mEGBDSiFZwYvpCJ1tmDz1Evz3lz59Je/ltH8DDdvHkT+452ojD+50xAGpK4JKefVTogZ+1dWZfIx9i+Q3uLW8IKjtDeyqRIR/xOEWiQLl0tIfpdt2aderDB5QLkpB+buOGf6WuENUL54PZR1weOC5rO9ZoVqBQZY7J/HNppF8zaODfLly6ls7hQBSYLHTkB/0q+Dt1NWGHiTUgoT9OFY3MmyIoBNxJc0zoCH+/5c+epoEWktZo+fZoqGNCFiUI6vyiQxxgvTKR0gysOSCsmdSCCf7G2oHuPnuyu0439hE+oCabOPp21gT+5NXGDKfr/Xv2fuq+dbWe9buLXX9NjDz1svcjrfl934w3Kb9wY2CWXXUbvvjNePcOMZVX9fvO11wifoeefT/95+aVKwaiO+kVhlsEDz7IbPDf+7bctm425+mr61yMPm6YMNDbAJPCKSy9jf3Jz9xxss27tOnrogQfU5rA4ffrFhApECJOj++69l6b++puxi2p9//zjj4SPIbfdeQfdeffdFquLsdzeN3Jj//XHn5ZVcN/x5/ukIT9LYYWAixcyy8Sxi42R6cbS2MGPzz/51LIGz5VnXnieRo4aZVlm9uOPqb/TPXfeWekZgJgFQxBg+8Y7bymLlLHM2fcnH32s/NgxHgSYxnLRFLgNxcXFqfeyo21BaPuf2V99VnHRlTtvvY1TAlb0m7e3rTUGeC89++ILNPyii+w1dbjs6uuuJYwbE/PqypkDByi/flf6weTv5utv4IqdjjP8uNJfXW7rs4R1zerV6iZCPSnkZMXnb84FeQk/8EAUjdQ6iALfxSZMmC73cB5VmCSSkpLUDYRMAyvZ927/Pr0gkKNs0gBJQzECM8HNgzKUTfmGL2YTHh6e7hKMHxHnEGiwVq9eZeoOYOwbuUSVgHRpaljRftu2LSp4CO4HICUgIN9O/pZ8rVBrFM/u4fuLCQ18LbOZsEIDoyPKb5kfxNBMQyOgK7AG4OVqYNeufTt1DS5kDZW7BBr9OC7qANJo5l+NfR5kjTKuZVzTOoJMCsuXL1OYNU9JUbl/YdnYx/7f4azlhN8i7jW8qIAtJo1IETZj+t/UrFk0r2umdgOijHvX3YKXzR2sfZw2a6a6t836X88uDN5OVkPYJ//qa6+tcCgI3Lv5trFudVv49eefVdXAm8feUmFfjv7Iycm2S1Zt23/2ySeqna55FJMnFO+oisC8POXXX2nsbbdZNn/umWfcRlYtnVr9eOv1N6gVW+oGc+o2M9nB7lpPPf64WTNlJUAqOGTX6Mf5wnEv6QjI/iMP/otas7YRGkczwYTv7tvvMA00hXLnumuuoYVMpnTG8hMTekcTjrPOPptuZ4KM7CzOBG4qE7+bzGT6LlrERX90BdfPg/fdzxi0cUm7j+fW5WOupI8/+EB3Vw7b3XjzzQ7XOVrx8n/+K2TVETg2y32WsM6aMYuGDB2qKi/BzAoT74TPPqMLhg1TL0s8wEASkTsU5NXfv6Gqfw4/O5gOoKGdyXk3ESUNEqsj8L/bwNo4lKI0E7y4ofFq3bYt5bNJFKUrqyvoswv7C57JlYNS2KQCwgkzZhs26anKRia+7chXi+pgINPIYoCHLTQ2ugKzKwgHMER1KOAPk+pqrobl7QLsYLqG3xtMeiB1GHvz5il0wfBhyPNlKtBeNuJzir7gVuAKdnAHwEwaZB/noG//fipVFDT87pDQJqHUlq81XHMYnzPByw3nccM6zirB17SOwHSICmcYLwLWWmS2ZNeATewmkqMycwAbaIjgWgFscU8eKDmgNMljrr5KTShVajh2g5jrpmO2N25MQKDNMZM1q9eYNan19TeNvdkSGGg9mEsuvZS1rOPc8kwx+n37jTfo4pEjeHJRPrEwltv7RolmaN5AZswEGklovPoPONOsqbp2z7vgfPpu8mTTtvYaWKdPw7VqVD6019ZdyzDp0yGsuvvDM+n7b79VH5QQf+Gl/7pEvjby5FuHsOpklDHGDP9OxEhkZmYai6r0DUsLPtdcdx099OgjTvuAS9G4996lC4eeVymlntMNeeXmzZtcwgz9XXf99TTh00+rpVjqzD77sEC6KshhLaKHgM8SVmj88NJJSExUmh1oM5GeB+4BoRxUgxyQ0AoFBwdZcrOigAAeyPCf2c7BWL/8+BOXkSzW9keELyfMuDCP4uVr5scKn8IOnTry7GmpW14uIBpIW5TBvoFw/gdhBfHsz07jIBRmAqKKBwHML9i+HfsfgljoSjD7A6OCEggRjh3ET+cFp9u/p9sh2AmzdyOpPtw2EDDUksmXjsDdBEFuCG7KYPI++tJLdDZTbYAXyCqwg6YiMTGJ/U2TiNmbdh/OGoaFhXMaqY6VIvLtbVPGLyAEG+TvzTPVsBjb49qDJeJnvmdSWMOKDyZh0O4DUxB5BBoim0Yw+7/hOoH7DbJ5oGAHfDGRjWH1qtXK4mH06+5vHesH9nmwtMTdu3Zrf+ER4XTlmKvs9ol7d+xtt9Kz/37a7vqqLMRE473x79JDbMI3E5isL7vycoKGUUde/b9XVDlss4kU+rr08surTFhRwdCQTz76SPu5bmxTle+DHHjqKcG9NXL4RTT+g/fZKtRTaze61z8CZl0RVLBzl3zMgY2Y9MJ87kzwnH78qSdp7I03OWtWad0hqwAquF61a9++UhvbBbAIXXL5ZWTtYmDbxuxvawvF39Ommx6f0d9+ntiL6CHgs4QVQVLz581jLWErRZrwEIcrwO9TpqqIWmXu5pdseHiECkLZw0n/EZ2MYCtEx2PbxQsXujyjwoMd5hSYQ80IK8hNe75ZQKbZiuYWATE9zKZWCDRZuKnhdO+qgLTGKr8ic22Uo74xIYBGwBcEhAtmY+vgKlwzMG+7KsAOWlJ8qiq4FnENuUNABCKY4LRh4ggNq5lA84mgFp1AQ+u+YImYx36s3Xt050nSADX5QbDfXs67iokTUp4hgh0TO1wXv0+ZogKvcI3ifsw9paHFufCUYCKpI2apvHT68GSbW9i07SyIZtQllygtax67OblLvvjsc0LACEykZjJy9GhtwgriNXPGTOWfaNZvJ44PgGYR27giA9g/PJ4npBBMxiZ++bUrm1e5re71VtUd4Jn12MMP0/c//6SUDGb96I5Ht52xP7NczUY73e/HH32Ufv19qukxDRg4UGnn4VqkK0ilZchn7OcL0msoKYzl9r6vv/FG+vKzCdqTeOs+wEOgOILg+TaOfbnNCLmxvZGmzPhbvh0jcMqh0XEDb16DyE8EHCGyEYn0UXZ1zpw5ylcQvnUgJ9BUwMF9KQe9pKSkckqqUFq2bClXwvrBQrag/XGl+hDSIBXvKzaFBoQS2ly8AKqbl83Y2RqOLkeQkCs+lMa27vrGDQl3C+C/NTfXXd16vJ/c3C2cA3WJws6TpMnsQOBrBR9KWAPcISDeCGxQadT4mjOTYs7sgMwFugISjHsEmJUy2YWWdRW7gUBjnJiYoHxbG3HgSFiTMKXpB2mF3/bC+Quoa1dOJ8f33B62TMD0tWRJ9aO1nY1b97zqtnO2L0+tQwDRpSauOjgft9x2q1uHAHL06cefaPWZwP7SypVGqzXRu+PGabYkuuaG67TbGg3hdmLIV1984ZYAGqM/Z981cR3hXv3808+cDcOyTnc8rs4Zdfu1DMTkB54HyM6hI4iXcEWsx4q4ga+/0pu84Joe5YLVzHpMN3M6MLzvISDXq9e44Kfv6smw3nE9+23+dvNiQOBD9d23k2nan3+qmRpSLjWNilKmVrxkd+/epVIWzZ09R5GrpOQkWrtmrfJtsk6y35J9cxLZB09HOwU4tjDxQQoQM4HmC5qn9Ix0La2FWX9YD7/BjRs3KFKu094TbaAtgxZjNQfPuDvS3RPjNfpE2VoVZMeaYaWBN1bU8DfSPG3asNEl0uhsiNExscqtAS4fOmbXfTzZytWcaCD6PpHdbqy1+CuWL6cvOSIbwWrNeUI2i7VnixYsVKlYoPELYjec+IR4Qt5ZZNTAZO0P9l37joP0SjjvrYhzBG6943ZTzRN6gJYTfsPulM85UEo388f1N9ygvWsEGMI8qyNDhgx16XkJ33qYmCGwQKEkty8KJp7nclxACzu+ojP+1tcw+sqxw89YR5Cyqjry7tvvWJRTZv3AwqDzDLXuB65iCJIz5O233jR+yrebEfBpwgoscrNz6NdffqXf+AOfRKRSQelHlFtFIEhXdoRGyUqYmF549jl6n/20QFohuDCh/bz51rH09PPP0chLRqvlZv8hVU+BSSlX6z6QAxZBN+4QzB635iJN0unUI+7o15U+oF2dwwFrIK2+JDBVI70MKgbhGGpLUAZW92GtM8b4hDgmrJk6TVUbBCHiGtaRS6+4TKWKwT2CQBvjYY7UM++OG08vvfAfVWENOUI7cTW5GCapKDQBdwcQGvjr/vLTTyqCG37nIs4RQDL0UUxEdQTPO+uoeJ1tzNrAlWPS1xPNmqn1yN/a75QZVGeDjz/6SKeZSus0hiPTdeXa66+zaLd+40wBOsFgun3XZLs+HIj56huv04hRIyvtdhlH6evEKVTa0IsXbNu6TWt0yRzgWh3B9TBRU8uazMV+LraDv7P9w3cVygLIXLbwLuOYFRHPIODzhBXEYy2byT/95GN65t//ZpeA2fTF55+Xp+BhExfWwyS5a+dupRGC1hWmL2hTU1JSCNoMJD9HBCQi73UEabJQVQoBWzqCJNOuBDeZ9YkE/vPmzVWzxprWFMIVARqYP//4U+XbMxurt61HSdQ/fv9dBeTVtFsFzhX8fuE/neMup2YGGJrMFi30/JgRMAhCaZvw3dF5SklJVdkkcI+MveM2Zb2A+wz8rhBEtWHDOlVUIIzdAw6XHVYvVaS/+oJ9wWbNnEn/fvJJ+ozNmXCBcNVn1tGY6vLy2+++09Q33vr4QW4QMOJO+fD997XJ0Q3s96crP//wI7uU6Pltj2blgY4bFVzBUFkIgsm8O1IT6R6Pu9uhkpkjwXvsuaefoVdfeYVjGMocNfOp5cjVrCPueG7AJUU33uImF1JTQeF14fDhlsMY99Zblt/yw/0I+DxhBSTQqK5asZI1OT/TnJmzaAn7KYKYlBwoUaZzJNo/wjc5bno81KAlSuKZFBIGDxw0SCXmhoYoijVIOoE08I1FqdKdO3dpnZEYNtkiICCEibM7BNHdixYuUlorjKWmSCsIHnKKIlXNSjYLHywpdcfh1Ggf0PwtW7yUkHsSvs81hR32g1Rq2O8SruSzr7DILceNhOPwvYpqGqXVH/JBYsKDSZuZQFsaGRWpzNOoCHP22WezBeMqSuIMBxDcS3iZIJcr7ieQceCLe28RBzTOmD6dM3H8TGvYV1dcAczQJoXhMKuXn/kWpMjtbXfcodNUuw2yP6BilY707N2L2p3KCW3WHqZuZG/REVRDsqdptN0W+TNhSocsYLcDw3pm284X/jaCdhyNFVWg3n7jzVq1DjkaW1WWw+9eR/Ly9ug0c9oGaS5R1lhHkCJzmGbxgetuutFSWRC8A0GpIp5DwGcJK0gntKQovwrByxNaH5h8jZcx8l6CJMAHxto/FXlEUW3o/AsuUC4BUOcj8jssPEz53unAjXx3uZybTkeQ/xR+LsaLXmcbZ21AEuDL+uWEL1QZQ5AEHL8nBf3D93IO+wN//cWXnNKooErRlJ4co07fCHjKz8+jr778SuUULawBtwZgd4B9NxGEhACAbL5uMA53SDJbCZKTm1semmZ9YqKFa1dHkponq0AqtMX9A39JPMjb870TyhpVQxqyxhW5jXEdgrRCcA/CvxnfxrWJ+ww+sSL2Ebjz7rssydmXcbUh3QpsF108gp9bCfY7reLS97lUpa4F4qZbbtbayy1jx1pM9zobjLnqdCCVvfZ4ByANliEoVOCrcsGwC9U95Kvjd3XcOHc9NFN1TXFTpbLxnLtY163ipltuMT0kPANHj77E0g6ZAUQ8i4DPElZofzCzRw5NaJkMHxJruKB5xQsUTvnInWkIXvLIKIA65EZkH9ahfrJuXtFs1lIhcAVaWx1JTUt1yc/QrE9oVlExCDnt1nMCeOTW9KSAjEBrhlk+/C/hD+qrArIIwj/xq6+UNh6TGk8KztW6dWvpg/ff43O1Vk2s3LU/TMZ0XVmQyik3N4cLaGzW2j3uBQRQGYJJHe61zl27qMIJxnLce2npaVwYAfdbZTcZrIf/KwK32nVoR1Hcr0hFBPCMwgTakDdefY3eflPPvIi8qCgT6k7BxMa6lKizvgex5t0sPRw08OcOHeKsm0rrgMmZnNbIkZx/4QWWIhF4JiGloS9KAgc13v/gg7449CqPGRkmjDRkzjqB0un7779z1kR7HayDkyd9o9W+RcsWqmSxs8aoQmeknoNP/9/TpjlrLuvcgIDPElYESD3PdYOfePrfdD5XtwrnXKe2gvQ6CDDBy7Sh/+n8lEj/07lzZ9vm6u8zztCDpIBNDFv5IalbwQova+TJdCV9lt0BWi0EWZ49c5Yikes3rFfmbUObZdWsyj/RFz4wZ8Pv8ttvviHrOtNV7thLNlzMbhXfclWdhQsWWLBzN37Abh0XrJjAOS4RSe8OfywDPvj4ocyhbhTtrl27aDu7BCDVi46cgTQtrAmxFfjaITjBED8/JqxsRiu/3ypnz0CximEXDacnn36annvhBRrG96tIRQTuvOduS9EHXI+zZszgQNJflO9vxZb2/4IrgSpEYX91lZa+y1pWHcFE5kYTLat1YAr8mXU1Xddcf53DISDA1pAvJ0wwfvrUN3LOvv/xR26NcfB2AFpyEPLDjzyiNcw3XnudUGnLXQItqG4O2ls40NSRBAQ25sIeYyyrx497x/JbfngOAT125rn9V7lnmCajo2NUuU2YpJ557jlVFs3a9F/IZLWwoFDN5IJCgpT7ANbDH68ZO0tXVxB0olviEX5WGekZ1K1b9+ruttL206b9RZ9+9DH7DP5dITF+pYYuLgB5K5/hfs/uBxNUajAXu/D65gvYTA8yiUh2OOW7k7Civ7/+4nPz8cc0e/Yst2OBNG7wt4KGTUdQp30nFwyorqDgBDSmcMeB9hQFAyL5nioqKuJqWKcJK+41VB9CBo4bbrqJAxtbqvtu+MUXV3cIdWr7rNYV69G//uqrluMb95aemRHZGFCn3Z2CmvCYqOrIhTwJceT/j6BAa9/cF599XmV20em3V69eBIJjK6gg2PFUkBKeUV+xe5SvCHAaev759OLLL9HX335DKWzxqy9y8ciR9AETdPgomwlM+O4OokP+V6TX0xGULj/n3MF2m15+xRXKvx8rUZ53qpvcFuzuTBZaEDhtJ7cs8o0filxwAfjAwCB+YQdQAD+8brj5JlV9BQmJYYZFxQtUFIKWC1rVosJ9FN2sqSrn6qhK1cmTJ7QBgLZqDScIPuvsQabbQAsBrVTvvn1U6gvTDVxogICWefxiQQ7MxYsXqUpfHdhdArhYuzzodgnfNWjhlnAqFQTOIAsDTIS6UZa6+/GGdkiEv5yLSiCZPpzmUfMcL0Ikxa8KdrjWEC2PQhWIkl/ORSpyNud4JOioD19LeNnh2tKR1atX0Q7OcKErJ/lYmMFXao6qMTBjxsTHqYDHtPRUVZL1ILtWGC4ymKBltWmtUluhdCtKtQJPkAtPlrOsNFgfWHDnPfdYrjUUPlkwb75l1JhIoUCATh13+EG+9eabtI0tP+6SD957jzAxMhOk2Lr+xhvovy+8WKkpJivG8xZlqufw5K2wcC9ddPFFljRplTY6tQC+jqjz/tADD1Roci3Xozfk+2+/9fpnE8qDzueiGSEc3Kt7vxrH50vfKgiUnw1B/IwI4Q+CmVE8J6t1a+rSrat6D5sdD94z77F2X7f0r1l/tuvfYVebi0eM0Jro33LrbZVcTTBRhzuAIRirSM0g4LOEdTP7ICKlBF6MeBGCYHTt1o2uuGqMCkYCyUIQ1sGDh1S+0NS0dJWsPYI1Qah+hYcGiBkixXGT4YF65OgRDhzR92eEO8CmjZtUeivjhezstDVtGqVSaCEAC5oud/qBwkWhmDVceFkhvx1mfagtn8yBM6FMFnQekvDtLOLo9ezszbSM87zCNImqYKWcbcGdmkdnGNXGOpBzZJlAEN1O1pojIT6wg5k7nK8NHexwLcFfGm4icJtAFgfkTvREcBq0mkjYjzyYOpoKY2zZm7Ndyp1byvXDjRKmMKNBgwq/VuCBeyiKMwicPH5CaXnz8/fSIb7fILgnW7OrwuVXXqkKB4DMGOQf9yTGIVKOADSFA7msqCGvv/I/46flezwnPn/5f69Y/nb0AxrtO9iX9cH7/+moicvLURMdLi3QNpkJysW+ySZc64kttO+jRo+ybGoEpqxbu04FPfbq3duyztGPoeefp4iwkV8VwX+DzjlHNce1/fGHHzna1GuW49wgBVddl3HvvVvlQ4SSaepvv9H/vfSyqoxX5Y5MNty1a6fyix3N16uZYLKNsr/T2VJmyMWjR1r8b3M5juWH79zjY2v0L9+OEfBZwgqi2LFjJ6JT2XzwQsSL8txzh6g8oYdVoMs6OsRan23btilNVDATN5C3kOAQRcCgDULkdk82OzWLbkZlTG6Nh6JjyE6vwYN5+3YmhxtAcDpYXsqnW1T8hapX8DPr1ac3/fjd924132NPIJyIzJ7CybOROgk3Wus2rZUZFjN74IO0MsH8acCk4+iRozyGg+oFg2NBYNVO1sBBCwItjytYVDxS53/h4Y3xGClucE78eTzOpIFfA0J2h3jW6h3h8wZtHoKlDI2es2111mHyAOKKB9PKlStZq9RDaVpVOjKMlTX5wRyAhG/4IZ/gF2UpY3aIPwZ2e3bvUYFw06dPo4L8vR4j+TD/9uvXjwNO4i2aK2fHCK3vmlWrCT6sIIy6spePwSAf+4v3q2sCFV1AWKE9CQ1poq7h1NQ0zu26VZ0TaMRS09NpyNDzCME4mAhimSEgtRvY31qkHIG7WLtq4AOXnqU8ybEVpJFDgQAEgpgJcpK+zVpWWETcJR++/wG9+NJ/TbsDIbuS/UrffecdS9vrbrheadmwAIEp1i/+T9iNSYewYsJz1bVX06v/V07ar772Gks2hekc6OJKmWHLwGr4B6xf8+bOVUQHGR0QNKwzEa7hYdbK7pDX+zH2acX178rzqTqDhZZ1OLsV4llqJmNvv63CdWtd4Q2FiERqDgGfJazZHOlcVnZIkQLjgQ/SihRSF3C0bd6ePKUtO8CkBnkn27OJHIQDUc8BjVmbyhHTuFFQ8g4+ZMg3CXMlcki6Img/jUlOZqtM5c9njMVRH9Dmns3agel//qUCDzyhuUSfeEAioh8CE24s54FNZs0ucnbCnAsigeNFXk4kkt/GhAOEy0gJ5mj87ljetGkzpYHDeCAwF1mnSbK3DxAkmOqReuz4seOceD+H1rEW3ROlYaGt/uXHn9QHgU2x/HJBiifghjKlwTzhgTYA2MHEjgnRHiaDns42AFxwfYHgDxx0lsWHyh5exjJcCxjrNH6x53OUrCuSl7+HcxkfUJYIlCIGEe/Tr69KcQUMECELlxto8aezJg4TnoiISHZJ6UswT4NoWIsaCxN83LsixJlKelCfvn0VFMDmNSvfVVt8xr39Fr3ESePNBNr32++6i/557z/MmmqvhwYJQWG49s3kSrZwfchuBJg8Y4J8Gfv6GTLunYr+uCipjWshjSc4ZjL60kuViRgT7ZGjTmtsP2PS6wuynZ8RD/zjPstQQVqvYbeGi0eOUNZBywof/wHtKI4VZPDMgQO0jgbXCQI2a4qsYlB4dv/EhSx0cv0iBSZ88REIOeS88yzXK579kyfq5XbVAkIamSLgs4R1I0ealjAZhfbIdqYKp/Y+XFsa5tlFbNaGiRcpY4I5bRVuCpg2t7AqPyc7h02qEdSQZ1lFXAEIGiiQWFcEuUn/nDqV4ISNG892LLZ9gTziBkhhczNe8IYGy7adO/8GkdrMGOADsSbVeFHWtCBQCC8d6wcaJhvOBEE9F404HayDFDYTmPx4grBajwPXC1LmGKVUaxu7ICaKmWyezeJPIGvszQQmU+Sanc6BeXBPcEUwGdvDEz9MbECEQJTxMlL7ZaUpsMA1n8RWAxCP/ZzWqh8/2FHX3Z75E/cqrveN7K4iQnQXk0BD/mTytmrFCuPPSt94uULLiny3ZnIeB/S8/fqbyrXHrK3OejwjPv34E3rokYdNm6Oi3wi+tzFZhrYVVhEIMgP8/tuUStt//tln9NgTT1RabrsAmSYuYr9DWDfgpgOBq8Jc1lr6osCS9fwzz9BrPAl55bXXlO+8Lx6H7ZihRUawHjTpX3KifqSdNBNoOf/3+ms0mvMJA5f/b+8s4Kwq0z/+CooBUiJI6VCChEojKiGIBWJ3766uiu3aHWv+d+1WEAsUlRClpENCKVFAKWkJSQt1/7/vC2c4c7kz99yZAWYOz+MH7517zz3xe0/83id+z44yIhGdJK9FxC+V/VN55BBW6mQC66bIQ36m9QXrtdfsEciZJWT/u53+zeLFSzwZXbpkqd+X4GE4QkVCkM4GhzaQ9+JI7/L/XvmuJIEXL1FcD9jibq899/LSFoQmq1Y90HuCILGrVq1M+7h83qd+y3YhvamMhzwXKBqCFSJ2+ki1znS/5wEU/Ev3t/mxPMphhPgh98G/MBFMtg2+D5blFYJbRJ/taAtw43VnGG04mXxRUJcKM/aPVqy00eWayM3NlTztNTq/SWehQQEkgYkWMnGkdaBlrB3x2q77Ft/XtWx5pKS26vvtDR86zBNUrk2MIsWxY8aklUfrfxjD/+Gxadpss2II+Dz31La5q4mH/XIo1J74Xfhvro8u110b/ijP73u8926k+xsbuuSyS3WP29N3RQs2HE4TCD7jtdf7H/jzK/xZdu8vuPiiLFJCb7/1VnaLFprPcSZcofa2pEvEyZiYXn3llZEdQCj3PC8PPHJRO8pIJemvosYoRn3MtTdc7xo0aOAXp0FRYZVSi3K8BXWZQktYCV1/OelL7xUFXAgERS+TVen91Zdf+gKQutK4O7LV0d6jSms2yCmV+kHR1XfKPa16YFW3t0KXhJi/19+5MS7OAcobXUn3J3m0Uhke3VZ6YJFfmioUnmpd9v2ugwDpJBQB0A4zileAc5HzftDAgb6KPzdIMdkjEkF4H0/q9GnT1Q9+o7a/u6uofGKun2WKTKwXiT26TSuvCwuhJTd8mjyGvA/IPTJwkyZMyrcuX7k5nu31GySK/vvMVjmqVNuhq1Vg65Q7XUX3pvYdOuT4L6p2KetFpD+KNzbYh1SvaGH23JJilGpZQvzPvPCcz9NkWbzveIiTGRGMXhHDqiglBMdEZKtPRHmiZNstSJ9xfTx4/30FaZfyZV8oAL5ZqSk4daIYqWGPPZE6VzrKuqIuQ6vbqPsXboGMjGR+1U9E3VdbzrlCmxLA4H2rTk/ffzfbFyMQruShSuU0hQtITyDWT4jyK1Vsz5bU1QF6wCI59IuKq/ByMksiL+sv3TBIG/ju+9wRVk5c+qXTzajCARW86kBOJxfeQSpdyUNctmypQ8A+rkYIr1TpMm4feeiKbgn7k8MayNzk9rj30kwcLd2MjIzMVVA0BUGK440Eb2pt6VG2a9c+kwhkHng2b/CM0lsdr2huGxYQvv9O11jDxo1Eqqr4VBoK9fDYFi26uwTPK/pwL2Hao1QIVrJkKfe1rgVaizZWR6y95AnGViktYbbCwrNmzcxmbwvvx03kKSVnL+p5R4U7KhSBgd1zylHNT8PLevW117gbrtlKjPO6/m5vvOHlfEgDSWVtQh2qXktRmNK9Wzetd2shVap1832v99+PjHeU9e3sZaZKlYW0iSjyZTt7X9PZPmH0Jx9/3N12R+p0EtZ7/Akn+K5t20vSKnHfSQ38tH//LDrBicsk/k0k9a23uid+bH/vAAQKrYcVbMhfRD4o6LlN3hx5kX9Jamfy5K98z/iDRVrryivFDaGsCkIoEpkj2aZfRVxLKEWAfFcksmgAQGU3BjkoXmJfn59H7l4UI7SDyDY5foFHKbvfsX5IK6LYzZq3iExAsltfQf685L6l1Cyhqc8/PVW5bfyj0KSixOfzYlTvUwAUrJPXw9W9rJTkluJoFXW8zaRcwL8oRph54ZYQfNRiBiYX5Fhz7gfpBkyouDYWzF/gZeT22KOYLxacrxs9eqr7ldtPEmiTVdR4qC9GIA8cXdvddy/qIC1BBIGmBVyrUTvDRTnGgrLMWeec7XeFCXAU6yIiuSOsw3HHuVrySuaXURjT5+Peaa2OOgI65OVktMwclCS/Nbvf/P7b7+7t7oU/HSDx+JAQi6OR69lbWrlRjQYYRBt2lL347HORIqPB/rz91tv52n0rWK+9pkYgGhtLvZ6dtsS4MWNdpcqVXM2aNT35xHOHp6dfv75ulGZ3dGA55ph2jtZpeDT/+ONPN0sagBCexk2aeM/Tp5/093qjAdHEU4t3iIc+upzL9C9KcdS40WNEQlt6T1gUuQzIcosjWvjiMKSoooYmdhrYudhw2XJlpYrQ3t+AEqvGc7G6zJ8QGgzCg8GH773zjlc6SFfpIfh9QX1l0gRRRX4tKGJJta+oYMxTKJYiiCiG16xSlcpeKusvSXxBLiG6XBNM9mgT+rd//N01OKyBz4ud+/0cT0ZL7lvSTVZqzo23/Es5xUXUKniUL+6i+jkopCN8SxOF8RH3Jcr+FpRlIOSofmB00UllJyr/mNAnRoj/qsv/qddfUv0sy/d1pGpyz333Zfks2R94WUk9uPaqq5N9navPXn/1VXfm2Wf5IrwoK3jjtddTTuBZT7euXb2yRJR1Dhig9CulusTN0A6Pq919+x2+gUA4spDdsXLePiqv7Lma7FBcvb2NdMDP+n8a6fwj7fDNN7pu712y9WeDQKEnrGiFosWHnA6dVrCmzZspPD/TzVBoctrUaZ6YIoKO+Dli54QmIbq0iBw9ZrQXyA9IDrmBdMX61623KEevku8g9UGPnm6EtpHK8PjOmPG1l2yCQEcxpJqWLV2mtICJ273iPcr+2DIFDwFSWUhtQZotqqHDSb5pVI9mS3m9ISJM4pDpIvdswfx5fkKHl/ULdWCqp0hF23btvMeMhhsZal6wdt1av0vs29jRo73MVwN5Wxs1bpy5qx/1+tBXDZOyETe76pouPhLDcTG5zcnwWl/VZSt5/FBV1HR9StfI0T/jzDN9+lOq30Km6xxSR2kh+ZOKQQQJb+iJHU9KtWmPB6H7KIZCwiTdAyluSWWFoVFAqmNI9v1kdcLCEzldz6y4Geky12ji9MFHH/p0uFTHh2QeaTJnnXa670KYavm8fo9iAA0qIMs5WQ+1AN6g7ohmOweBQp0SEEC2YN4CX1gyUP18IaV4pOgeQ1HTtKlTJVu1h88NCgqvqlWv5oarqv+Jxx5zLykcEEgWkfdaWzf362+6yWtuIt9TS8QzimB3sC/jx45TMdikyCEGCDIP+9P0ADIzBJIhcPoZZ/oOaYHHMtky4c8otoLUjB8/Pvxxju/J964lrzXnPPq4N9x0ozv44Do+NQAv61x5IZ5/+hn3f48/4UaNGOm9JRRccU3VkceQNq4QZHLIW7dp472rXIv9+vT1ZBWFgLhZRrUMd/75F2QeVqBYkvlBwptzzj8vMyqAd/XlF7LqkiYsnuOfL6v7VRTjnLn2hhuiLBp5mddefjmS17SrvKvpRI26v/lmyn2A1M5QY484Go6L29SljHa8cTRSP65TuJ/7QhTDcYTcVZCeFOU3ictE/S1pgXTZyslI+3sjD528clq3fRcNgULvYeUwUQyg4h9NPzw/5EiSW3fU0a3cDz8sUBOB5b4Qao7CmIi/U+28WmFKHrbh/L4Kkgxq2rSZa6hcSMLX3Owp7KlUqbLX/4tSuAL5pYiqrh7i9I9OZWwD7cIjVPntNeymTsl1gUyqbW3v75GyaXdsO2G/vyu2RduOhgx44qLmAudlH2sdXNvr6pFmgf3x5x++TS0tUpEtK2zGZAaPZzNFDCroPIxKWKfjrZo00edTRzlmSCo5xaSo4GHgbzxdjZQWs2olWqzL/XVC6GyJclR3l6QVHdtIv0FFAEma5WpUQU4xZBc1A5QFRo8a6QbqIfC9rruoBUlR9jenZaI+oKKoLOS0HSa39z7wQJZ+5IiRZ2fkBneRoH9g5IJyX8qtkUIUtUjnGHnFyZfPL83SbxS6Rp4saHqQ7BhyI/szSA4H7p8QlewMPdiCZFHPN1VG7JDdjro/RdLcncjrTbHiKV9NdvdLd/ffjz4aCQ86od1x913u4QcejLQ8CxUJ7QO59FHteSkGdDj++Gy9rJsl2NZEWl1RtBujmiIvZtEQiAVh5VA3yk1PO9L1CjsStsJTuo/aaBKGpFIazbVff/vV9zZHUQBd1sRwaUZGNZGD5r6bUgAfxLXc/uV8bl+UFoB0FUI4ecRBB7oa8s4GxDdYX7JXunMhBdP51FPcEj30VkiwPTeamcnWvSM/o4NYK3nX8NSh24lBDMpIPDwq2crL/tasVdO32A3IEa8TVAg3X53OChthheDvrw5bnZULSqOFAM+c8CHnGi8A+qeEFcOTsZx+V0mTOCZ4gXIDY0WuLMVys6QwEC4mYp37Vd7Pd4z7afVPbvqUqa6sfkvBCF3UKHgkFedb/W6c0gjmzZmblpctp/2M8l1wDKmWDdoCp1ouu+8f/b8ntmkrSueb7KyLmgRA7APLDw3RV+XpfPI//wlWmePrrRL9P6VjpxyXSefL1195LUfC2r1rt1xNUshDz66inPs6RL0gGZreUSzqeRllXTkts5fuwVEMHed0jElsFNt7r9TL0R2qTp1DvDJElHVeePHFfnLWq2e09JKiuncGhhMlqn0vJaEhgwe740RaE42IyOuvvJL4cbZ/7xUBh+DHUbENlt+VX7eObAxQIPzEgxJ5qqpVqroyekD8IdKy+qfV7sdlP0q+6k/vcUM/EvWAcfKyBgZBqHJgFXeI0gjCxsMbjxGEMgph5bcLf1go79JoFcm0lHzNYZHIGgSBNIYZ0zf328ZTHBSBhfenIL/fXd45ZMLAKor0TX4fCxiGi5IIPTFmO+phkV/HgzeDzj5tRP5bqYAwWdeoZNsiFYCwPN4vHu5RrYbGi3M80bgWqspj+pWKqsITqINVeU7bhIWKXnwnObii8rwu0jlfiomJzoGfJGEFeUsnHJy47dz+vU+J4pF+WqVy6jajyVaEF/o+eVaPPW7bKuYgtSjxdxTMXXDhhZkfkx6RH8UknyjdAq9tRkZG5rqze1NHndHOOf9810OEMD+M3NvpCs0HQurhdSJh9vbbb4U/ivz+/fd6uCuvvjrpOU/+YEGz4hHPt33klNgRRmOcKEbXx3Qs6nr3LVUy0mrp8kWqHR7UKEaR4ZJFSyLlfNeuXSdzlURU07HnlSJ4rBQKEh0svaX5S7vzqFay1L5RF3UlipeIvOyuvmAafuvCAxWhe8KXeFzJYeVhStoAn9PdioctigHhMDUelwryaO2///7bHOi+qoSO0kc7+CEP9wXafs/33nOIgkMkUhmhWDwwFI41Uhi4uJLOzXZNBEooAtCoaRNV5f/Dt6FMVQgASpxja6WNSveVBfPnpzXZwTNK1CHRUNIor1QEPOeBIX3Vuk1rJwkNFS/O9qQUrytkjWuN9Ave7wyyijeFiWgUoyiMVp9R7SCRwn9efZXrP3BAUrIK/ksWLtpmdRUrVnL/efrpLI0ekmG9zQ8jfpCdIH+yn99y261SeYheuJdsHeHPkCtKZm93755r2R/UWD7+6KNtVrtR+soFkbDWb3DoNvua7ANSaEpILm57GhPdevXrRdpErdoHR84NJZ2Fuo8oduih0c4vnDHXSyMY2bMohurOU8894+qpGVBOhqwlzVUCO+e8c9NynsxWsfZQdQYMG4orr0TsMhf8rlHjJsHblK91I45ZyhXtAgvEkrDmNG7jx37hlqmt6+ENG/lc1oC0Fhdh3VspBIkzK9aFhy7Qk8xp3eHvCEGPkZcVeaAN6zeEv8r2PdvGo3WBenA3V2qC2a6HAA8dvA4XX3KJow1rsvMxGSrr1693o1WlP2nCRJ8Gk2yZ7D4rKWmmZCFCn8+q6yIIb3Gt1KpZyxcJLhQ5GyuljYJikNVnX3jee6aj7FOxPYu5R5943L3Ts4d77Mkn3P3y+Nx17z3Kl7vb/7v7vnvdk+rz/kb3N91nQwa7gZ8PcdffeGMWD354O8nSeDp2Ptn16vPxNvuEx5yuWOneU8LbC96TZxzViHo8J4yYEOeH9e/XT5OjrISDhhJ5lf0hnSBxwtOnd58CV519zfXXu1atW0WCktSoN995208AI/0gzYXIqX5U5zHpQ1GMnPWXJFGWKhzNhLWb9jtqzjcNIM4655wou+An2F2kHMBkJIpxn3rjrTfdqaef7vWiE3/DdhM7ZTHp5hqvH6GeJFjfC88+6yUtg7/7S/ZySQ7pPsFywevhjRq62++6M/gz5eutt9/u9eNTLmgLuFilBEQZz3WS4UG25zfls3Y6pbN7/VVVsir/tUTJfXUR7JN01lmmbBlfFR1l/eFl1q9f58hVo5oamY5UnjLICsugbtDp5M7ax98dnUIKonEDo2NPuZBHem+6T+nMSRVAAABAAElEQVTvVMe5o46H/cAzfoK0L5tsmQD8hSdSXm9ylaLmeO6o/WU7bZQWcrLOS6r2o+LIw/2HBT+49955161bn750VIa8J2XLlt3mMDkfS5QoofBsSWmvrpKndW/XsXNn3wKZlIOf1YRjZxn7RqtEHtB4i6rrNepDNbzPFLXxL6+G9Fdgd99/n8+DI60jOzvhpJN8gQdFa8uXLfcpHA+oGCWV3aYH4eHqkkXYHU3LnLaRbF102HtXD3A0P8dIN/pJKaXkxd7s1jWLJiw5qFE0q3PaJuTg8yFDMnMJ8V5313Z2tt16553q5Lav1/zm+kw37Qnv57DRo7xXkc6LFP1CxKl5SNeY9NDbnuuTFB3SsNJNfcIbOV5SWkQj6Wi3du0a11Vec4gf980DRGp5FqWjn40n9IGHH3L/uOJyN0vHuEET6ak6vvfeTp6KgkfztltvdU+LJEaZnHPcjzz+mLv/wQfdl19O8ilfVeTk4fqneDmZoXvc6+OPfItgoj80wLj3rrt91DXZ8hQVDhs61LVr395Hrl7WJC8noznHueed5w7KOMjftyH56Rjc4OXXXvMpB9wPFsybr2YtM9wHPXums5pdYtnYEFZuHhnVqrvGTTd7HLgZzFel8g/KYaQQJTD/cFfuHQ/cVq1bS3ann1v42y9uj6K7iyBshoMbJCLge6rgihsCxgMyXaPoh5atY+T5IsxLBXUU41iaKCTMfnMTKYi6fNyYuKBr16mdeUjkL5ZXWkXgtc78Yie94QbIjfe444/Tzel3vxeMCXqZCNwXNMLaUGFqesCjUpHOw5B8USryc9uCdbctFcwUFvCA+UPnfzDxoMq2iDyrkMGqVSr7phxTpkz211VijjUdsqjwzqie4YkUBVjoC89XigLFiPlphyo3/BoVMhUUWxrSYD3kkLqRiCQTktp62PEvami0nSr+icLk1XiIM6Z5Jay9en7gc045X/Dy0yggP4x0g6D4hQYw6ArvbOM+gscuL8aYM8kKPKEjR4zI1epoJ366pBDTIZPJNsR9PDgHKdocN+4Ld/k/r0i2aFqfcY4G5yl1BdkRVlY6eMBA9+LzL7irpWkc1YiQcA4Hij5Rfgep5x8Rz3s0+cjJXpBiAOoaSGWmOvfAiwZFeTWILv9aKsLGhM0I67aIxoawMtAnn9rZtWhxhNtNshb0+128aLHkrmarTetk/zAPCMpCeaMozqKivfkRzd36DevcH8o7DQpLIKzMgAmXVK9eQ4UkRXJV8QrckM4hgwb7nuulSpbKokCw7XBs/QTPZfMWLbRv6+VNWe2WL1mauX9bl9o57yDvkNMDRU6Y3RdUg7CW1Iycf4FBXHlQF9W/gmKE86pUreJOkie4mfrSh6vJU+0jqSdorn6uvKvg/E71m8TvIfGQyzUb18gLMdcr8LAPPFzpDPenvscD20I30nLl9vOd4ohSBMZ1coj0V/H8UUhBw40yWh5vdoUDKrj31XgjasFisM5Ur3TfqlOjZqrFdsr355111nbb7rFtj9lu687NiqkNePWllxWmPc199ulnbr06AeWHTVYu9GfqrpaRkeG6vZ4/JDiv+3WMZBILioHz4SnyOXO7rzvjunr2qacc/wqKofV7SM1akXbnX2qywj+z7Y9AbAhraT0gj5L+KjNXHrQYxJNq+9qqmh48aJD7etp031Zy5YqVXuqKvDO8hN+qV/p65dFs+n2T/x0P2gXyCjF7LV++gghPtMpH/+Mk/yMkUlmEhApjWr5GDd3Qcpb9Wy5B6b59+7o1Iq4BqU6yGfuoECKANxoyeLy6rKASAdmLanhEKXQa9vnQPImp4ynlPx6Cc9XOlUYbjdSBDWK/6Y9NfjKH9NUx7du5RZoEzpo10+sYs+8QWwp5qKw94siW/u/g+iOawb8hg4fkO2GNipEtt/0R6N6tm0L23fJ9QzeoKMfMEDAEDIEAgSLBm8L+ineJ1qg8IINQJQ9O8lpOPuUUd8tttymxua1CneW9B5aw8Ah1u0LcnwpgxIbxNPFbwi30PydEwud08UGEPi/2uR7adDBZooIv1hvsY07rxENICOpSKQegiVlaubS5SU3IaRv23c5DAM8qY9pcou4XSWuQIqso4+vPUZ1DeC3xQg37PGtVa7pH9Of//vI/wWv+lyIN6BNzDbCdTXikdy/mMqplSDuxjm+DTEenokoVqKAct7Yisbfecbvvw02UIyCr/Nan1kgK5nddm2aGgCFgCBgChkBeEIgNYV24aKF7u/tbPlcukQxC/Minuf7GG9yFqmKkyGTF8h99fgpeqmrqxLSHuvcEvc75/cYNG9Vo4Hc9dCXGLqmVxCYD6YIOoR6lfKV3VIT126+/RSKsbIN9x4t16+23eQ9cWb03iwcCiPXT6/3mW2/x8lWMdRTj/CQn9F0VMowaOSrX6SrBtlYpCsH5TmSBlIkNeo8xseKaIOSPVut65X4NGTxIxQHLvVrAxbqWrlPbTyZVifvOPm5Q1OIddSYij9zMEDAEDAFDwBDICwLRnpB52cIO+u0vG372BU7Dhw3LJJ7BpvFa4fkpLXH0kzp1dP+86iofvly+fJl0Iyf7ZgEQWv4mN88T1p83OoTw8SRRiEI+bKIREkVOJ6pR2Ut7xPeklQlpjWLsO2SAFrGXXHapzxWrLE0/s8KNAMVJp6lo4uJLLvUFOoxxVO8qk6zu3d5UK98v3FrlsEY1ztVkBXF4/SmYoeiwqCZuG1X9D1mlWOpHTewOyjhQrYbruQkTJ7iVSqNp3aaNu0LX0PEnnqiq6ZL+2krcd3JrhygNZ7YaC+S1ajzq8dlyhoAhYAgYAvFFIDY5rOR28pDs3+8TR5Vu8CBl6AhNohdIwQgyMC1UzESRCVI9VPCffe45vuKQKuvJX37lc/WoJETmak9VUc7TQ5r2rmGDcDTTeipJdP1z6TRGqdAmXQF1gk/7f+IOlARG0yZNIxVhQQYgGuTndpKsEF1KPlF6wRyRAbPChwCFah1P7uSrUA9UC98gjB7lSPB40vJ0kFpULlkSrZsUYv/16jfwHnqq0b9Qu1paAAeGjMuKFT8q0lBd6gR7e28rk7YvpSEMOaZbW6nSpbyg+0GKRpyo4rCmzZt5TVI0FFeuXOUlXYLj8KkASs9B1H6ddDlZ1/awmXO+3x6rtXUaAoaAIRA7BHZGMV1+gxgbwgow5KDS3QryWL5C+cx2k3iL8G5+L4LH51XloWymBy5eJSRT1q1b7/Xm6BQ0fdo0Vec3l8bkRl9dvkGvc/RgnD9P1dNbjBSCYyVxQkFUuXL7+4c8eX2rtY1UD2dSA2bPnO0++uADV1JSV/UbNIgsYQQhqFmzpieveH779unjpbsgwjvD/lBBDpgXJmN/IVQaqB2+20w60G/sKAmUdsr9RGUhMZSe004RYkc/88MPejn0+qJg7/NkFVnorDzutse09XneaBmiQxtU7s9ToRXFVoiJly5VWh5RyamtWeu3RYee2spdZTLI9XKaRLvRLSWvm2JCChj3L7+/PLJVM4n3Gi3LNThN11KUfczpmO07Q8AQMAQMAUMABGJFWDkgZKTwmtJDGM8Qlc4+pL5/ORVZDfOe1fqHNvDFVp1PPUUP4tUisrOle9pUHWAauxHDhruvJdq7l/L2aKH4/ezv3MQJE7OENVtJcLljp05e0gnyg17geHmtWC6KrAtSMLR/Iw0BvU1ITFTlAEgrIsnFRUB43+fj3r4VZm4ljcAsXYOUk++I1/k35TwWJvtDhBWh9kCXdUftO5q+1TXRQf6HUDq6lYlh9Jz2BU/nHJFUJNLSaSaBKHVjtQls3+FYnwuNcDzqF7/8+ov6yr/rN4nXlo5stBFFEWN3pQWgtUrUAlULdBQnjJ/gtYFJqaEgcZLI6IzpX3tFgTPPPjuTeIMrZJqJ4I48J3PCzr4zBAwBQ8AQKPwIxI6wMiQDBwzwIUoUAipWquiJXYbIwimnnup69ujh3lIhCML8l1x2mfcYkQ+4tMpSL6CMl/Td7m+7uvXrqtPUb264SC6i7IERXj3q6KO1/gz/EaQRuaoLLrpYhVmrvHRWVOmp99/r4fZW96DixUvI2xY9NAzR4djOv+ACkY+9XS91xPju++/cH5t2nKeVicFD998fwGKvOSDAOVOzVi133vnnS2v1JD8ZymHxbb5iUkQqyaCBg1wveeajGt7V6vLI0y6xjLysnKsY527r1m3cB+/1zJRJY6JGGs2J6sBEs4UeyrM+5dTT/H7PnjVLntml7u+XX67Iw8+e6KJt3FDSV6eedrqPWLBe9nPp0qVKWRjnZeT4bHtaHEJc2xMfW7chYAgYAnFCoGi5MmXvi9MBcSx/kiuqIqlNm36XIHvVzP7fCMjTHQMvEt7QiRPkNZJQe5GiRXy7NqqhafM2buwYN2LoMO/NwouItA9GSBeFgZOVR4qnKvCQ4cGtoA5Py6SXulQP9nVro4lnE8qnLd7v2k9IBJ6sdAwZooPlnS2l361atdJLZqXze1t2+yPAOdKsWXN3yaWXekWAPffaM/O8ibp18q9pedn344995X3U32XonDr2uA7uBHl0iTQE5yvElbD/aEUifpYiAB5zyOZ8nYvDdN6TdlDNn+en+POKLk5EAYg4PPnE417LmNQAIhRcB0FaA/v5viaEfXv3VqRhfdTdtOUMAUPAEDAEDIGUCMSSsHLUv/zys/f2IF9F+BVpKAgnpBRySf4gD23SB3iwr5Be5GoJ81euXEX/KrtRat1J3l44P3QPFWA1b95CHtajfCefAF2IAOuGeC5QZTV93aPar5InWqEQ6zqFZQ+qluFJQUAAUq2D7UJEyCEknFtU+0BeouUNpkJux3xPq9ITOp6o3txnq2VwEz+2AWmMsgeQSEggouyoX9BuOB07TB5QyCrnd3i7vN8oTymeU1Qx2A7GuU7qQXGlEVzd5RpfdEhjAj4ruW9J16dPb61rc9HVYQ0P9xMsyC+/I5+VyAX7yTWXKpc7neOwZQ0BQ8AQMAQMgdgSVgqt8PIsl2Ykofq/JI5ORyFyU/cpXtyhZ7rffuW8fA9/05JysSqn8b7Sn3jp0iUisSvcL9JgDQzFgGbNm/vvCZ8mGh7PGV9/7QtOEr/L6W+8XCuVikChV5UqVbV/+3gCnNNvwt9BwsuW3c8Lz1PN/aPI98b1mz1n4eXs/Y5BgFC8l60643QRxpNc/fr1vBc8na1DEpn8vPfue2740KEOL2e6JJBz9bgTTnAlSpTYZtMUcM2aNdOT1vCkbD9N7uhX37HzySpAnO6JKK2OK6rrGqS1RcsjXC2lN3D+sz9cW8Plle398UdurDrK/bh8edr7uc3O2QeGgCFgCBgChkACArHMYQ2OkQcqupEDJQG0XA9SvKgHHZThJXpIDyi7X1kVZh3uUwMoaPlRUjzFRErr1a/v+7qjvQrppUgqMNpWykGV1MqJAEMcfUvLNKrnyXlFZqhv7z6eUOMVO0g5rZDrqAZR5Vgg4ZCJEeri9e03MyVRZKHZqBjmx3IlSu7rNUtps3pMu2O87BnnQzr2y8+/SAN1nvtEEm2fSAmCoqh0yaqfxOj8ZpKWzIrsViRLmgDL8JvaB9eWbNVJfvI0Yfx43/K1UpXK3oPftFlTr5yxWHJaRATW/LTGF/wNGjhQ5HbqDs2hTnZM9pkhYAgYAoZAfBGINWENho0Q+UQ9fJHaQbqHdICMahk+vAlZHK3w/zSFNEkBoLtVnbqHuA7HHedatmzpZXvmzZ0TrMo/lMWDvRFKpeMQnthK8kCRn7ifSAIFLng50zEICbqw3bt2lXe1qHIPj9P+ZURWD2BbkG1yD8+/8EIVZVVU287+burUaW6NUh2iFoKls8+27FYE8KrSuarBYYe5TpKtaiMlCZQBwqH4rUtn/w7PKmR14GcD3FsqBgxPlrL/1bbflFUhIGSVkD1pJ0sku1bhgAqelJJyQsSBIr2ACLOfVQ6s6loe1dLVOeQQL5nGNTNfXl5SVvge4s3vSFPA+0v6yRJFIjZYvuq2A2CfGAKGgCFgCOQrArsEYQ0QQ3KKf+TuZWcLf1jgur76mm8+gPwQEj7LVUhFtx4e7oRS//hzczU+odRZM2e6l1540d1z/32eDEMS9leObLqENdgftvP6K696MtFZqgbpqAewDogFUlmdFNKlUKbHe++qYnuwkdYA4O3wGpDVDgq/n3X2WV7uLDebYQIECewnwf23lQ+aW7LKtsnbxtvPObpUxYAP3f+Au+7G673uL9+TMsO5/NeW2dc+UqpoeeSRroMk2tBk7fZGV7dcJBejLfHng4f4f/4D+58hYAgYAoaAIbCDEYhNa9b8ws0/4JUa8PKLL3qVgXbSr2wtwXVsk4Ty58z53heh8DceJwqd5qmpgCeyIgdlypT13bT4PreG1/YDyRe99eabXtMyt+upUbOGu/yKK1yXa7qIvNbK7WrsdykQqFOvruty3XWSffqH70aWYvFsv2by81b37q73hx/liayyATq6MXn6848/lVKwRl5SPP67ZVb0//oLuq5zvKIGy7c/tr07tsNxPmLwwnPP+1Qa88qDjJkhYAgYAoZAQUBgl/KwRgX8d+mvTpLkFQVUhzds5Nq0aevm6eE+S57Z2bNmu1Xq7vPblnD9biIBOKmQ0uINaQHkAubVVsurNXTIEHXhWudOOe1UX+xFyD8dg1AjO9Tu2GN9isAn/fr5dUKIzfKOAJ5sOp6hXVqvXj2vHBFonaazdsT2x4wa7fr27eO+lNwazSzyapyDpCT8T//hueU8LaLiKTzwePHRVV2g1AMmaA0OO9S1atPaE9yJEyf4pgFIwpkZAoaAIWAIGAIFBYFdmrAiRcXDPMjjCwaFvyl0QZOyQoUDfNvW0846U2Lq77kflLuH8gAPfbpTkW9aUlXYpA6QGwthQdc10chrrVipsie0kN6fN27YZrvh37APy+TppasREl0/rf7JHaEK7TJb8hLDy+b0nv0hPNxEBTN0PapStYobKZH4udLctE5EOSGX/XeQwWrVa7g2x7T1EmcH164tCajiaeercu7RMniMqusH9P/UTfnqK3/eZb/lzd9AOpGeqq3tIuS/ZPHibX7HOci5uen3TT4Pu7jOUc53DDUKzi0aY9RR21U0VWm/umjRQl/xn11OKr8vr3QXWhgHKTKb98j+bwgYAoaAIWAIbF8EdinCSsehWrUO9iL9iKDjgcKTRHiUBzCkc+HCH3w4FNi/nDTJlS9f3ldNt27d2ldFv6OQ7XcinPWlJAAJxYu5n9q+UrBFcRdEM5EAswwNCloccYTyS/f2TQuGff65bzCQKuy6RuRghCr+f1JF9s8irqyjUqVKfrtRTw0IDsfbqHFjSV9V9F2yvhj3hZs+fbpXJ4i6nl19OXCsLMLfoEEDPw5Ht2rlPdi58apyriCjNnbMWNf7o4/cN9/MiFRlT74sjSJotdqwSWNf8ESji5EjRmzze85Dzu81a37K1CFmDFdKiopzmHPihI4d3ZHq3Ea72s/l0Z/85ZeZw1xO+r4HHniQ/y3eZEg5neOYPKG6MVfKGt99pw5rRBfMDAFDwBAwBAyB7YjALkVYS4tgItlzbIcOEtsv74uTeNjiMYJw8vD9Ytw4N33KVLdCnaPWiiT2Vxgdz9K555/nq7+nq4f6NEn4ICFVVV2x9tijmG/NShrB/1TIQuU1Xq2wVate3R2n0PEx7dt7DyyV5OS8jte28OSmMtb5lcjzKgnHs6+E+GluQE/4dAxiVVWdv+j9jkfN96WXQgLEiVarZtkjAFk7UON9VOtWvmNVfZHW3BBVtkDV/g8qrqKQqVevXm6xxPujWumyZXxx1N/+8XeRx0rKdf3d7Ssprfmq3IdABsbnkOI/dU6iPFBOkypSSpDMIg/7a9JdpDd8/Ikn+DxXpN8+/aS/3zcmaeX2L69z/DDXvEULtWet6SdnEGWOmYKtxfLqsvyyN97w106wXXs1BAwBQ8AQMAS2BwKxbRyQDCzasp6gfEPIBuF8PGY8gGkcQCcsqupbtDjCNZLnipy/1fIkEbJdv369f9gjml6jRg034YvxvroaT+fu8p5+LU/ln3/+pTaptdXecq4bLzmgZeqpHtilf/+bO7pV60wvVylpwFbXembM+NotXxpdaB1yC9GAXCJMDwHnGPiXjiFrBNlpcGgD9ZqvoVDwInnNVpunLAmIYAtZbdi4kety7bV+wkPqBxima3g8IXtzVIXfvWs398H7H6gCP3r3Kjz1jZo0cXfcdacfP/6GhO4t76fc+v68DPYJhQCiCTTIGCcv7h677+HqH3qo1CuWe4/9BunzXnfDDX49H33Qy30uj/8aTdpoTcyk7souV0uL+CQv8cY68LByzODBK95ZzsOZM79Vs4D0JNyCfbRXQ8AQMAQMAUMgKgK7FGFdLVJWXd7ODBVMQVKTGQQWTxIe1FbypiErRQgVUrpu7Tp3tD4jZxGtTEgeIVPyCP8nfcqMatXcuLHj1J5yqPtdeq4YD/iLL73Uk+Egh5DPEPf3bVml10oqQlTbJM/Z4oWLpEww3+0nUkKIP7eePsj2AcrRRc4IjU6aLECKIVVmm1Ug6miS87d//MNd+rfLlLdaXc0c9soVWQVPPJ6jR45yr7z0su8KtVFe9nSMVqunnn6aT+0IyCO/h7TiRf30k08yV+e7pklXla5U36tgEH1eNIhHq7iL3FW0eqvq+zck4TZixHBflHfu+ee7f1xxuU8RwMtKK2LO1WS2XDmwo0aM9Hqx5OKaGQKGgCFgCBgC2xOBXYqwUskv/5AjNw9PaTLjAQ2xJNxON6xKCr1DWncvurv7Svl9+2f+djff5Wfx4kVecJ1CnJXyln0xdqybPXOzzivrQsD9tDPO8HmnwcM/2Aaeq5VKPVigQi4IR1RjWTy/aHYS+qWoCi9bul4/ludYad2JxxXB+BLynPkuRko92JWtes2aPgXk3AvOd43l1YTAFduzWNoYgyGEDo973759Xa/335eXfHraYvu+ZWp7tUzt1Ml7N4NzifUzjkw0Bnw2UCkpm9NRSEvh/KWZwd46lymWGv/FF151gslYNU2uJk2c5IYMHuyOOLKlcllPck2aNvWkFo8yk6DwNthOYExoRo8e7T5VoRi6xWaGgCFgCBgChsD2RmCXIqyASSi0qMgnJK+CHuKJRpETIdT5qqJfLw/Y3nvtLe9TBU/oICyTv/xKJGB/31ud1qn0TicHlLD6lMlT3LfqpkVzAowCmTL7lfFtXpNtC6IIESZ/FrH2dIycWSq9f5RXlO0VK7anVwEg1SEdg5RAeNgXSA1eOPJ7i2k9yxXqDQhQOussrMuCxb7yfLdp29aHw9se084rRJDCke5kIMBg7Zq1Km6bJnLX33sjv/l6hi/yC76P+tpWbV6Rz8Jjmkgk2TcmMAM+/dTrqLJO0g9+U0thtr923VpftEeaC22H6Wi1SHmzQ5UG0FxFfEeroPBA5Tb/pqgAk6dvZnyj4sOFPtIQzpMOigmnKY+b3G4kuPDWmhkChoAhYAgYAtsbgV2q6AowIYfjxo1VSsA+3otKLmk4pA5BW6HiJiqr/zfbaZk9XUmlCJQuWcpVqVLV67BOmfyVrxI/pM4hPrdvxLBhCvWOFnldtg0Z2SQSkF2Ine3Wa1DfNZ3bzO9TdnJC2Z0EFIzNUKrC0sVL3DJtu23bY3x+LpXckJpEYpPdeoLPISfk+ZJmkFEtw3sVKTCbN3e+2yiiH2crXmJfed2ru8MaHu7athNRFQ6lNe65MYgd/yhMokhvuOTRRkrpgclQboz9aCrvZ11pvWZHnP/663/e2x5e/9JFS5S2ssaPI4QXWbRKklZbqnaqUyZP9utigkKKyTczZngv7W+//qbPd/P51uHJCseDp5jCxH69+7gJ4ydEKhgM74+9NwQMAUPAEDAEcovALudhBShyUSF5COiT2wdRC0KgSF2R21dCBAbpnqlTprgv5HH9ShqZ5O3VrVfXzfx2psjHWl95fYjC6BRATZo4UURXDQV+2xraDzxSnTp39t7LgEBCYAnhEo6nmIWiru+/+z5LoVY6A8px0CXpB3nF2AYV4XvJQ7qbPG/ZEZyc1k96Q+UqVVxDVZEXL17Cd136XUoFEJi4SRiBPwT9cOWHnnzqKe6c887z4XI+z41B6iB9S5UCgMezp7R7x0pnNZ2Uj8TtMg4oQzC5Av+gCDA4n9jm0iVL3Pvv9lAxYNb8YxpZ0OXsKhVRHaTc7Xnz5iqPdZS0W5dIKaK2GzlyhHKuh7mvp0330m6Me/Pmzf3Yk1LA+cP5ikd/sX7TvVs3N1hpBD/qWjAzBAwBQ8AQMAR2FAK7JGEFXCqcZ8+e7R/+5PMR9oUA8ICGsB2gDlEQhWOUN9jiiBaeBK5Q+H3QZwO892za1GneY1ZHXlYILrmuPyxY4AuofNerLSMIwSCUGy6OQrCfrlPkRSLo/pcUBtgfSG9ebIUKuPCIfj/7O1/kRWEXpDggNumsGxwg8hB0pI1Kly7lVqnIbKW8z9l5jNNZf0FYlpSNQyXddN75F7jLJBOFTi0pFbkh+YFXlRA559VjD//b53guWbI4z4d6+plnusZNm/hCPfAn75TUAIrmMMgkygP9+vTJsi3O48MbNXT/vOpqSVPV8jnP77z1trpqjfIqF5MklcaE7fgTTnAXXHyRbyDQWBiQGsLEjfOG4+IcpsjvpRdecJ+JhAcpL1k2Zn8YAoaAIWAIGALbEYFdlrCC6V/yTJGr992s77xnFHWAgLgGmPPQhlQiZdREpOFECa13OK6DHuKbfGMBSCphdLysdevWU+h8g/Jf5/vwabCOWrUP9qFYCCRGk4InHn/CEwNIKz3fixYt4gtZUA5IZjQ9CDy2yb4PPtvc2WiF+1Lk909525Dr4pjyYpCXDHnnGknaCYKzatVqpVasLrTEFSzr1q+nivgrvEcVySrGJjdENYwrRXB9Pv7YPf/sc/KYfyfN05/DXyd9z/lF/jFe0kSDUFMMd9ElF3vCiQbvBEmm9VVIvoN0fRkXbLG8q2NVBEVRVWB4iI8/6UR35dWbyeqypcvcq6+87HOzjz/xREl0XeOLyo486khXo9ZmndUgyhCsg/ON85viqm7SW52onNXszs/gN/a68xA4oOIBSmlp6HPQ6Y6XF69+qqMgAnWwtJwPqXuIvw9QqBk2NKpprMI5a1b4ELDxLXxjtivs8S6XwxoeVB7IeIumTpnsVqta/ysVVFFBjd5qOVX3ozXJQxzZIP5RmIS3CbWAeSrKWiCx9hUqTCIM+8+rrnLkfXZWWJn1Dhow0IfS2d7kryYr5NxIYfbKfvOE6stKAH6GCnCQSqI9JrqZlUVO1ip3lkYBgVEd3qZNG3dUK3UjUorCFK1r2tSpvh1nsEz4lW3jrf1OXtYPer7vCQpFRE3VaYvq79wYGBAehiDRGpamA3iDhw0d6uYolYGUhMJgkDg8jTRwgHwTIuehytjmxWgCgTYveFB5jxYv45CTVVaRE8L8DXVeVKxcyX2l3w2QeD9tVgOjvSoheiZSaK6SZkB3srKq/OccCgwFAs7JwPCsHitCSypKFW0HndQXn3/ejVVHrD+kHjBv7lwf3m8o72siSYU4r9M1gbeeJhpTlBLznTzG5Lky0dpR1vWt7rqequW4OTz97Vq1znGZgvblAw8/5IvcwvvF9Xzd1V3CH6X1vmPnk93Nt9zio0LBDxnH/pI5u/fOu/Lt+uRaOevcc+SJP0PXzkHBpvwrEn/fKi2J1JLRo0a662+80W1Yv8Fd/re/ZVnO/ii4COzK47s9rsuCO9KFd892acIaDNtmAXTlgMpD9r0ezrNnzRIxK+MyRCYrqOMP2puEeldJL5UWqZNVdAVxhOzyj45RZcqUdceoWAcShIwVhTKjR45UGHahL4yi+hpvBKQRZQDSDSCsEAPyYCEkSGYV+baodmsrYa2kAqpTTjvVtwNdq22xLB7dUco9nCSPV3ZGy9c58vJR6MM2yHFEtgiCnEhSsltH4ueEy0mVQGGBV44VHCargGe2HlbojBZEw4NZ+5DaPk+VsD+V8nQKA4fcGoSUycvcOXM9eR+lsZ42ZaqaTaxKucpmSrFA4/ewwxt6z3XpMqXd/pog0Qo4K2Et6tNG9hShxhNLSgrV/VT2cw5hkEsULb795luf/lFVusBHtz7an4uQXYqpUAMYOvRzL6XFer4UuYXskffMtvHCM7Z0waKxAJMx2q+Sv43XeGdMSJ5UBIL2x9dcf52fPIZBRbf42aeelr7s1s5e4e8L8nvSOZgUM+kLLNBsDv5O5/Wqa7q4a6+/3v+ECTQGmeTcPlkTFs7TW2+62X+el/91OOF498BDD2UpRGQCM1YyfuvXrVcHvc25zy1btnSX//MKv6lx+s4s9wgw8ez29luuZ48ejuYe29N29fGNel3uyDFJNd4FaV9S7Wt+fW+ENYQkN2BkqfiH0SAgKMDiu4UKj67Qw3Lt2jVZvKCQ1p7vvuvJ6wkKt9Ixq3wFhfr3KysvR3/97gefWwphpdKbkHQ9ddsaN2acCM5qHw4mJWAvSWjtpgrtwDghq4h0NFCHIghFBf2NPBZEsbzIbbE99nRT5Z35eeOGbD16tHMdoQr1ud/PkfTWYnekmgTQ3Yo0hNx6FnkYsg9472oLI4qBCFXPmjlLKRYLsmATHMvOeAXnqlUPUremOupg1kJ5oCLs8trlhahyHExe8ChBmGgEMFyeVWTNILDZGUSx+L4l3OEiK51POdU1a9E8U1YN8giGXk5M5Pp3yVFhRfQbvMJ4U5kMkIaxXl6rBjp3ghQRvKVTRZTx8rKOjid3UuvYDl7i7NtvvvE5pwM//SyTdEJgaAW7fNlSKVOMc+VFlOluxXbWr18nbeEf/EQnu+PYUZ+jfsE/dGQDQhZs++3u3V3P994L/ixUryOHj/DXzT333Zfn/a6p+wwpH4Fd+Y/L3dy5c9yIMaPVCOQA/zGT6Lxal+uuc1eLGHMOBzZkyBB39223e9WV4DPub/c+8ICfGAaf2WvuEbjz3nv85IaUnO1pNr5OKi7RrssdNSZRxrsg7UuU/c2PZYyw5oAinlb+RTG8tH1V9ELeGKShoXLJLrjoIt3ki8jDNUSdsr5WnusXnoCSE1tNhI+QL0U0tMTEJIaUZVNl5PkidzashckCeDePO+EE/90Lzz0niaKpOcpOkWKA1+xN5SGOkSfwjLPPdk2bN/NeRkhxbgkcv4OcE7ptKfH5QQMHem8zck7ktO0sjyve6lLyHCJDdmyHY0XgjvVFcXyeFyPMysRlsYj/hPFfuPd79nTz587LkagG26MhA95dWp4eLMwgiIGRO0v6CSkjFLf9KI94YEFqATqraO5yLgStYVGaIG/1G7XrJa2kw/HHu7POOdsX8eGt7de3j/t88OeZBDhYJ6+cE0xm+BdM0MLfF5T3eO8SjeMuzEaoPD8ML334nF6xcoVf7Uhd42eedZZ/H/4+N9vE85ZIVvHa33Td9VLD2KqIwrrx8l964UXuvQ/e901IcrM9+81mBEjpOUPFltvbbHy3IpzqutxRY7J1j7J/V5D2Jfu9zP9vtibC5f+6d7k1+up/dTN66YUXFYId6nNeCY9dIw8FxVWfDxnsK7zxVOABqV6juvdKElLHiojchq2UCoHKl9+2uQHLQDQPPewwyRV18fqh4d9l9x4P4EyFgR958CH3xKOPupEjRnhvYXbLR/0cJQI8rrQw/e8zz7gzRYjxDO8sO1D7cs6557qnnnnaXXLZZZ7gsY95NSYjaKo+9u9/u0ceetirMeTkVQ22x3jXqFlT58G1rp487IxdMjugYiVPtMPfBV6tyZJVI5f04NoH+7xblkE2a5j2Bw/pdcoZpH0sKSeEt3xFvzwzgbc2vM7C9D5Rpot9/99fWSd2hel48nNfa9aslWV1bY85xv+9Yf1Wkk9L6dwa6SIP6VwPzkHWs1oFl1dfceU2ZDXYBukj13W5JmkRYbCMveaMACTysSeeyHmhfPjWxjc6iDtqTKLsUUHalyj7m5/L5P0pnp97U4jWFdzEAw9YsOv8PVOh2PfeeddXiVPwxGenn3WmWyAvJy0t6bR1nG5KLVoc4fNhqbz+Vdqd5JyGJbH4nBzFZMb28XCiTEBxDd5MQsBRjPxWOnYtWbLUh7Q7qWiD/NbcemMCLPAUkgpxoSSSWrdpLe/feE/Sp0sCbEcYuYHt5VFtKh3RqkpXCDpUBfuXm31g7MB2ooTy+8pjOW3KNB9OTxz3nNZd/9AG7tTTT3M1RVpzkhmjovpn5ZEGtklKFIR4Icr8QxqtZcsjRUJ/90oBEyZM8F7Vg0RYmSzhmR2owi0k02bPUtcLs0wEmCBeogKgQzUWpNXMlWd8yKBB7n3lByYak4tLRP5p34zEF/rLeBXnKv0CCbrAKEpboNx0jHOu3bHt/W8OVB7xOqVXTFSaDA0jvhi3Vb3BL5xP/+OcCBve9R/UKrfjySf7j4kGPPLgw+FF0nqPJnGgbBL8sJ8m5DSeyMnAhBShZHa2JpLNhJVvd61JJClRA5Sy0vujjzIjMmB+yWWXZvk5aThP/+e/mvDVd+deeL5rpiJSIgPd3ujqoyhZFtYfAwcMULONZsrnbuXbYpOy00OayFwfySzKfvE76hNq1qqZZRWM8cTxE935F13oTpCEIU1pXn/lVV2jvTOXi3p+XKwJ9i233Zol6tW6bRuds+X9uniukFeOpXNO+x8k/C+/x5fr5oyzz/JNV4IIEl3z3lIKD7UhgV197bWaWG+NMPH5l5O+9Dn0p55xmm9+Q73I4IGD3Ltvv515XrBcOdVP3HDzTX4btCOfocjl7NmzfL72Hbfd5u/5LZTjHzaim5+qAPHCiy+W6krF8Fc+QjVi2PAsnyX+kc6YRB3n3GKQzr4kHkcc/jbCGmEUixYp6nPpGsijWUkV3XhLuXDwiCLiTrOAOcoRpTsWkk/LVOBEEdL7PXpKH3OuO1oV/qVLlXZKoPQesBUrfvSFOmi38ptffvlV3rNZEv5flMUzQZoB32dnEMS999nbF9jQppWLPBxOzu53fM66F+ihDUlC8ogqc8g1RVnBzSan3yf7DmII6aWIp7huJvurYK1+g/q+uGzkiOFquDArx9SFZOtM9VmJkvuqEK2uJ8iQdzy95D0mplGkWk+y78GIcR0xfJj7SjdUcnTBC8If1SrKa3rkkUdp/9qoeC+5ZzVYF+v+ZePPwZ+avFA4972bp3MIuTLyb8urTTBpAORRI/APxrSTJY/6tZdfUSHeBF88WEZFfwdIdWI/5VGXLl3Ge/P33HMvn9uKnBuFS6QYLFm8SPnV09xPK5VLncZxZe5kIXhD61k8VuSUk9ZBJzvC6fxDY/imLUVLHMqlf/+bu+GmmzLzu5kwohjSuEmTbY701pv/5QsaH//Pf5R20t5/T24zD1IKn8g1vuzvf3f33XOv6/HOO9v8Pq8fUEgZNiadPT74wH/EOYIW8DcRJ7Hh9QTvk4Wkh6uAL4o9rCjO/roPBMY95cmn/uvvVXzmJ4I6pw+pW8wdedRRnuyTg8vnSGKRxkPqS9hINbrjrrsy7098T0rOsR06+N+ElwX3sLHOI1QQdvcdd0o9pWfmV+nsFz9CAu+000/PQijpFsdkiPtnYNffdKMnrKw/6vnRRh7y2++8I1hF5iuFbPzDKPCFsKZzTmeuKOFNfo4vjU0efeJxn9rEZog88Tyg/oKi4SuETzBxw7HAMzHsSLhAqSRFdy+aiSv3Os7n/XTt/WeLtxm1lA/79M7M/WdCltGpozvJdfRHhtIOToqzzznHS1EGh8tyEFYUYs6Q4whHT2Dk/edEWKOOCc/eqOPMtnODQdR9CSY0wTHG6dUIa4rR5KI6WB2BOFloj1lWN2HyDCEJfEcIbJM8XsskObVRuWkQ2OXLl6lye76XIZqmSmsKppiVk6fIBQiBoMAK7wLFMtyIf+T3G7aG8tgtvlu+bLl/5cYHQcUgUmyH9RFeZtZIMRUn6iB5FqIaBIULbfzqcW7+gvkiL0u894NUA4rN8hJGh5xVUeUwHtfqIsHot85QriWV9Ajrs928GHJfdTQuDRoc6uofpn8KtYNtbr3E4X3Bo7pABXbTVdCGF/MLFccxPrkhdA2bNHItjzrS5x1TXMXYcQNFIo3wPQaJIi+TCn3GPDAe3qg8fK4CFyYBjDM5z4z5QQcpB1o3ePYJrdQ5KgCbInm2ipoEUVwGYapQ4QC/nRIq9sIbQ5EdRJ79gOCu0TZXr1zlaiofl1QHUg7YZpyM437q2Wcysb79llv1gBrmPlPaBAT2JD3whmtCQrvZo1q18vJQwQMNabgzTj1VD5fDXPcEwokEGV678y64IJOszlXzhhM7HOdq1a7tevfrm/lgvOOuO90opd8s1vL5aXh8k9mF553vvbt8R5rOFVddmWUxxv/mG270KiZZvgj9wYOfospE+2ZLQWri54l/41ELazg89MgjmWSVe+aJxx7n5QA/HTjAkwvI3oOP/NvdpUIuNKSvvPwK11/fhY30hETjeumoQteJihgV27NYlq9Hjhip831Pf08LvrhNhJBJHTn9WDr7xfJ3336H9zofp3zxwPCqJhrXN5bO+YF281133OEulz50mKwjt/jRh738PQLHQjrndOJ+BX/n5/hCBJ9+7tnM5wWe7PvuvtvdoX9oSPP8e1qyepyXOHIulxf57R7veUIa7A9jh9OFySQTvcDwtL/+6qveUXDFlf/MJKsoUJAvDVF+/P+e9Nc3Xls89dwTn/i//wtWkUmM79Vkh/Ohs67pwHZzuwVvk75GHZO/X355WveB3GAQdV+SHkhMPjTCmmIgy4oEMTM/XaEgCFiiQSQxwrWBeXkgeVHxjKFjSQESXkxI5l7ycjUQwfqfHhoLFb4jBAwp/UIXYKJBbPCiIFtUR3JWAWFlPZy83PjZNwjMIfXq+hszxTbpkkEu8CV6mH7Uq5ebNWumPE9tfHcvbppIWLHd8Gw4cT9z+huShCeZfxBhvKAT9UCCvEK0VquYKCpJYh8YjwyRMaSp0JY9XNJQyIHl1dgHHuRIgPEwI9w7asQopVnMiLx/ifuAJx4vKAVRkFLI6hilhJSSt51OVQFhZbsU96GzmyznlNauqAqggYsWK+tjsjJ9mtqp/vKzVCI2+lxkxg1yhdcAmaycPLo8UGlKgFXR5IRziWP/aUsBYOKxFNa/b1GYMMCZwjW6gTHWo0eP8mFEjuufV17pCevpZ56RSTL5HOk4Ui3Q2MVzCv4Yofh/ifAxsblPVfGBgTuV8hQf8eClOBLjGuChnl+EFcJy6+23++YRwbbDrxRAko6ADVLaAxJ6pMpgpJMQVuaekpNlaIKZaH5ipQlUukYY//gTT8j82XR59IO0AnL9O21JYcDjRxiYCUVY3i3zh3rDvh+h48tQJIX7I5JtvylCtUYRB663wEhJgBRg78nrjOYwxrlwvULK6N7mZr9YB9dJojFZmarUp5M6nuTHG08oFiZHqc4PjrtXz4XuLNUAhAkr6Q29er6fucl0zunMHyW8yc/xBc+wc2PAp/391rp36+YJK39AkB+UBvHZ0vDFiPCEjXvfOXrG/qTz6ys5CoJJI9cOzg7kE7nnB8a9n5SGHlLnefiBB92/H3tUnvfNDoAVWld2lrjd7JYLPkdGMMqYpDPOwX0gcV+iYBBlX4J9j+OrEdYUo4oOKjeacvuXy3ZJHoDczAnbUunP34gwl9JNkpD4zyIDPPTwiCBdBKGcKU8FVeyEPn7+WeF5PUA2KQScKFG1SrqekCdC3dxsIY94EVfJM9bt9df9+mofXNtvD1ILYRk6JFrYLvGA2G+EvwlD076T3FhCNzyoIebhm1Lib6P8zQP8AOXuQr4QS6fN51TdiFasWul+FXHProCJ7RJmBbtGevgiHVZPuKJ1mx/GdiF9K4XpSJGUT5Snh3ctsQo63W3RThW9TbzxPFzRPX35+Rfc5fJ47VFsD786yCodscj3paAlbBD0fYrLOypPaUZGNX/jZl8hsKyXiQteerz3J3c+RROXvRRW292fI39qvWwTDwLei5wmHRWEK15wxiduhDUgauBKVzvOcSw8qSMNZk9hxyQibFxjgeHpDggrXupD5NEneoIUWNiC7mOMa9j2lGcnP4yCi0APlfGnKQmpDWHD2zP086E+QoAUX/c3u2US1icefcy99eab4cWTvkdXOtEgEUH6SeJ3Of192lln+PMvWGaxIjmBLVUefdhIvYC4JbsXjNY96f577vGLIwu4eNFCKaxsJtCJeNOCOLDP5U0PCCufBRq4udkvfv+H7vWJRv4kpOpxeYErVarsZsz42i+yPc6PdM5pyHwyy6/xpSNWuwTptHnz5vtNQvZ4LgbkE/1wnl1M9MK54CxMoShOHYyJfTh3em9FI7FgPbznmXDfgw9Ik7qFu++uu90YRUqCnO4/QnnmLBu2P/SM3R6Wm3HODQbbY98L0zqNsKYYrW+mz3CjKo2UfFMtn3eSbHFuloRqCSPjXaVAgNxT8kkJ/5dRtW0lebNoRjBFigAPa1b4D6kH0D6RixTiesO/bna91dZzuB40EI3AlumG/uEHH/ocSEgrM040VOvKo1pBXl2kqq6RxAyKA3iAW8jjmlvCGmwTbxui8SgKfPj+++6c88/TQ7F1pkcuWC63rxDNo44+WoUTzVU48Y37+MOPfPFX4HVJXG95ESkeyqeedpokoWpnErDE5XL7Nx7u4XpIvvfuO+6H+T8k9XLmZt0k/wdqCXjK33jtdVdN44S+Lx4HDI8XKSFIUC3fcsMOtsUkoV37dl4mDY8TEwk81SVLltKD+15fXNNED3jI75TJX7nVP/3kH+AUaPBA4lwhx5nzBk95+IYfbIPX6dO/VneiUX68w58X9vccN+HIwJATGyCJOSxD3wUGmee8gtDiCU1qu+2W5WPyjbGH73/QS5VVPehAr8dLwdUDUpDgeg9b0SJb8+bCn6fz/nBNgP/z1FOZE8enlDv7Tve3HF3BiLQEBil49oXnvceK+ws5ohjkNgpZZdkfFZ5NZtxnpk6ekuyrbD9jUh420lECS5wUUqCYnRGdCAyN3pwsrC7BuIaNgkwmFvm1X8gSQlYxJnzhSV9+nx/pntOkNSWz/BpfVEuYWIeNyV1g3N8CVRTuR9SBfKUoYE5Gjn3YglQPGu2Q1xo2nBdE7kghId1gZ1l+j3N2GOys4yso2zXCmmIkCJdPUAEDFxskg+T9xAc/FyytSyGrhAJx9VOtSwebokUX+mRy3jMDZMZZUTPwXupcUq9+Pd+5ajd5LggjEqI6Sg+XsWPG+jwrcnrwNCDy3vX1N3wu2kEK07M95K5OUS7Ok0p0xytIbg75tRl6EPNwZL/zajxMKBp7+aWXvQTW0crxO+roVklTI9LZFvsPhnidKJiqdG1ld7x0ZXkgIdM0T5XxWE1NEpDqOaKlmh3oQVlKRI3fQDDyw+hsNkp5bpA1vJ+E+pj954fRXauqxopCAR6YEGJyd++Wh4jQJRjg7aOt6uuvviY91JWZY1ZJkmBNmzX1zQ4ojMIbj5d1jwrF/ESCqliIMGkV5J3i8YO0UqT1x5+cL8v8eUqxEJEBcGN7ica5SCgTov7lxEmZ3sfE5Qrb3yd16qQ8PykB6NwNGxMAqoqx4JX3Gzdu8AodgxU+D6Sh+JxrHWMsw+FmvO8L5s/33+FJ66KUAopgbr71FneF3m8vu1+dpsJRjm/0AMceuv8B97FyZrlHBYa3/OXXXnP333ufl3bDa3Xbv/4VfJ3ydf6W40tcMCOjWtqEtXLlrKlUkJjsDCWN7Gze3JzTGLL73boQgWIZ7j3VqmVIhzp/9ov7SHaW3+cH0YCwpTqnw8uG3+fX+EKgEy1xEhL+vq68rKkIa3j58PtXXnrJp4+Qex42uhZ2697dXSLd851FWvN7nMPHZ++3ImCEdSsW2b5jxjxeYXmIAcSqgbwAQdEVP4IM4Mmpc0gdl1EtQ8T0tyzFMyyzadPvXj+STlaQoheee95NVBiY1q3IgRCS3VcV7zO/nenz4Cia4T0XNzNW2rweqtxXtsPDiEIaKjAzdMMYM2q0l0witw1SXVoX9GqR5iD8yfZza+RUkt+6ds1aX5SFRAihbjpHoZSQlyIniCceQP6RU3mA0i+owoWggWkd3dxqVEdaqHxmHmJujyP4HQ/LH5f/6Cvtyff9VnJFqDMkFrwFy+fmlQlDOeXV7qtj4hghhWNHj1VXsIN1TJvb87JeyCoi7xQQMJmhaIyQaG0Vk1GUt1D7xTm1bu065RbP8nnLeEZKK92Eavbd5UlT5YrvjrbHHlvJShGdY+RzMYkhjxUPBXgGxnnBOinSonhokiS7wl6RYLnC+nqhCj2IUqDwEDauu1tvvjlpuJnlBn02UG1F/+mvKf4m75JcSJQEgjxY1tFVqTiBMdb3Pni/Gjac4z9igvnIww+7Sy69NJPwBsvm5RXCTFe5sCGhgz4vBXevaFLZ5dprwl/70Pe7PXv4z/6tqv2AZGdZKJs/wI/8ezqzhY18wnTt11+3SrXx2/B9iY5tYcuJ7PywpfVsePlI70PnfrA86T/5tV9LEtQagm3wmt/nRxg71p/qnGaZZJZf48t6wpaYmhGeYLHcLwnnQvi3qd6TxnORihxffPmlba4tSOwVunbDih+p1pef3+f3OOfnvsVpXVnvFnE6snw+FjynPNx5sOMpg0gw2w1Crczaw4VXyTaPRwuZHDRQ8eosmLfAzZMnY+XKFT4MTTixZKmSKsza03tS6CJF+Jxt04JzyOAhvk0nMkmE/5FvwptEVSZtQmn7CoGk9zySW3/+L+9e1uA4IHTovFIo9Z0KviBhhGLIm0XXMuzdCX6TziuElX/MllknuZd4B5kk5IeRpkFSO2R4ikKaEyeM9yH2xBtufmyLSQlhMG7W5KVCnDYqT/mMs8/0ucacK4F3d5jyjQnxU1AAKUGDkslQMY0jMi942ZHu2UxM5nhyWlPf8eDCW0HxFeuLYpCpzUVl8/1Yfjlpooj7BJ83HeX3hWEZohSH6bxEA5McurBxbdANDImbsFWWt5pCiA06x88782z38GOPeE8r195/nn4qc1FUAW771y0++hF8+K87bsskq3xGjihheghrfhrKA4nW+ZTO2lZ3n/P33NNP+1xmJtSJxrgPHLBVf5SHK132IDs5GfJPiYT1hBNPck/9339y+pn/rpV0mAnhQjJIlQryRvkSndLAwikbfAbG2dkmRQ9yY0S/wkYBXZDClR/7RWQjO8vv8yPdczq7/eLz/BhfJkthY4LO5A6PPpPkRMKKek5ujWcuKTundursr9GwUgPrPPKoI/2qE0lzbreXzu/ye5zT2fautGz+xFZ3AcQgCOReDR440L0kiQ76mSMhhcwIIV8KIHgwJM6AgQavHjdIKsGHDx/uW3pCVvFe4lGg0Knvx73dCH0H2YTUsj5yXxs3aezDw0cefZT3GhHCpnqZhwA3BORgIC6sH1K7UTqeFPGEc7jyc3i4EU1VHi6an2927eq7LdHLHmIFIU92/OlsH4JKjhmEPK9klX1hn9g3coc//aS/FxrvqlApzQy2B1nlWOnEtFHjRy4wwv+MDd5x0j0wKrTR/hszepQK7Tb5G22z5s3UurWRPOSbNVU3SK6nmLymEFoaFjBZCM4XFCLQ+OVcIDzN+pOFWTl+zknOJ84XPMoDBgz04/biC8/7XGcmInkdM39Q2+l/TFwSDbKVzAjd33bXHf66WC4vOtGLxDG+9obrfYFV8PtTlBc9RLJWR2wRG4dokScHnl3f6OrwTN539z2uk4ha+9ZtspBVtkdFd9iogKcQhVSQsOWYxrLtIYZ/6t8n01cmMvPCKy9nRjluufFmX8yY+GNIw0cqJMQjS+7mk5nQGgAAFq1JREFUK11fd/1CBDZx+eBv7klEVMJG5AeR/ZwMqaHnXnhRFfKn+MWQsgsbKSqBIdcWNho0YLS0TrQiSTylicsEf4e7BpJOFLY5mmxjudkvfpd4ToajF3wfWG7Pjz8TioYCZwDh99yc08H+JL7mx/iihBO0Fg/WT2oadsABFf21GHzOtRi11Xnwm/DrhZdc4m67805/baLycP+992aZdO0josxYJBaaBfixrgOliBLZQtdlTmOS23GOvB8JC+a0LwmLxu5P87DmYkjRzPxc3s6Rw0e4g6pleH1BZDYqV67ivQc8nLi5ii944kglNxXF6AKiMZgs/Mo6Cfvz/THt20sT8mhP2PCuQTgaN27ivY/c0EdpuXXqsX5Sx44KD5f2ItvfSH4JIsPFGU76z8XhRfoJJGfShIk+n42K3RMUPm2ralG8UuRV4fXL7kYeaQN5WCggqtwg8aZQMU0bU/KM/tiUOy9NOrtD/jAkefGixd6TTgifKuUiwoTP+8vDBxEAJ1IAyEeePn2ayOWfLkMPpRIlNvoxpngOkppoTHKofMZT31RElwI2vN2cCzyoderpvNss08UEA68MKgTk0VI1nsqzlri9nfl38RKbpWrC+0AzhkTDow154zrE0EIGOxQfgnA9n4NvDxUSvvv2Oz6siCeUh+gXylNnHXcpx5h0G4qVukr/MRlRZD0YuriMYdieffEFHyEprmhB2E6Q3BGFWK++/PI2EzEiKqnsO+0jufHkJYeNdIUPJdWF9BHHHKgUhJfhPdt+S7nKgdG6N5Vxnlx/zTWuV++Psyhy3HnP3X4yFu7kxLpQWuhyzbVqlvA3nwoTeEvfeestd/Fll/poFMuhZBIY0ZTAuF4pTMSItiRadseWuBx/X3VNF53z45Tqs9E3FggvgxQWlpv94ndhDzF/F8tm/HJ7fiSGzSlsQjbpkccfc+dpgpTOOc3+ZWf5Mb7c68jBv+mWrfnRFJuiCX5Yw6zpJB9ImisouCumdKWwBUoAfJaIJ+lRGEounP9EQJCVe0/X8KGHHa5i3M26qnj0uffTmCdsPIs6ntxJhdMH+05k4e/C51SigyR8XeY0Jterc1du7gO5wYB9z2lfOD+CAsDwccblvRHWPIwkFzyV2wsU5nir25uuqPKxCO8iaYXEEASEcP8GkctfRDJou4rHKyfDK0f7wBEjhksMv74v5miiAhwKQpjNEvY/VDeC3RXWe1/5aTVq1JT+ajOJbxf3HYtocbcjDQymy4OCp6+3qv1PlZZlexFu8mmjhqrze38JCUE4wAzB7SUijr9K3oWb64405MgI9TM+pE0MFsmcLy84N04IBFXbAz8b4MOlFPvMmfO9f4h+owK835TrlepcWS3JM4grRHRvec7IX0antqg8kD9v/EUTo598i9A/RdJZF/8Ksjc1PDaQ77p16/tWmOHPeY+YOI0tiEIQ5q9V62AvSRVUE7MMyg/YS/L0HSlCT6pJYMjrPPjvh/2fEPp7JCgOLqQLQFaxAyoe4EaOG+s9tExA8dRv0PaYfDKuzz/znFswf75/MAfFWfwOj2Egr8PfgXFNlCheQtf1CN+sIPicV9JAaN38Sd9+OUqp3Sv5nudEiBMnghAa/gXGpOh3pcCEtTyD73ilbe8Lzz4f/ijb90x2zj3jTPf4f/+TKejOpBixdogpovbsD/J9kE/Gg+NnXwfp3Ma4pz2mxgGPPPaY/x68/q33RBGCkDzn5uOPPOpTCAgpk6OdaKedcbp7Q+SIJh6prHbt2u7TgYN8BINrLzAKO1kHlu5+8RuUGtBWDhvRk0aaeCLRF77H5Ob8YEIzUK1qg85WbAedWv6REgYZIQc/6jkd3s9k7/NjfN/s2s1rAlNTgV0uBRxySikKDgxHy4vPP+f/ZN8P3bJs8D1Fpjg+mkvyMJH8cexEyIK21fxNh7/BgwYq/WfzNlnPxx9+6FfHJHO6VCTCDQie/O9/g01leUU3mYJJ0sVoIhK28HWZ05hwnkPQ07kP0IEuNxhQUJvTvsSZrDI2RcuVKXtfeJDsffoIQJC44aIE8LNm9Hg4yd0iVEvTADwHkNWouTUsx7pYD+ErCm54QLdp28bf8CGuG3WzP1SeCuS25uvB2a9PX53In/qq8WQPzPSPKvoveNjT7QuvL/mttDFllsuMFfK+o4gr5JlZdu+PPnZvS2uSgiYKxsBjZxA1SNOsmd+qQ806n29cq2Yt7wGfKUKKFiVpD9yof5KWJPnR41XgQ44wnoR0zhXOLY4R8hXk5+GNw5uBN5ZzM+r6oo/69l0Srx6kBY99onE+kctWT5M3SBK5v0V3L5plMXq5cw5CapkUNGnaxE8awgtx3dx0/Q3+fOVzQuzHSec0bBAwQuqQNLx+5KwjxM9Dl0Ybq1atdPX18CWXmDzpiRMmuOu6XCON4f6upbrP8RvILoV11151tfvvM09vI80D4WsnQovqA7nh2RkpPxRi1teDPRkuPHRfeO45eZ9u8ilL+2gSi+c9THAppLri738XLlulpbLbXvA5EaGP3u8l2bTVflLEZJR1UnQJSaEIFVJIrj2pLtd16eKL1YLf84oXe7LSp5DdAxMmDbSnxNjvf914o+vfr5//u5Nyc2+57Vb/Pvy/wxs29Fq4gSh/+LtL/nZZFq8sqRncMxkXjBQP2tXectPNPmUq+G06+8Vv6HiWOBGAYNN+9OOPPsr0IAbr/0mYpXN+sJ+cBzgg2A7kjefHSE107rz1Vn+OpXNOB/uR02tex5fzu4/SRypWruQn40SUkFnj3s89mdS5K9W9a+P6DX43Xnr1FZ/2Fd4nJttcW6epeQDXXNiq16jhz1ciJqyXiGONmjW81CPayGD23DPPSOKxa+bPaJxBpDIYf75AqpGOc+EGBDyjyIWF3CZKZoWvSxqO5DQm6Y7zM88/mysMuL+kOj8yQYjhm91qV6+xWUk7hgdXGA6JAhsePtzECYVQbEPFN7mNkI0NIiGQES5+pIy4qLkRo7O5YMF8L6WFlBGhbh6YdPP5Rjc8RM/Ds/0diQU3HCSdasvrw82BB3zduvXyVZIqOB6IGPhwEVOYBGHFg7Ro4Q87LfQNwSF0hbg8mrmQHWSnGNd9pIhAcR0NIRYu+MGHnBdpdr5WxIqQdHE9oAhT7bE7Atu/+5sxN+T1yjXlwYK33ix3CNBrvHHTxhqD4n5iwPkSNojHBx9/5BuFhD/P7v0F556XmdNKURSTDa7ZsFWvXsOTWlJ+8ss4v9CUrV6rphdYx0sGkU3mecRTTZrEfvuV87Jx+dFnHMJarVqGoigVvUf6B53HX4sgRk1FouUrRBfPNSSC/c+rDR8zOrOzGOu65uqr3Tgpc5AyA+mD8AXh6Oy2tT32K9hWbs4PCFNNjTMRvGSpQaw71TkdbD+d17yML2khtMsm15nzggYxiddEOvsSXhbdXPSOScXDm4m3G7WTL5T6kZi3GvyOvO2K0iun1iQ/zrNUY5KbcQ72Nd3XVPuS7voKw/JGWHfQKFGZS7hxf0k0ldPDo2y5/fzNBmJDoQ3yQxA9iAqpBX8pnQACSutNbrR4IZhZU9ABaWVZKpshQ3gPITUUWrEdbMPGDW7CuPH+IbaziCuSNTzUGjdu7HM4EczP0E0Hgp5XryuzepL9586d5zVUIR+TFTb6UXl8OyJPNdlpA5GgyYFP0RApYlyCMWG8uMHgcaAqnfGiYp9xoqsLOqt4AyCtXmtWXkTOAR5Um+QhReMXsXomJHj1Vkl9gJSAFQoPbpDnYmeNcTIcCutnTUTsXnzlZT8OjNMSecHpXIfHm2p2r+CwxWPHMdL3Pdwys7Aedxz2OxlhHTxgYBwOzY7BEDAEtiBgOaw76FQ4UDNDQkOEwvCUVlL4ZHPXopLbSH+EdwlihgcRckO4nbyoGcoH4mFK8QJhTIgs8iIQGDyv5FSRJ0laAmG8P3/dsbmbwf5DHBfLe0JYftjQocoVUuMBVcqDAQ9/iCteFohcFPMEUB5Vwt4QeJotUGk/atQoH27i+51pTB5o70nlOB6vr5VHtXbtGu/dqiKPALqsVPtTIENOJfnIFO3V0hjSdhZSlBORh0SRYkI6AZ2z6JCFZxkvRn54znYmdgVh29dcd21mCPE+FV99+P4HWXYLz/f/qdNU+2Pb+8nI+HFfZPne/th5CCTeQ4KJ+87bI9uyIWAI5DcCRljzG9Ek68NrRg9wWoviYczOAsIFMcEDR+4hZJV+2YsWL/KeRG7MeOQIb0BWSPgm9NziiBb6vLEvfkKNADJEdeXu8tRtW2ee3R5sn885LrzEn/b7RDluI7y39cSTTnQtRV45FohaTsQ1IKpgAWEbPWq0cj4/9VXy5HcVFCuye1GftwvphEhXr1EjM0d13JgxXsUBrV6OhyYRCJfPFIGFrtNsolTpUv733tOOt13/sOBhzN80F+AfkxIMZYgP1TWN3unZhQ39gva/lAhw3QXW9ph2vl0wRQ4YkkINhDnFRdgkqXnYJMFDsdP/R2SD+0jYwoVW4c/tvSFgCBReBCwlYAeMXTt5ZC665BKf1J2TBw0iwz88ZuPGjvOVkQuk2bleBVb7i6T8oSp3cr9oX0iVJIUkfXr3kYbrxz4XEv1SSGD7DseqvesY11ufL5IWZUEzUgWomCYf6eTOp3hZMKqys8MGL/PSJUvd50MGu74qLiPnkwIjCH1BswwRGyq+D2/YyFexMg6oFJAacdY5Z/uxoSgNCSIK6qZPm+6K6YG7Qt7wkgo3Z6gZBcUlTEAomMECwprsWMEGyao3XnvVE6xky9hn0RAgAvJ/T/3X0f4Y41okeoHnPCx500dNCVAWyC5vLtrWbKn8QAAN6keefGKbLmCkzgwdMsQ98tDDvmgpP7Zl6zAEDIGdi4B5WHcA/hQe4VnLjpAFu0AB0ffffe9W6mZLYRVhY8gKBVn0Rt9N4XMKCHzoWLmq5LhCiAiP42mka9Y6PWARiIcArVQf+oJopApQhPKLiqOo5kdXFjkT9PvIcw08i3iZZ82c6SusaZE5X8f347IffTESZKIg2hIR6549ejgqmfHOrV79k/d8IrUyetRIVaQe5xsF/CGiSTEC48axIMOFzA/jTacsqvznqiocApzTecN3FBXUr9/ACGseTwgKh07qcJw0kFu5jGoUFR3gi+fw7FO9z7VJSgfqH2YFA4HddP6jCMC/RON+WUT3STNDwBCIBwLmYd0B44ig/kkSDqdiMadQFd4yWnkiDEwon6ryolvCzHh4CJtjEDm6a/WTd7XHu+968ldQCVwUeCGoB4qY1a5TW1JF9ZXiUFfuLee+mfG1T3uYJWJb2ATvE48bIkrO7vkXXeg6qp0uFcmEMjEmKmhCMgGhzSPqABBaqtZTFagtW7rUjZNcD+kWSLaYGQKGgCFgCBgCcUTACOsOGNUSylds1KixwvVH+t7clStX8cLKOXnOststQl1oB6Lr2K93352qApDdPub2c1IFyqtaHkFlSBxFSz/+uHynVf3n9jiy+x2FIBWrVBJhPdl7WdESRa4sHWNiAjZ4YJeo6IrmA2Ml6TP5qynS5jXJq3SwtGUNAUPAEDAECg8CRlh34FjRUrJhk0YKf7dwderU8d4zvGh7K8yPlzEIheONCwyvK/ma66W/SciYFqzDPh/qvWl45cwKHwKMeau2bVzbtsf4Tj/7Su4MPV7SPcKTmMBrTq4u3lekrTZKBgs9VkLTpEnQnpXORmaGgCFgCBgChkCcETDCuhNGF08bMlfNWx7hGqo4p2atmj7HFW+bz2EMEVZIKhJJE8Z/4fNVZ82clVIAeycckm0yFwgQ7j9YExc6IlFkhRQZZDawwJuKjBek9DtJl01WwdZ4tQZdtHCRaa8GQNmrIWAIGAKGQOwRMMIa+yG2AzQEDAFDwBAwBAwBQ6BwI7C5iqdwH4PtvSFgCBgChoAhYAgYAoZAjBEwwhrjwbVDMwQMAUPAEDAEDAFDIA4IGGGNwyjaMRgChoAhYAgYAoaAIRBjBIywxnhw7dAMAUPAEDAEDAFDwBCIAwJGWOMwinYMhoAhYAgYAoaAIWAIxBgBI6wxHlw7NEPAEDAEDAFDwBAwBOKAgBHWOIyiHYMhYAgYAoaAIWAIGAIxRsAIa4wH1w7NEDAEDAFDwBAwBAyBOCBghDUOo2jHYAgYAoaAIWAIGAKGQIwRMMIa48G1QzMEDAFDwBAwBAwBQyAOCBhhjcMo2jEYAoaAIWAIGAKGgCEQYwSMsMZ4cO3QDAFDwBAwBAwBQ8AQiAMCRljjMIp2DIaAIWAIGAKGgCFgCMQYASOsMR5cOzRDwBAwBAwBQ8AQMATigIAR1jiMoh2DIWAIGAKGgCFgCBgCMUbACGuMB9cOzRAwBAwBQ8AQMAQMgTggYIQ1DqNox2AIGAKGgCFgCBgChkCMETDCGuPBtUMzBAwBQ8AQMAQMAUMgDggYYY3DKNoxGAKGgCFgCBgChoAhEGMEjLDGeHDt0AwBQ8AQMAQMAUPAEIgDAkZY4zCKdgyGgCFgCBgChoAhYAjEGAEjrDEeXDs0Q8AQMAQMAUPAEDAE4oCAEdY4jKIdgyFgCBgChoAhYAgYAjFGwAhrjAfXDs0QMAQMAUPAEDAEDIE4IGCENQ6jaMdgCBgChoAhYAgYAoZAjBEwwhrjwbVDMwQMAUPAEDAEDAFDIA4IGGGNwyjaMRgChoAhYAgYAoaAIRBjBIywxnhw7dAMAUPAEDAEDAFDwBCIAwJGWOMwinYMhoAhYAgYAoaAIWAIxBgBI6wxHlw7NEPAEDAEDAFDwBAwBOKAgBHWOIyiHYMhYAgYAoaAIWAIGAIxRsAIa4wH1w7NEDAEDAFDwBAwBAyBOCBghDUOo2jHYAgYAoaAIWAIGAKGQIwRMMIa48G1QzMEDAFDwBAwBAwBQyAOCBhhjcMo2jEYAoaAIWAIGAKGgCEQYwSMsMZ4cO3QDAFDwBAwBAwBQ8AQiAMCRljjMIp2DIaAIWAIGAKGgCFgCMQYASOsMR5cOzRDwBAwBAwBQ8AQMATigIAR1jiMoh2DIWAIGAKGgCFgCBgCMUbACGuMB9cOzRAwBAwBQ8AQMAQMgTggYIQ1DqNox2AIGAKGgCFgCBgChkCMETDCGuPBtUMzBAwBQ8AQMAQMAUMgDggYYY3DKNoxGAKGgCFgCBgChoAhEGMEjLDGeHDt0AwBQ8AQMAQMAUPAEIgDAkZY4zCKdgyGgCFgCBgChoAhYAjEGAEjrDEeXDs0Q8AQMAQMAUPAEDAE4oCAEdY4jKIdgyFgCBgChoAhYAgYAjFGwAhrjAfXDs0QMAQMAUPAEDAEDIE4IGCENQ6jaMdgCBgChoAhYAgYAoZAjBEwwhrjwbVDMwQMAUPAEDAEDAFDIA4IGGGNwyjaMRgChoAhYAgYAoaAIRBjBIywxnhw7dAMAUPAEDAEDAFDwBCIAwJGWOMwinYMhoAhYAgYAoaAIWAIxBgBI6wxHlw7NEPAEDAEDAFDwBAwBOKAgBHWOIyiHYMhYAgYAoaAIWAIGAIxRsAIa4wH1w7NEDAEDAFDwBAwBAyBOCBghDUOo2jHYAgYAoaAIWAIGAKGQIwRMMIa48G1QzMEDAFDwBAwBAwBQyAOCBhhjcMo2jEYAoaAIWAIGAKGgCEQYwSMsMZ4cO3QDAFDwBAwBAwBQ8AQiAMCRljjMIp2DIaAIWAIGAKGgCFgCMQYASOsMR5cOzRDwBAwBAwBQ8AQMATigIAR1jiMoh2DIWAIGAKGgCFgCBgCMUbg/wGjDcTZbud4hAAAAABJRU5ErkJggg=="
HTML_TEMPLATE = HTML_TEMPLATE.replace("__LOGO_DATA_URI__", LOGO_DATA_URI)
LOGIN_TEMPLATE = LOGIN_TEMPLATE.replace("__LOGO_DATA_URI__", LOGO_DATA_URI)
VERIFY_TEMPLATE = VERIFY_TEMPLATE.replace("__LOGO_DATA_URI__", LOGO_DATA_URI)
INVENTORY_TEMPLATE = INVENTORY_TEMPLATE.replace("__LOGO_DATA_URI__", LOGO_DATA_URI)




if __name__ == "__main__":
    print("=" * 50)
    print("  Legal Notice Generator")
    print(f"  Detected CPUs: {_CPU}")
    print(f"  Machine profiles (pick at generate time in the webpage):")
    for _name, _prof in MACHINE_PROFILES.items():
        _marker = " (default)" if _name == DEFAULT_MACHINE else ""
        print(f"    - {_name:4s}{_marker}  render={_prof['render_workers']}  "
              f"[{_prof['label']}]")
    print("  Open browser: http://127.0.0.1:5002")
    print("=" * 50)
    # threaded=True is required so the /status and /download polling hits
    # the same process as the background worker thread.
    app.run(debug=False, port=5002, threaded=True)
